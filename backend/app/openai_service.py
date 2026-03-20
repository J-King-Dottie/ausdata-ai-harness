from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from functools import lru_cache
from html import unescape
import json
import logging
import re
import shutil
import subprocess
import sys
import tempfile
import os
from pathlib import Path
from threading import Event, Lock
from typing import Any, Callable, Dict, List, Optional
from urllib.parse import urlencode

import httpx
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from .config import get_settings
from .harness.parser import HarnessParserError, parse_harness_loop_output
from .harness.prompt_builder import (
    build_sandbox_codegen_messages,
    build_loop_payload,
    build_model_messages,
    load_system_prompt,
)
from .harness.state import build_chat_history_payload, compact_artifacts, compact_chat_history, compact_loop_history
from .macro_data import (
    build_macro_shortlist,
    get_macro_candidate_metadata,
    retrieve_macro_candidate,
    run_macro_query,
)
from .mcp_bridge import (
    MCPBridgeError,
    get_dataflow_metadata,
    list_dataflows,
    resolve_dataset,
)
from .storage import ConversationStore


settings = get_settings()
logger = logging.getLogger("abs.backend.harness")
if not logger.handlers:
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(
        logging.Formatter("[%(asctime)s] %(levelname)s %(name)s - %(message)s")
    )
    logger.addHandler(stream_handler)
logger.setLevel(logging.INFO)
logger.propagate = False

OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"
WEB_SEARCH_URL = "https://html.duckduckgo.com/html/"
WEB_USER_AGENT = "Mozilla/5.0 (compatible; Seshat/1.0; +https://dottieaistudio.com.au/)"
PRIOR_RUN_WINDOW = 5
HARNESS_RESPONSE_SCHEMA: Dict[str, Any] = {
    "type": "json_object",
}
PLAN_APPROVAL_RE = re.compile(r"^\s*(yes|y|ok|okay|proceed|go ahead|do it|use that|add it|sounds good)\b", re.IGNORECASE)
PLAN_REJECT_RE = re.compile(r"^\s*(no|nah|stop|cancel|change|revise|refine)\b", re.IGNORECASE)
CORRECTION_RESET_RE = re.compile(
    r"\b("
    r"wrong|mistake|made a mistake|double[\s-]?count|double[\s-]?counted|"
    r"corrected|redo|re-do|recheck|re-check|rebuild|re-run|rerun|"
    r"that's not right|that is not right|you messed up|you are making mistakes|"
    r"correct totals|with the corrected totals|with correct totals"
    r")\b",
    re.IGNORECASE,
)
WEB_RESULT_LINK_RE = re.compile(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', re.IGNORECASE | re.DOTALL)
WEB_RESULT_SNIPPET_RE = re.compile(r'<(?:a|div)[^>]+class="result__snippet"[^>]*>(.*?)</(?:a|div)>', re.IGNORECASE | re.DOTALL)
STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "by", "for", "from", "how", "i", "in",
    "is", "it", "of", "on", "or", "the", "to", "what", "which", "with", "over",
    "last", "years", "year", "there", "many", "much", "jobs", "job", "data",
}
CLARIFICATION_KEYWORDS = {
    "why", "driver", "drivers", "cause", "causes", "explain", "decline",
    "declining", "falling", "trend", "over", "happened", "happen",
}
COMPLEX_ANALYSIS_KEYWORDS = {
    "compare", "comparison", "ratio", "per", "highest", "lowest", "rank",
    "ranking", "versus", "vs", "relative", "relative_to", "productivity",
}
ABS_QUERY_HINTS = {
    "abs", "australian bureau of statistics", "state final demand", "labour account",
    "supply use", "input output", "asgs", "seasonally adjusted", "trend",
}
MACRO_QUERY_HINTS = {
    "gdp", "inflation", "unemployment", "cpi", "interest rate", "policy rate",
    "current account", "trade balance", "trade", "imports", "exports", "debt", "house prices", "exchange rate",
    "world bank", "imf", "oecd", "comtrade", "un comtrade", "productivity",
}
NON_ABS_COUNTRY_HINTS = {
    "japan", "china", "us", "usa", "united states", "uk", "united kingdom",
    "germany", "france", "canada", "india", "korea", "euro area", "singapore",
    "brazil", "indonesia", "new zealand",
}

EMPLOYMENT_INTENT_TERMS = {
    "employment",
    "employed",
    "employ",
    "job",
    "jobs",
    "worker",
    "workers",
    "workforce",
    "labour",
    "labor",
}

CORRECTION_EXPLANATION_RE = re.compile(
    r"\b(what caused|why did you make|why did that happen|explain what went wrong)\b",
    re.IGNORECASE,
)
MAX_CONSECUTIVE_RECOVERY_FAILURES = 3
GPT_5_4_INPUT_PRICE_PER_MILLION = 2.50
GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION = 0.25
GPT_5_4_OUTPUT_PRICE_PER_MILLION = 15.00
AI_COST_SURCHARGE_RATE = 0.10


class ConversationCancelled(RuntimeError):
    """Raised when a conversation is cancelled mid-generation."""


_CANCELLATION_LOCK = Lock()
_CANCELLATION_EVENTS: Dict[str, Event] = {}


def _acquire_cancellation_event(conversation_id: str) -> Event:
    with _CANCELLATION_LOCK:
        event = _CANCELLATION_EVENTS.get(conversation_id)
        if event is None:
            event = Event()
            _CANCELLATION_EVENTS[conversation_id] = event
        event.clear()
        return event


def cancel_conversation_processing(conversation_id: str) -> None:
    with _CANCELLATION_LOCK:
        event = _CANCELLATION_EVENTS.get(conversation_id)
        if event is None:
            event = Event()
            _CANCELLATION_EVENTS[conversation_id] = event
        event.set()


def _release_cancellation_event(conversation_id: str) -> None:
    with _CANCELLATION_LOCK:
        _CANCELLATION_EVENTS.pop(conversation_id, None)


def _ensure_not_cancelled(conversation_id: str, event: Event, stage: str) -> None:
    if event.is_set():
        logger.info("Conversation cancelled cid=%s stage=%s", conversation_id, stage)
        raise ConversationCancelled(f"Conversation {conversation_id} cancelled during {stage}")


def _conversation_runtime_dir(conversation_id: str) -> Path:
    safe_id = "".join(ch for ch in conversation_id if ch.isalnum() or ch in {"-", "_"})
    if not safe_id:
        safe_id = "conversation"
    return settings.runtime_dir / "conversations" / safe_id


def _ensure_runtime_dirs(conversation_id: str) -> Path:
    run_dir = _conversation_runtime_dir(conversation_id)
    (run_dir / "artifacts").mkdir(parents=True, exist_ok=True)
    (run_dir / "sandbox").mkdir(parents=True, exist_ok=True)
    return run_dir


def _clear_runtime_dir(conversation_id: str) -> None:
    run_dir = _conversation_runtime_dir(conversation_id)
    if run_dir.exists():
        shutil.rmtree(run_dir, ignore_errors=True)


def clear_dataset_cache() -> None:
    if settings.runtime_dir.exists():
        shutil.rmtree(settings.runtime_dir, ignore_errors=True)


def reset_conversation_runtime(conversation_id: str) -> None:
    _clear_runtime_dir(conversation_id)


def reset_all_conversation_memory() -> None:
    clear_dataset_cache()


def _extract_openai_output_text(response_data: Dict[str, Any]) -> str:
    output_text = str(response_data.get("output_text") or "").strip()
    if output_text:
        return output_text

    output = response_data.get("output")
    if not isinstance(output, list):
        return ""

    fragments: List[str] = []
    for item in output:
        if not isinstance(item, dict):
            continue
        content = item.get("content")
        if not isinstance(content, list):
            continue
        for block in content:
            if not isinstance(block, dict):
                continue
            block_type = str(block.get("type") or "").strip().lower()
            text_value = block.get("text")
            if block_type in {"output_text", "text"} and text_value:
                fragments.append(str(text_value))

    return "".join(fragments).strip()


def _summarize_openai_response(response_data: Dict[str, Any]) -> str:
    if not isinstance(response_data, dict):
        return _truncate(str(response_data), 600)

    def _present(value: Any) -> bool:
        return value is not None and value != ""

    summary: Dict[str, Any] = {}
    for key in ("id", "status", "model", "output_text"):
        value = response_data.get(key)
        if _present(value):
            summary[key] = _truncate(str(value), 240)

    for key in ("incomplete_details", "error", "usage"):
        value = response_data.get(key)
        if _present(value):
            try:
                summary[key] = json.loads(json.dumps(value, ensure_ascii=True))
            except Exception:
                summary[key] = _truncate(str(value), 240)

    output = response_data.get("output")
    if isinstance(output, list):
        output_summary = []
        for item in output[:4]:
            if not isinstance(item, dict):
                output_summary.append(_truncate(str(item), 120))
                continue
            entry: Dict[str, Any] = {}
            item_type = item.get("type")
            if _present(item_type):
                entry["type"] = item_type
            item_status = item.get("status")
            if _present(item_status):
                entry["status"] = item_status
            for key in ("id", "role"):
                value = item.get(key)
                if _present(value):
                    entry[key] = value
            content = item.get("content")
            if isinstance(content, list):
                content_summary = []
                for block in content[:4]:
                    if not isinstance(block, dict):
                        content_summary.append(_truncate(str(block), 120))
                        continue
                    block_entry: Dict[str, Any] = {}
                    block_type = block.get("type")
                    if _present(block_type):
                        block_entry["type"] = block_type
                    block_status = block.get("status")
                    if _present(block_status):
                        block_entry["status"] = block_status
                    text_value = block.get("text")
                    if _present(text_value):
                        block_entry["text_preview"] = _truncate(str(text_value), 160)
                    refusal = block.get("refusal")
                    if _present(refusal):
                        block_entry["refusal_preview"] = _truncate(str(refusal), 160)
                    for key in ("id",):
                        value = block.get(key)
                        if _present(value):
                            block_entry[key] = value
                    if block_entry:
                        content_summary.append(block_entry)
                entry["content"] = content_summary
            if entry:
                output_summary.append(entry)
        summary["output"] = output_summary

    return _truncate(json.dumps(summary, ensure_ascii=True), 1200)


def _safe_int(value: Any) -> int:
    try:
        numeric = int(value)
    except Exception:
        return 0
    return numeric if numeric > 0 else 0


def _extract_usage_totals(response_data: Dict[str, Any]) -> Dict[str, int]:
    usage = response_data.get("usage") if isinstance(response_data.get("usage"), dict) else {}
    input_tokens = _safe_int(usage.get("input_tokens"))
    output_tokens = _safe_int(usage.get("output_tokens"))

    if not input_tokens:
        input_tokens = _safe_int(usage.get("prompt_tokens"))
    if not output_tokens:
        output_tokens = _safe_int(usage.get("completion_tokens"))

    details = usage.get("input_tokens_details") if isinstance(usage.get("input_tokens_details"), dict) else {}
    cached_input_tokens = _safe_int(details.get("cached_tokens"))

    return {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "cached_input_tokens": cached_input_tokens,
    }


def _compute_run_cost_breakdown(*, input_tokens: int, output_tokens: int, cached_input_tokens: int = 0) -> Dict[str, float]:
    cached_tokens = min(max(cached_input_tokens, 0), max(input_tokens, 0))
    uncached_input_tokens = max(input_tokens, 0) - cached_tokens
    ai_cost = (
        (uncached_input_tokens / 1_000_000) * GPT_5_4_INPUT_PRICE_PER_MILLION
        + (cached_tokens / 1_000_000) * GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION
        + (max(output_tokens, 0) / 1_000_000) * GPT_5_4_OUTPUT_PRICE_PER_MILLION
    )
    surcharge = ai_cost * AI_COST_SURCHARGE_RATE
    final_cost = ai_cost + surcharge
    return {
        "ai_cost_usd": round(ai_cost, 6),
        "surcharge_usd": round(surcharge, 6),
        "final_cost_usd": round(final_cost, 6),
    }


def _build_run_cost_payload(
    *,
    input_tokens: int,
    output_tokens: int,
    cached_input_tokens: int = 0,
    model: str | None = None,
) -> Dict[str, Any]:
    cached_tokens = min(max(cached_input_tokens, 0), max(input_tokens, 0))
    breakdown = _compute_run_cost_breakdown(
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        cached_input_tokens=cached_tokens,
    )
    return {
        "model": str(model or settings.openai_model or "gpt-5.4"),
        "input_tokens": max(input_tokens, 0),
        "cached_input_tokens": cached_tokens,
        "output_tokens": max(output_tokens, 0),
        "pricing": {
            "input_per_million_usd": GPT_5_4_INPUT_PRICE_PER_MILLION,
            "cached_input_per_million_usd": GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION,
            "output_per_million_usd": GPT_5_4_OUTPUT_PRICE_PER_MILLION,
            "surcharge_rate": AI_COST_SURCHARGE_RATE,
        },
        **breakdown,
    }


def _call_model(
    messages: List[Dict[str, str]],
    *,
    reasoning_effort: Optional[str] = None,
    usage_callback: Optional[Callable[[Dict[str, int]], None]] = None,
) -> str:
    openai_messages: List[Dict[str, Any]] = []
    for message in messages:
        role = str(message.get("role") or "user").strip().lower()
        content = str(message.get("content") or "").strip()
        if not content:
            continue
        if role not in {"system", "user", "assistant", "developer"}:
            role = "user"
        content_type = "output_text" if role == "assistant" else "input_text"
        openai_messages.append(
            {
                "role": role,
                "content": [
                    {
                        "type": content_type,
                        "text": content,
                    }
                ],
            }
        )

    if not openai_messages:
        raise RuntimeError("No model input was generated for the harness loop.")

    payload: Dict[str, Any] = {
        "model": settings.openai_model,
        "input": openai_messages,
        "reasoning": {
            "effort": reasoning_effort or settings.openai_reasoning_effort,
        },
        "text": {
            "format": HARNESS_RESPONSE_SCHEMA,
        },
    }

    response = httpx.post(
        OPENAI_RESPONSES_URL,
        headers={
            "Authorization": f"Bearer {settings.openai_api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=settings.openai_timeout_seconds,
    )
    if response.status_code >= 400:
        if response.status_code == 429:
            raise RuntimeError(
                "OpenAI rate limit reached for this request. "
                "Please retry in a moment or ask a narrower follow-up."
            )
        raise RuntimeError(
            f"OpenAI responses error {response.status_code}: {response.text.strip()}"
        )

    response_data = response.json()
    if usage_callback:
        usage_callback(_extract_usage_totals(response_data))
    text = _extract_openai_output_text(response_data)
    if not text:
        raise RuntimeError(
            "OpenAI returned an empty response. "
            f"Response summary: {_summarize_openai_response(response_data)}"
        )
    return text


def _call_model_text(
    messages: List[Dict[str, str]],
    *,
    reasoning_effort: Optional[str] = None,
    usage_callback: Optional[Callable[[Dict[str, int]], None]] = None,
) -> str:
    openai_messages: List[Dict[str, Any]] = []
    for message in messages:
        role = str(message.get("role") or "user").strip().lower()
        content = str(message.get("content") or "").strip()
        if not content:
            continue
        if role not in {"system", "user", "assistant", "developer"}:
            role = "user"
        content_type = "output_text" if role == "assistant" else "input_text"
        openai_messages.append(
            {
                "role": role,
                "content": [
                    {
                        "type": content_type,
                        "text": content,
                    }
                ],
            }
        )

    if not openai_messages:
        raise RuntimeError("No model input was generated for the sandbox code step.")

    payload: Dict[str, Any] = {
        "model": settings.openai_model,
        "input": openai_messages,
        "reasoning": {
            "effort": reasoning_effort or "low",
        },
    }

    response = httpx.post(
        OPENAI_RESPONSES_URL,
        headers={
            "Authorization": f"Bearer {settings.openai_api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=settings.openai_timeout_seconds,
    )
    if response.status_code >= 400:
        if response.status_code == 429:
            raise RuntimeError(
                "OpenAI rate limit reached for this sandbox code step. "
                "Please retry in a moment or ask a narrower follow-up."
            )
        raise RuntimeError(
            f"OpenAI responses error {response.status_code}: {response.text.strip()}"
        )

    response_data = response.json()
    if usage_callback:
        usage_callback(_extract_usage_totals(response_data))
    text = _extract_openai_output_text(response_data)
    if not text:
        raise RuntimeError(
            "OpenAI returned an empty response for the sandbox code step. "
            f"Response summary: {_summarize_openai_response(response_data)}"
        )
    return text


def _compact_user_only_history(messages: List[Dict[str, str]], limit: int = 6) -> List[Dict[str, str]]:
    compact: List[Dict[str, str]] = []
    for item in messages[-limit * 2:]:
        if not isinstance(item, dict):
            continue
        role = str(item.get("role") or "").strip().lower()
        content = str(item.get("content") or "").strip()
        if role != "user" or not content:
            continue
        compact.append({"role": "user", "content": content[:600]})
    return compact[-limit:]


def _retain_non_scratchpad_artifacts(artifacts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    kept: List[Dict[str, Any]] = []
    for item in artifacts:
        if not isinstance(item, dict):
            continue
        kind = str(item.get("kind") or "").strip()
        if kind in {"sandbox_result", "sandbox_output"}:
            continue
        kept.append(item)
    return kept


def _retain_non_scratchpad_loop_history(loop_history: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    kept: List[Dict[str, Any]] = []
    for item in loop_history:
        if not isinstance(item, dict):
            continue
        step = item.get("step") if isinstance(item.get("step"), dict) else {}
        step_id = str(step.get("id") or "").strip()
        result_data = item.get("result_data") if isinstance(item.get("result_data"), dict) else {}
        kind = str(result_data.get("kind") or "").strip()
        if step_id == "sandbox_tool":
            continue
        if step_id == "compose_final":
            continue
        kept.append(item)
    return kept


def _should_reset_after_user_correction(state, user_message: str) -> bool:
    text = str(user_message or "").strip()
    if not text:
        return False
    if CORRECTION_EXPLANATION_RE.search(text):
        return False
    if not CORRECTION_RESET_RE.search(text):
        return False
    if not state.loop_history and not state.artifacts:
        return False
    return any(
        isinstance(item, dict) and str(item.get("role") or "").strip().lower() == "assistant"
        for item in (state.messages[-6:] if isinstance(state.messages, list) else [])
    )


def _build_correction_reset_message(state, user_message: str) -> str:
    recent_user_messages = [
        str(item.get("content") or "").strip()
        for item in (state.messages if isinstance(state.messages, list) else [])
        if isinstance(item, dict) and str(item.get("role") or "").strip().lower() == "user"
    ]
    current = str(user_message or "").strip()
    prior_user_messages = [message for message in recent_user_messages[:-1] if message]
    prior_question = prior_user_messages[-1] if prior_user_messages else ""

    parts = ["A user correction invalidated earlier derived results."]
    if prior_question:
        parts.append(f"Most recent user context before the correction: {prior_question}")
    parts.append(f"Current correction or revised request: {current}")
    parts.append(
        "Treat prior derived artifacts, rankings, charts, and intermediate calculations as stale. "
        "Do not reuse them. Re-retrieve or re-derive from source data and validated totals only."
    )
    return "\n\n".join(parts).strip()


def _reset_context_after_user_correction(state, user_message: str) -> str:
    correction_message = _build_correction_reset_message(state, user_message)
    handoff_entry = {
        "step": {
            "id": "compose_final",
            "summary": "User correction detected; previous analytical scratchpad cleared",
        },
        "progress_note": "Rechecking from source after the correction.",
        "result_summary": (
            "The user called out a likely analytical error. Current-run artifacts and loop history "
            "were cleared before recomputation."
        ),
        "result_data": {
            "kind": "correction_handoff",
            "user_correction": str(user_message or "").strip(),
        },
    }
    preserved_loop_history = _retain_non_scratchpad_loop_history(
        state.loop_history[:-1] if state.loop_history else []
    )
    preserved_artifacts = _retain_non_scratchpad_artifacts(state.artifacts)
    state.loop_history = preserved_loop_history + [handoff_entry]
    state.artifacts = preserved_artifacts
    return correction_message


def _compose_best_effort_final(
    conversation_id: str,
    user_message: str,
    state,
    *,
    usage_callback: Optional[Callable[[Dict[str, int]], None]] = None,
) -> str:
    completed_runs = _recent_completed_runs(state, keep_runs=PRIOR_RUN_WINDOW + 1)
    recent_messages = _slice_from_recent_runs(state.messages, completed_runs, "message_count")
    recent_loops = _slice_from_recent_runs(state.loop_history, completed_runs, "loop_count")
    recent_artifacts = _slice_from_recent_runs(state.artifacts, completed_runs, "artifact_count")
    payload = {
        "task": {
            "user_message": user_message,
            "reason": "Loop limit reached. Compose the best possible final answer from the evidence already gathered.",
        },
        "chat_history": build_chat_history_payload(recent_messages, recent_full_limit=6, older_compact_limit=3),
        "loop_history": compact_loop_history(recent_loops, limit=4),
        "available_artifacts": compact_artifacts(recent_artifacts, limit=4),
        "plan_state": _build_plan_state(state, user_message=user_content),
    }
    messages = [
        {"role": "system", "content": load_system_prompt()},
        {
            "role": "user",
            "content": (
                "You must stop using tools and write the best possible final answer now.\n"
                "Use only the evidence already gathered in this conversation.\n"
                "Do not mention loop limits or internal harness mechanics.\n"
                "If the evidence is incomplete, answer to the best of your ability, "
                "state what the evidence does show, and end with one short caveat only if needed.\n"
                "Do not ask the user a clarification question in this fallback path.\n\n"
                f"Current state:\n{json.dumps(payload, ensure_ascii=True)}"
            ),
        },
    ]
    try:
        return _call_model(messages, usage_callback=usage_callback)
    except Exception as exc:
        logger.exception(
            "Best-effort final compose failed cid=%s error=%s",
            conversation_id,
            exc,
        )
        return (
            "I couldn't fully resolve every part of that, but here is the best answer from the evidence gathered so far.\n\n"
            "I found relevant ABS data and narrowed the likely series, but the result was not fully resolved into a clean final analysis. "
            "If you want, narrow the request to one measure such as jobs, filled jobs, output, or a specific period."
        )


def _next_artifact_id(artifacts: List[Dict[str, Any]]) -> str:
    return f"artifact-{len(artifacts) + 1:03d}"


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _make_artifact_record(
    *,
    state,
    path: Path,
    kind: str,
    label: str,
    summary: str,
    source_references: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    artifact_id = _next_artifact_id(state.artifacts)
    record = {
        "artifact_id": artifact_id,
        "kind": kind,
        "label": label,
        "summary": summary,
        "path": str(path),
    }
    if isinstance(source_references, list) and source_references:
        record["source_references"] = source_references[:12]
    state.artifacts.append(record)
    return record


def get_latest_export_artifact_path(state) -> Path | None:
    target_id = str(getattr(state, "latest_export_artifact_id", "") or "").strip()
    if not target_id:
        return None
    for item in reversed(state.artifacts):
        if not isinstance(item, dict):
            continue
        if str(item.get("artifact_id") or "").strip() != target_id:
            continue
        path_value = str(item.get("path") or "").strip()
        if not path_value:
            return None
        path = Path(path_value)
        return path if path.exists() else None
    return None


def generate_latest_export(conversation_id: str, store: ConversationStore) -> Path | None:
    state = store.load(conversation_id)
    request = state.latest_export_request if isinstance(state.latest_export_request, dict) else None
    if not request:
        return get_latest_export_artifact_path(state)

    try:
        _build_answer_export(
            state=state,
            conversation_id=conversation_id,
            user_message=str(request.get("user_message") or "").strip(),
            final_answer=str(request.get("final_answer") or "").strip(),
            run_loop_start_index=int(request.get("run_loop_start_index") or 0),
        )
        state.latest_export_status = "ready"
        state.latest_export_request = None
        store.save(state)
    except Exception:
        state.latest_export_status = "failed"
        state.latest_export_request = None
        store.save(state)
        raise

    return get_latest_export_artifact_path(state)


def _safe_sheet_text(value: Any) -> str:
    text = str(value or "").strip()
    return text if len(text) <= 32000 else text[:31997] + "..."


def _parse_chart_spec_from_markdown(markdown: str) -> Dict[str, Any] | None:
    text = str(markdown or "").strip()
    if not text:
        return None
    match = re.search(r"```chart\s*(\{.*?\})\s*```", text, flags=re.DOTALL)
    if not match:
        return None
    try:
        parsed = json.loads(match.group(1))
    except Exception:
        return None
    if not isinstance(parsed, dict):
        return None
    series = parsed.get("series")
    if not isinstance(series, list) or not series:
        return None
    normalized_series: List[Dict[str, Any]] = []
    for entry in series:
        if not isinstance(entry, dict):
            continue
        points = entry.get("points")
        if not isinstance(points, list) or not points:
            continue
        normalized_points = []
        for point in points:
            if not isinstance(point, dict):
                continue
            x = str(point.get("x") or "").strip()
            y = point.get("y")
            if not x:
                continue
            try:
                numeric_y = float(y)
            except Exception:
                continue
            normalized_points.append({"x": x, "y": numeric_y})
        if normalized_points:
            normalized_series.append(
                {
                    "name": str(entry.get("name") or "Series").strip() or "Series",
                    "points": normalized_points,
                }
            )
    if not normalized_series:
        return None
    return {
        "type": "bar" if str(parsed.get("type") or "").strip().lower() == "bar" else "line",
        "title": str(parsed.get("title") or "").strip(),
        "xLabel": str(parsed.get("xLabel") or "").strip(),
        "yLabel": str(parsed.get("yLabel") or "").strip(),
        "series": normalized_series,
    }


def _source_preview(source_references: Any) -> str:
    return "; ".join(_source_reference_lines(source_references)[:4])


def _source_reference_lines(source_references: Any) -> List[str]:
    if not isinstance(source_references, list):
        return []
    lines: List[str] = []
    seen: set[str] = set()
    for item in source_references:
        if not isinstance(item, dict):
            continue
        line = " | ".join(
            part
            for part in [
                str(item.get("provider") or "").strip(),
                str(item.get("dataset_id") or item.get("series_id") or "").strip(),
                str(item.get("title") or item.get("indicator") or "").strip(),
                str(item.get("url") or "").strip(),
            ]
            if part
        ).strip()
        if not line or line in seen:
            continue
        seen.add(line)
        lines.append(line)
    return lines


def _load_artifact_payload(record: Dict[str, Any]) -> Any:
    path_value = str(record.get("path") or "").strip()
    if not path_value:
        return None
    path = Path(path_value)
    if not path.exists():
        return None
    try:
        if path.suffix.lower() == ".json":
            return json.loads(path.read_text(encoding="utf-8"))
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None


def _rows_from_artifact(record: Dict[str, Any]) -> List[Dict[str, str]]:
    payload = _load_artifact_payload(record)
    kind = str(record.get("kind") or "").strip()
    artifact_id = str(record.get("artifact_id") or "").strip()
    label = str(record.get("label") or "").strip()
    source_text = _source_preview(record.get("source_references"))
    rows: List[Dict[str, str]] = []

    if isinstance(payload, dict):
        if kind == "macro_query_response":
            series = payload.get("response", {}).get("series") if isinstance(payload.get("response"), dict) else []
            if isinstance(series, list):
                for index, item in enumerate(series[:200], start=1):
                    rows.append(
                        {
                            "section": "raw_data",
                            "artifact_id": artifact_id,
                            "artifact_label": label,
                            "source": source_text,
                            "detail": f"Series {index}",
                            "value": _safe_sheet_text(json.dumps(item, ensure_ascii=False)),
                            "method": "",
                        }
                    )
        elif kind == "abs_resolved_dataset":
            series = payload.get("resolved_dataset", {}).get("series") if isinstance(payload.get("resolved_dataset"), dict) else []
            if isinstance(series, list):
                for index, item in enumerate(series[:200], start=1):
                    rows.append(
                        {
                            "section": "raw_data",
                            "artifact_id": artifact_id,
                            "artifact_label": label,
                            "source": source_text,
                            "detail": f"Series {index}",
                            "value": _safe_sheet_text(json.dumps(item, ensure_ascii=False)),
                            "method": "",
                        }
                    )
        elif kind == "sandbox_result":
            rows.append(
                {
                    "section": "calculation_output",
                    "artifact_id": artifact_id,
                    "artifact_label": label,
                    "source": source_text,
                    "detail": "Derived result",
                    "value": _safe_sheet_text(json.dumps(payload, ensure_ascii=False)),
                    "method": "",
                }
            )

    if rows:
        return rows

    return [
        {
            "section": "raw_data",
            "artifact_id": artifact_id,
            "artifact_label": label,
            "source": source_text,
            "detail": str(record.get("summary") or "").strip(),
            "value": _safe_sheet_text(
                json.dumps(payload, ensure_ascii=False) if isinstance(payload, (dict, list)) else str(payload or "")
            ),
            "method": "",
        }
    ]


def _chart_table(chart_spec: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    series = chart_spec.get("series") if isinstance(chart_spec.get("series"), list) else []
    if not series:
        return [], []
    x_values: List[str] = []
    for entry in series:
        if not isinstance(entry, dict):
            continue
        for point in entry.get("points") or []:
            if not isinstance(point, dict):
                continue
            x = str(point.get("x") or "").strip()
            if x and x not in x_values:
                x_values.append(x)
    headers = [str(chart_spec.get("xLabel") or "Metric").strip() or "Metric"]
    headers.extend(str(entry.get("name") or "Series").strip() or "Series" for entry in series if isinstance(entry, dict))
    rows: List[List[Any]] = []
    for x in x_values:
        row: List[Any] = [x]
        for entry in series:
            point_map = {
                str(point.get("x") or "").strip(): point.get("y")
                for point in (entry.get("points") or [])
                if isinstance(point, dict)
            }
            row.append(point_map.get(x))
        rows.append(row)
    return headers, rows


def _summary_calc_table(calc_rows: List[Dict[str, str]]) -> tuple[List[str], List[List[str]]]:
    headers = ["Calculation", "Description"]
    rows: List[List[str]] = []
    for index, row in enumerate(calc_rows, start=1):
        metric = str(row.get("detail") or f"Calculation {index}").strip() or f"Calculation {index}"
        description = str(row.get("method") or "").strip() or str(row.get("value") or "").strip()
        rows.append([metric, description])
    return headers, rows


def _safe_sheet_name(value: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\\[\\]]+", " ", str(value or "").strip())[:31].strip() or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f" {suffix}"
        candidate = (cleaned[: 31 - len(tail)] + tail).strip()
        suffix += 1
    used.add(candidate)
    return candidate


def _safe_export_filename(user_message: str) -> str:
    words = re.findall(r"[A-Za-z0-9]+", str(user_message or "").lower())
    meaningful = [word for word in words if len(word) > 2][:6]
    stem = "-".join(meaningful) or "analysis-export"
    stem = stem[:64].strip("-") or "analysis-export"
    return f"{stem}.xlsx"


def _write_table(sheet, headers: List[Any], rows: List[List[Any]]) -> None:
    if headers:
        sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _flatten_resolved_dataset_for_export(dataset: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    if not isinstance(dataset, dict):
        return [], []

    series_items = dataset.get("series")
    if not isinstance(series_items, list):
        return [], []

    dimension_keys: List[str] = []
    attribute_keys: List[str] = []
    for series in series_items:
        if not isinstance(series, dict):
            continue
        for key in (series.get("dimensions") or {}).keys():
            if key not in dimension_keys:
                dimension_keys.append(str(key))
        for observation in series.get("observations") or []:
            if not isinstance(observation, dict):
                continue
            for key in (observation.get("dimensions") or {}).keys():
                if key not in dimension_keys:
                    dimension_keys.append(str(key))
            for key in (observation.get("attributes") or {}).keys():
                if key not in attribute_keys:
                    attribute_keys.append(str(key))
        for key in (series.get("attributes") or {}).keys():
            if key not in attribute_keys:
                attribute_keys.append(str(key))

    headers = ["seriesKey"] + dimension_keys + ["observationKey", "value"] + attribute_keys
    rows: List[List[Any]] = []

    def _label_or_value(value: Any) -> Any:
        if isinstance(value, dict):
            if value.get("label") is not None:
                return value.get("label")
            if value.get("code") is not None:
                return value.get("code")
        return value

    for series in series_items:
        if not isinstance(series, dict):
            continue
        series_dims = series.get("dimensions") if isinstance(series.get("dimensions"), dict) else {}
        series_attrs = series.get("attributes") if isinstance(series.get("attributes"), dict) else {}
        observations = series.get("observations") if isinstance(series.get("observations"), list) else []
        for observation in observations:
            if not isinstance(observation, dict):
                continue
            obs_dims = observation.get("dimensions") if isinstance(observation.get("dimensions"), dict) else {}
            obs_attrs = observation.get("attributes") if isinstance(observation.get("attributes"), dict) else {}
            row: List[Any] = [series.get("seriesKey")]
            for key in dimension_keys:
                value = obs_dims.get(key)
                if value is None:
                    value = series_dims.get(key)
                row.append(_label_or_value(value))
            row.append(observation.get("observationKey"))
            row.append(observation.get("value"))
            for key in attribute_keys:
                value = obs_attrs.get(key)
                if value is None:
                    value = series_attrs.get(key)
                row.append(_label_or_value(value))
            rows.append(row)

    return headers, rows


def _apply_export_theme(workbook) -> None:
    summary = workbook["Summary"] if "Summary" in workbook.sheetnames else None
    if summary is None:
        return

    title_font = Font(color="234233", size=16)
    section_font = Font(color="8F6A3A", size=12)
    header_font = Font(color="54745F", size=11)

    for cell in summary[1]:
        cell.font = title_font

    for row in range(1, summary.max_row + 1):
        first_value = str(summary.cell(row=row, column=1).value or "").strip()
        if first_value in {"Presented data", "Calculations"}:
            summary.cell(row=row, column=1).font = section_font
            continue
        values = [str(summary.cell(row=row, column=col).value or "").strip() for col in range(1, summary.max_column + 1)]
        non_empty = [value for value in values if value]
        if len(non_empty) >= 2 and row > 1:
            for col in range(1, len(values) + 1):
                summary.cell(row=row, column=col).font = header_font
            break


def _artifact_call_structure(record: Dict[str, Any], payload: Any) -> str:
    kind = str(record.get("kind") or "").strip()
    if not isinstance(payload, dict):
        return str(record.get("summary") or "").strip()

    if kind == "macro_query_response":
        api_request_url = str(payload.get("response", {}).get("api_request_url") or payload.get("api_request_url") or "").strip() if isinstance(payload.get("response"), dict) or isinstance(payload.get("api_request_url"), str) else ""
        if api_request_url:
            return api_request_url
        source_urls = [
            str(item.get("url") or "").strip()
            for item in (payload.get("source_references") or [])
            if isinstance(item, dict) and str(item.get("url") or "").strip()
        ]
        return source_urls[0] if source_urls else str(record.get("summary") or "").strip()

    if kind == "abs_resolved_dataset":
        retrieval = payload.get("retrieval") if isinstance(payload.get("retrieval"), dict) else {}
        dataset_id = str(retrieval.get("datasetId") or "").strip()
        data_key = str(retrieval.get("dataKey") or "").strip()
        params = []
        for key in ["startPeriod", "endPeriod", "detail", "dimensionAtObservation"]:
            value = str(retrieval.get(key) or "").strip()
            if value:
                params.append(f"{key}={value}")
        base = f"https://data.api.abs.gov.au/rest/data/{dataset_id}/{data_key}" if dataset_id and data_key else ""
        if base and params:
            return base + "?" + "&".join(params)
        if retrieval:
            return _safe_sheet_text(json.dumps(retrieval, ensure_ascii=False))

    return str(record.get("summary") or "").strip()


def _write_raw_artifact_sheet(sheet, record: Dict[str, Any]) -> None:
    payload = _load_artifact_payload(record)
    source_lines = _source_reference_lines(record.get("source_references"))
    source_line = source_lines[0] if source_lines else str(record.get("label") or "").strip()
    call_line = _artifact_call_structure(record, payload)

    sheet["A1"] = f"Source: {source_line}".strip()
    sheet["A2"] = f"Call structure: {call_line}".strip()
    sheet["A3"] = ""
    sheet["A4"] = "Returned data"
    tabular_headers: List[str] = []
    tabular_rows: List[List[Any]] = []
    if isinstance(payload, dict):
        if str(record.get("kind") or "").strip() == "abs_resolved_dataset":
            resolved_dataset = payload.get("resolved_dataset") if isinstance(payload.get("resolved_dataset"), dict) else {}
            tabular_headers, tabular_rows = _flatten_resolved_dataset_for_export(resolved_dataset)
        elif str(record.get("kind") or "").strip() == "macro_query_response":
            response = payload.get("response") if isinstance(payload.get("response"), dict) else {}
            tabular_headers, tabular_rows = _flatten_resolved_dataset_for_export(response)

    row_index = 5
    if tabular_headers and tabular_rows:
        _write_table(sheet, tabular_headers, tabular_rows)
        for column in range(1, min(len(tabular_headers), 10) + 1):
            sheet.column_dimensions[chr(64 + column)].width = 24
    else:
        if isinstance(payload, (dict, list)):
            sheet.cell(row=row_index, column=1, value=_safe_sheet_text(json.dumps(payload, ensure_ascii=False, indent=2)))
        else:
            sheet.cell(row=row_index, column=1, value=_safe_sheet_text(str(payload or "")))
        sheet.column_dimensions["A"].width = 140
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")


def _run_has_exportable_support(state, run_loop_start_index: int) -> bool:
    current_run_history = state.loop_history[run_loop_start_index:]
    for entry in current_run_history:
        if not isinstance(entry, dict):
            continue
        step = entry.get("step") if isinstance(entry.get("step"), dict) else {}
        result_data = entry.get("result_data") if isinstance(entry.get("result_data"), dict) else {}
        step_id = str(step.get("id") or "").strip()
        kind = str(result_data.get("kind") or "").strip()
        if kind in {"retrieve", "raw_retrieve", "macro_query"}:
            return True
        if step_id == "sandbox_tool":
            tool_input = step.get("tool_input") if isinstance(step.get("tool_input"), dict) else {}
            artifact_ids = tool_input.get("artifact_ids") if isinstance(tool_input.get("artifact_ids"), list) else []
            created_ids = result_data.get("created_artifact_ids") if isinstance(result_data.get("created_artifact_ids"), list) else []
            if artifact_ids or created_ids:
                return True
    return False


def _build_answer_export(
    *,
    state,
    conversation_id: str,
    user_message: str,
    final_answer: str,
    run_loop_start_index: int,
) -> str:
    current_run_history = state.loop_history[run_loop_start_index:]
    chart_spec = _parse_chart_spec_from_markdown(final_answer)
    artifact_map = {
        str(item.get("artifact_id") or "").strip(): item
        for item in state.artifacts
        if isinstance(item, dict) and str(item.get("artifact_id") or "").strip()
    }

    raw_artifact_ids: List[str] = []
    calc_artifact_ids: List[str] = []
    calc_rows: List[Dict[str, str]] = []

    for entry in current_run_history:
        if not isinstance(entry, dict):
            continue
        step = entry.get("step") if isinstance(entry.get("step"), dict) else {}
        result_data = entry.get("result_data") if isinstance(entry.get("result_data"), dict) else {}
        step_id = str(step.get("id") or "").strip()
        kind = str(result_data.get("kind") or "").strip()
        if kind in {"retrieve", "raw_retrieve", "macro_query"}:
            artifact_id = str(result_data.get("artifact_id") or "").strip()
            if artifact_id and artifact_id not in raw_artifact_ids:
                raw_artifact_ids.append(artifact_id)
        if step_id == "sandbox_tool":
            tool_input = step.get("tool_input") if isinstance(step.get("tool_input"), dict) else {}
            input_artifact_ids = [
                str(item).strip()
                for item in (tool_input.get("artifact_ids") or [])
                if isinstance(item, str) and str(item).strip()
            ]
            for artifact_id in input_artifact_ids:
                if artifact_id and artifact_id not in raw_artifact_ids:
                    raw_artifact_ids.append(artifact_id)
            created_ids = [
                str(item).strip()
                for item in (result_data.get("created_artifact_ids") or [])
                if isinstance(item, str) and str(item).strip()
            ]
            for artifact_id in created_ids:
                if artifact_id and artifact_id not in calc_artifact_ids:
                    calc_artifact_ids.append(artifact_id)
            calc_rows.append(
                {
                    "section": "calculation",
                    "artifact_id": ", ".join(created_ids[:4]),
                    "artifact_label": "Sandbox calculation",
                    "source": "",
                    "detail": "Derived calculation",
                    "value": "",
                    "method": _safe_sheet_text(
                        str(tool_input.get("sandbox_request") or "").strip()
                    ),
                }
            )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    source_lines: List[str] = []
    seen_source_lines: set[str] = set()
    for artifact_id in raw_artifact_ids + calc_artifact_ids:
        record = artifact_map.get(artifact_id)
        if not record:
            continue
        for line in _source_reference_lines(record.get("source_references")):
            if line in seen_source_lines:
                continue
            seen_source_lines.add(line)
            source_lines.append(line)
    sheet.append(["Summary"])
    sheet.append(["Question", user_message])
    sheet.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")])
    if source_lines:
        sheet.append(["Sources", _safe_sheet_text(source_lines[0])])
        for line in source_lines[1:12]:
            sheet.append(["", _safe_sheet_text(line)])
    sheet.append([])
    chart_headers, chart_rows = _chart_table(chart_spec or {})
    if chart_headers and chart_rows:
        sheet.append(["Presented data"])
        _write_table(sheet, chart_headers, chart_rows)
        sheet.append([])

    calc_headers, calc_table_rows = _summary_calc_table(calc_rows)
    if calc_table_rows:
        sheet.append(["Calculations"])
        _write_table(sheet, calc_headers, calc_table_rows)

    widths = {"A": 18, "B": 44, "C": 20, "D": 20, "E": 20, "F": 20}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    used_sheet_names = {"Summary"}
    for artifact_id in raw_artifact_ids:
        record = artifact_map.get(artifact_id)
        if not record:
            continue
        raw_sheet = workbook.create_sheet(
            title=_safe_sheet_name(
                str(record.get("label") or record.get("artifact_id") or "Raw data"),
                used_sheet_names,
            )
        )
        _write_raw_artifact_sheet(raw_sheet, record)
    _apply_export_theme(workbook)

    run_dir = _ensure_runtime_dirs(conversation_id)
    download_filename = _safe_export_filename(user_message)
    path = run_dir / "artifacts" / f"answer_export_{len(state.artifacts) + 1:03d}.xlsx"
    workbook.save(path)
    source_references: List[Dict[str, Any]] = []
    for artifact_id in raw_artifact_ids:
        record = artifact_map.get(artifact_id)
        refs = record.get("source_references") if isinstance(record, dict) else None
        if isinstance(refs, list):
            source_references.extend([item for item in refs if isinstance(item, dict)])
    record = _make_artifact_record(
        state=state,
        path=path,
        kind="answer_export",
        label="Excel export",
        summary=_truncate(f"Summary workbook with clean tables and raw data sheets for '{user_message}'.", 300),
        source_references=source_references[:12],
    )
    record["download_filename"] = download_filename
    state.latest_export_artifact_id = record["artifact_id"]
    return record["artifact_id"]


def _persist_shortlist_artifact(
    *,
    state,
    conversation_id: str,
    kind: str,
    query: str,
    candidates: List[Dict[str, Any]],
) -> Dict[str, Any]:
    run_dir = _ensure_runtime_dirs(conversation_id)
    safe_kind = re.sub(r"[^a-z0-9_]+", "_", kind.lower()).strip("_") or "shortlist"
    artifact_payload = {
        "artifact_type": kind,
        "query": str(query or "").strip(),
        "candidates": [item for item in candidates if isinstance(item, dict)],
    }
    artifact_path = run_dir / "artifacts" / f"{safe_kind}_{len(state.artifacts) + 1:03d}.json"
    _write_json(artifact_path, artifact_payload)
    label_prefix = "ABS dataset shortlist" if kind == "abs_dataset_shortlist" else "Macro indicator shortlist"
    record = _make_artifact_record(
        state=state,
        path=artifact_path,
        kind=kind,
        label=f"{label_prefix}: {str(query or '').strip()[:60]}".strip(),
        summary=_truncate(
            f"{label_prefix} for '{query}' with {len(candidates)} candidates.",
            300,
        ),
    )
    return record


def _truncate(text: Any, limit: int = 180) -> str:
    value = str(text or "").replace("\n", " ").strip()
    return value if len(value) <= limit else value[: limit - 1] + "…"


def _json_text(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _strip_html(value: str) -> str:
    text = re.sub(r"(?is)<script[^>]*>.*?</script>", " ", value or "")
    text = re.sub(r"(?is)<style[^>]*>.*?</style>", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = unescape(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _to_list(value: Any) -> List[Any]:
    if isinstance(value, list):
        return value
    if value is None:
        return []
    return [value]


def _clean_string_list(value: Any) -> List[str]:
    values = value if isinstance(value, list) else [value]
    result: List[str] = []
    for item in values:
        text = str(item or "").strip()
        if text and text not in result:
            result.append(text)
    return result


def _tokenize_query(text: str) -> List[str]:
    tokens = re.findall(r"[a-z0-9]+", str(text or "").lower())
    return [token for token in tokens if len(token) > 2 and token not in STOPWORDS]


def _score_text_match(query: str, *haystacks: str) -> int:
    tokens = _tokenize_query(query)
    if not tokens:
        return 0
    corpus = " ".join(str(item or "").lower() for item in haystacks)
    score = 0
    for token in tokens:
        if token in corpus:
            score += 3 if re.search(rf"\b{re.escape(token)}\b", corpus) else 1
    return score


def _query_explicitly_requests_census(query: str) -> bool:
    text = str(query or "").lower()
    if not text:
        return False
    explicit_patterns = (
        r"\bcensus\b",
        r"\bpopulation census\b",
        r"\b2021 census\b",
        r"\b2026 census\b",
        r"\b2016 census\b",
    )
    return any(re.search(pattern, text) for pattern in explicit_patterns)


def _is_census_dataflow(dataset_id: str, title: str, description: str) -> bool:
    corpus = " ".join(
        part.strip().lower()
        for part in (dataset_id, title, description)
        if str(part or "").strip()
    )
    if not corpus:
        return False
    return bool(
        re.search(r"\bcensus\b", corpus)
        or "census of population and housing" in corpus
    )


def _detect_plan_reply(user_message: str) -> str:
    text = str(user_message or "").strip()
    if PLAN_APPROVAL_RE.match(text):
        return "approve"
    if PLAN_REJECT_RE.match(text):
        return "revise"
    return "revise"


def _lookup_concept_label(metadata: Dict[str, Any], concept_id: str) -> str:
    for concept in _to_list(metadata.get("concepts")):
        if not isinstance(concept, dict):
            continue
        if str(concept.get("id") or "").strip() != concept_id:
            continue
        return str(concept.get("name") or concept.get("description") or concept_id).strip()
    return concept_id


def _normalize_plan_context(plan_context: Any, *, fallback_question: str) -> Dict[str, Any]:
    context = dict(plan_context) if isinstance(plan_context, dict) else {}
    question = str(context.get("question") or "").strip() or str(fallback_question or "").strip()
    selected_dataset_ids = _clean_string_list(context.get("selected_dataset_ids"))
    allow_raw_discovery = bool(context.get("allow_raw_discovery"))
    await_user_input = bool(context.get("await_user_input"))
    normalized = {
        "question": question,
        "selected_dataset_ids": selected_dataset_ids,
        "allow_raw_discovery": allow_raw_discovery,
    }
    if await_user_input:
        normalized["await_user_input"] = True
    return normalized


def _build_plan_state(state, *, user_message: str = "") -> Dict[str, Any]:
    pending = state.pending_plan if isinstance(state.pending_plan, dict) else {}
    status = str(pending.get("status") or "none").strip() or "none"
    plan_context = pending.get("plan_context") if isinstance(pending.get("plan_context"), dict) else None
    approved_plan = plan_context if status == "approved" else None
    payload = {
        "status": status,
        "approved_plan": approved_plan,
        "pending_plan_summary": str(pending.get("plan_markdown") or "").strip()[:1200] if status == "awaiting_approval" else "",
        "pending_plan_context": plan_context if status == "awaiting_approval" else None,
    }
    route_hint = _detect_provider_route(user_message)
    if route_hint:
        payload.update(route_hint)
    return payload


def _normalize_provider_route(route: Any) -> str:
    clean = str(route or "").strip().lower()
    if clean == "abs":
        return "aus"
    return clean if clean in {"aus", "macro"} else ""


def _detect_provider_route(user_message: str) -> Dict[str, str]:
    query = str(user_message or "").strip().lower()
    if not query:
        return {}

    has_abs_hint = any(token in query for token in ABS_QUERY_HINTS)
    has_macro_term = any(token in query for token in MACRO_QUERY_HINTS)
    has_foreign_country = any(token in query for token in NON_ABS_COUNTRY_HINTS)
    mentions_australia = any(token in query for token in {"australia", "australian", "aus"})
    has_comparison = any(token in query for token in {" vs ", " versus ", "compare", "comparison"})
    has_macro_provider = any(token in query for token in {"world bank", "worldbank", "imf", "oecd", "comtrade", "un comtrade", "uncomtrade"})

    if has_abs_hint and has_macro_term and has_foreign_country:
        return {
            "preferred_tool": "macro_data_tool",
            "provider_route": "macro",
            "routing_reason": "The query mixes Australian domestic wording with broader non-Australian macro comparison, so macro is the safer default hint.",
        }

    if has_abs_hint:
        return {
            "preferred_tool": "aus_metadata_tool",
            "provider_route": "aus",
            "routing_reason": "The query appears focused on Australian domestic data.",
        }

    if has_macro_provider:
        return {
            "preferred_tool": "macro_data_tool",
            "provider_route": "macro",
            "routing_reason": "The query explicitly names a macro data provider.",
        }

    if has_foreign_country and not has_abs_hint:
        return {
            "preferred_tool": "macro_data_tool",
            "provider_route": "macro",
            "routing_reason": "The query names non-Australian countries and is more likely to need global macro sources.",
        }

    if has_macro_term and mentions_australia and has_foreign_country and has_comparison:
        return {
            "preferred_tool": "macro_data_tool",
            "provider_route": "macro",
            "routing_reason": "The query is a cross-country macro comparison and should prefer macro provider retrieval.",
        }

    if has_macro_term and (has_foreign_country or has_comparison or not has_abs_hint):
        return {
            "preferred_tool": "macro_data_tool",
            "provider_route": "macro",
            "routing_reason": "The query looks like a non-ABS macro comparison and should prefer macro provider retrieval.",
        }

    return {
        "preferred_tool": "aus_metadata_tool",
        "provider_route": "aus",
        "routing_reason": "Defaulting to Australian domestic data retrieval first.",
    }
def _summarize_tool_input(tool_input: Dict[str, Any]) -> str:
    if not isinstance(tool_input, dict):
        return "invalid tool input"
    parts: List[str] = []
    for key in (
        "route",
        "action",
        "candidateId",
        "datasetId",
        "dataKey",
        "searchQuery",
        "query",
        "url",
        "allCountries",
        "startYear",
        "endYear",
        "startPeriod",
        "endPeriod",
    ):
        value = tool_input.get(key)
        if value is None or value == "" or value == [] or value == {}:
            continue
        else:
            parts.append(f"{key}={value}")
    artifact_ids = tool_input.get("artifact_ids")
    if isinstance(artifact_ids, list) and artifact_ids:
        parts.append(f"artifact_ids={','.join(str(item) for item in artifact_ids[:6])}")
    candidate_ids = tool_input.get("candidateIds")
    if isinstance(candidate_ids, list) and candidate_ids:
        parts.append(f"candidateIds={','.join(str(item) for item in candidate_ids[:6])}")
    countries = tool_input.get("countries")
    if isinstance(countries, list) and countries:
        parts.append(f"countries={','.join(str(item) for item in countries[:10])}")
    code = str(tool_input.get("code") or "").strip()
    if code:
        parts.append(f"code_preview={_truncate(code, 120)}")
    sandbox_request = str(tool_input.get("sandbox_request") or "").strip()
    if sandbox_request:
        parts.append(f"sandbox_request={_truncate(sandbox_request, 160)}")
    return "; ".join(parts) if parts else "no key inputs"
def _build_discover_payload(search_query: str, limit: int = 40) -> Dict[str, Any]:
    logger.info(
        'ABS FTS discover start query="%s" limit=%s',
        _truncate(search_query, 200),
        limit,
    )
    payload = list_dataflows(
        force_refresh=False,
        search_query=search_query,
        limit=limit,
    )
    flows = payload.get("dataflows") if isinstance(payload, dict) else []
    candidates: List[Dict[str, Any]] = []
    explicit_census_query = _query_explicitly_requests_census(search_query)
    for item in _to_list(flows):
        if not isinstance(item, dict):
            continue
        entry = {
            "dataset_id": str(item.get("id") or "").strip(),
            "title": str(item.get("name") or item.get("id") or "").strip(),
            "description": str(item.get("description") or "").strip(),
            "flow_type": str(item.get("flowType") or "").strip(),
            "source_type": str(item.get("sourceType") or "").strip(),
            "requires_metadata_before_retrieval": bool(item.get("requiresMetadataBeforeRetrieval")),
        }
        if not entry["dataset_id"]:
            continue
        entry["score"] = _score_text_match(
            search_query,
            entry["dataset_id"],
            entry["title"],
            entry["description"],
        ) if search_query else 0
        entry["is_census"] = _is_census_dataflow(
            entry["dataset_id"],
            entry["title"],
            entry["description"],
        )
        if entry["is_census"] and not explicit_census_query:
            entry["score"] -= 100
        candidates.append(entry)
    candidates.sort(key=lambda item: (-int(item.get("score") or 0), item["dataset_id"]))
    trimmed = [
        {
            "dataset_id": item["dataset_id"],
            "title": item["title"],
            "description": item["description"],
            "flow_type": item.get("flow_type") or "",
            "source_type": item.get("source_type") or "",
            "requires_metadata_before_retrieval": bool(item.get("requires_metadata_before_retrieval")),
        }
        for item in candidates[:limit]
        if search_query == "" or int(item.get("score") or 0) > 0 or len(candidates) <= limit
    ]
    logger.info(
        "ABS FTS discover complete query=%r candidates=%s shortlisted=%s top=%s",
        search_query,
        len(candidates),
        len(trimmed),
        [item["dataset_id"] for item in trimmed[:3]],
    )
    return {
        "search_query": search_query,
        "datasets": trimmed,
    }

def _build_raw_metadata_payload(dataset_id: str, metadata: Dict[str, Any]) -> Dict[str, Any]:
    dimensions = [
        item
        for item in _to_list(metadata.get("dimensions"))
        if isinstance(item, dict)
    ]
    concepts = [
        item
        for item in _to_list(metadata.get("concepts"))
        if isinstance(item, dict)
    ]
    codelists = {
        str(item.get("id") or "").strip(): item
        for item in _to_list(metadata.get("codelists"))
        if isinstance(item, dict) and str(item.get("id") or "").strip()
    }
    concept_lookup = {
        str(item.get("id") or "").strip(): item
        for item in concepts
        if str(item.get("id") or "").strip()
    }

    dimension_order: List[str] = []
    anchor_dimension_rows: List[Dict[str, Any]] = []
    for dimension in sorted(dimensions, key=lambda item: int(item.get("position") or 0)):
        dimension_id = str(dimension.get("id") or "").strip()
        if not dimension_id:
            continue
        codelist_ref = dimension.get("codelist") if isinstance(dimension.get("codelist"), dict) else {}
        codelist_id = str(
            dimension.get("codeList")
            or codelist_ref.get("id")
            or ""
        ).strip()
        codes = codelists.get(codelist_id, {})
        concept_id = str(dimension.get("conceptId") or "").strip()
        concept_meta = concept_lookup.get(concept_id, {})
        concept_name = str(
            concept_meta.get("name")
            or concept_meta.get("description")
            or concept_id
            or dimension_id
        ).strip()
        code_values = [
            {
                "code": str(code.get("id") or "").strip(),
                "label": str(code.get("name") or "").strip(),
                "description": str(code.get("description") or "").strip(),
            }
            for code in _to_list(codes.get("codes"))
            if isinstance(code, dict) and str(code.get("id") or "").strip()
        ]
        dimension_order.append(dimension_id)
        anchor_type = _normalize_anchor_type(dimension_id, concept_id)
        if anchor_type:
            anchor_dimension_rows.append(
                {
                    "anchor_type": anchor_type,
                    "dimension_id": dimension_id,
                    "anchor_description": concept_name,
                    "position": int(dimension.get("position") or 0),
                    "anchor_codes": code_values,
                }
            )

    def _wildcard_template_for(anchor_dimension_id: str) -> str:
        parts: List[str] = []
        for dim_id in dimension_order:
            if dim_id == anchor_dimension_id:
                parts.append("{" + dim_id + "}")
            else:
                parts.append("")
        return ".".join(parts) if parts else "all"

    anchor_candidates_by_type: Dict[str, Dict[str, Any]] = {}
    for row in anchor_dimension_rows:
        anchor_type = str(row.get("anchor_type") or "").strip()
        dimension_id = str(row.get("dimension_id") or "").strip()
        candidate = {
            "anchor_type": anchor_type,
            "anchor_description": str(row.get("anchor_description") or anchor_type).strip(),
            "wildcard_data_key_template": _wildcard_template_for(dimension_id),
            "anchor_codes": row.get("anchor_codes") or [],
        }
        existing = anchor_candidates_by_type.get(anchor_type)
        if existing is None:
            anchor_candidates_by_type[anchor_type] = candidate
            continue
        existing_dimension = _anchor_dimension_from_template(str(existing.get("wildcard_data_key_template") or ""))
        existing_rank = dimension_order.index(existing_dimension) if existing_dimension in dimension_order else 999
        current_rank = dimension_order.index(dimension_id) if dimension_id in dimension_order else 999
        if current_rank < existing_rank:
            anchor_candidates_by_type[anchor_type] = candidate
    anchor_candidates = list(anchor_candidates_by_type.values())
    anchor_candidates.sort(
        key=lambda item: (
            -_anchor_priority(str(item.get("anchor_type") or "")),
        )
    )

    return {
        "dataset_id": dataset_id,
        "anchor_candidates": anchor_candidates,
    }


def _normalize_anchor_type(dimension_id: str, concept_id: str) -> str:
    text = " ".join(
        part.strip().upper()
        for part in (dimension_id, concept_id)
        if str(part or "").strip()
    )
    if "MEASURE" in text:
        return "MEASURE"
    if "DATA_ITEM" in text or text.endswith("ITEM") or " ITEM" in text:
        return "DATA_ITEM"
    if any(token in text for token in {"CAT", "CATEGORY", "SUPG", "SUPC", "PRODUCT", "COMMODITY", "INDUSTRY", "SECTOR", "FLOW"}):
        return "CATEGORY"
    return ""


def _anchor_dimension_from_template(template: str) -> str:
    match = re.search(r"\{([^}]+)\}", str(template or ""))
    return str(match.group(1) if match else "").strip()


def _anchor_priority(anchor_type: str) -> int:
    priority_map = {
        "DATA_ITEM": 100,
        "MEASURE": 90,
        "CATEGORY": 80,
    }
    return priority_map.get(str(anchor_type or "").strip().upper(), 0)


def _summarize_raw_metadata_payload(payload: Dict[str, Any]) -> str:
    return _json_text(payload)


def _tool_result(summary: str, result_data: Any = None) -> Dict[str, Any]:
    payload = {
        "summary": str(summary or "").strip(),
    }
    if result_data is not None:
        payload["result_data"] = result_data
    return payload


def _normalize_result_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.netloc.endswith("duckduckgo.com") and parsed.path.startswith("/l/"):
        redirected = parse_qs(parsed.query).get("uddg") or []
        if redirected:
            return redirected[0]
    return url


def _build_abs_source_references(dataset_id: str, title: str = "") -> List[Dict[str, Any]]:
    clean_dataset_id = str(dataset_id or "").strip()
    clean_title = str(title or "").strip() or clean_dataset_id
    if not clean_dataset_id:
        return []
    return [
        {
            "provider": "ABS",
            "dataset_id": clean_dataset_id,
            "title": clean_title,
        }
    ]


@lru_cache(maxsize=1)
def _load_custom_domestic_flows() -> Dict[str, Dict[str, Any]]:
    custom_path = Path(__file__).resolve().parents[2] / "CUSTOM_AUS_DATAFLOWS.json"
    if not custom_path.exists():
        return {}
    try:
        payload = json.loads(custom_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    flows = payload.get("flows") if isinstance(payload, dict) else []
    result: Dict[str, Dict[str, Any]] = {}
    for item in _to_list(flows):
        if not isinstance(item, dict):
            continue
        flow_id = str(item.get("id") or "").strip()
        if flow_id:
            result[flow_id] = item
    return result


def _base_dataset_id(dataset_id: str) -> str:
    parts = [part.strip() for part in str(dataset_id or "").split(",") if part.strip()]
    if len(parts) >= 2:
        return parts[1]
    return str(dataset_id or "").strip()


def _custom_domestic_flow(dataset_id: str) -> Optional[Dict[str, Any]]:
    base_id = _base_dataset_id(dataset_id)
    if not base_id:
        return None
    return _load_custom_domestic_flows().get(base_id)


def _is_custom_domestic_dataset(dataset_id: str) -> bool:
    if str(dataset_id or "").startswith("CUSTOM_AUS,"):
        return True
    return _custom_domestic_flow(dataset_id) is not None


def _resolve_dcceew_dataset_python(
    *,
    flow: Dict[str, Any],
    data_key: str,
    detail: str,
) -> Dict[str, Any]:
    source_url = str(flow.get("sourceUrl") or "").strip()
    if not source_url:
        raise RuntimeError(f"Custom flow {flow.get('id') or ''} is missing sourceUrl.")

    parser_path = Path(__file__).resolve().parents[2] / "scripts" / "dcceew_aes_xlsx.py"
    if not parser_path.exists():
        raise RuntimeError(f"DCCEEW parser script not found at {parser_path}.")

    logger.info(
        "DCCEEW python retrieval start datasetId=%s sourceUrl=%s retrievalMethod=%s",
        str(flow.get("id") or "").strip(),
        source_url,
        "python_backend_requests_fetch_wsl" if os.name == "nt" else "python_backend_requests_fetch",
    )

    temp_path: Optional[Path] = None
    try:
        with tempfile.NamedTemporaryFile(prefix="dcceew-aes-", suffix=".xlsx", delete=False) as handle:
            temp_path = Path(handle.name)

        if os.name == "nt":
            windows_temp_path = temp_path.as_posix()
            wsl_target = subprocess.run(
                ["wsl.exe", "wslpath", "-a", windows_temp_path],
                check=True,
                capture_output=True,
                text=True,
                timeout=30,
            ).stdout.strip()
            download_script = (
                "import requests, sys; "
                "url=sys.argv[1]; out=sys.argv[2]; "
                "r=requests.get(url, headers={'User-Agent':'Mozilla/5.0'}, timeout=(20,60)); "
                "r.raise_for_status(); "
                "data=r.content; "
                "assert len(data) >= 4 and data[:2] == b'PK'; "
                "open(out,'wb').write(data); "
                "print(len(data))"
            )
            download_cmd = ["wsl.exe", "python3", "-c", download_script, source_url, wsl_target]
            retrieval_method = "python_backend_requests_fetch_wsl"
        else:
            response = requests.get(
                source_url,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=(20, 60),
            )
            response.raise_for_status()
            payload = response.content
            if len(payload) < 4 or payload[:2] != b"PK":
                raise RuntimeError("Downloaded DCCEEW payload is not a valid XLSX workbook.")
            temp_path.write_bytes(payload)
            download_cmd = None
            retrieval_method = "python_backend_requests_fetch"

        if download_cmd is not None:
            completed_download = subprocess.run(
                download_cmd,
                check=True,
                capture_output=True,
                text=True,
                timeout=120,
            )
            logger.info(
                "DCCEEW workbook download complete datasetId=%s retrievalMethod=%s bytes=%s",
                str(flow.get("id") or "").strip(),
                retrieval_method,
                str(completed_download.stdout or "").strip(),
            )

        command = [
            settings.python_binary,
            str(parser_path),
            "resolve",
            "--xlsx",
            str(temp_path),
            "--dataset-id",
            str(flow.get("id") or "").strip(),
            "--agency-id",
            str(flow.get("agencyID") or "").strip(),
            "--version",
            str(flow.get("version") or "").strip(),
            "--name",
            str(flow.get("name") or "").strip(),
            "--description",
            str(flow.get("description") or "").strip(),
            "--data-key",
            data_key or "all",
            "--detail",
            detail or "full",
            "--curation-json",
            json.dumps(flow.get("curation") or {}),
        ]
        completed = subprocess.run(
            command,
            check=True,
            capture_output=True,
            text=True,
            timeout=120,
        )
        logger.info(
            "DCCEEW python retrieval complete datasetId=%s retrievalMethod=%s",
            str(flow.get("id") or "").strip(),
            retrieval_method,
        )
        return json.loads(completed.stdout)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(
            "DCCEEW parser failed. "
            f"stdout={str(exc.stdout or '')[:1000]!r} stderr={str(exc.stderr or '')[:1000]!r}"
        ) from exc
    finally:
        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except Exception:
                pass


def _build_domestic_source_references(dataset_id: str, title: str = "") -> List[Dict[str, Any]]:
    custom_flow = _custom_domestic_flow(dataset_id)
    if not custom_flow:
        return _build_abs_source_references(dataset_id, title)

    clean_dataset_id = _base_dataset_id(dataset_id)
    clean_title = str(title or custom_flow.get("name") or clean_dataset_id).strip()
    reference = {
        "provider": str(custom_flow.get("sourceOrganization") or "Custom Australian source").strip(),
        "dataset_id": clean_dataset_id,
        "title": clean_title,
    }
    source_url = str(custom_flow.get("sourceUrl") or "").strip()
    source_page_url = str(custom_flow.get("sourcePageUrl") or "").strip()
    if source_url:
        reference["url"] = source_url
    if source_page_url:
        reference["page_url"] = source_page_url
    return [reference]


def _search_web(query: str, max_results: int = 5) -> Dict[str, Any]:
    logger.info(
        'Web retrieval request kind=search url="%s" query="%s" max_results=%s',
        WEB_SEARCH_URL,
        _truncate(query, 280),
        max_results,
    )
    try:
        response = httpx.get(
            WEB_SEARCH_URL,
            params={"q": query},
            headers={
                "User-Agent": WEB_USER_AGENT,
            },
            follow_redirects=True,
            timeout=20,
        )
        response.raise_for_status()
    except Exception as exc:
        response = exc.response if isinstance(exc, httpx.HTTPStatusError) else None
        logger.error(
            'Web retrieval error kind=search url="%s" query="%s" status=%s error="%s" body="%s"',
            WEB_SEARCH_URL,
            _truncate(query, 280),
            getattr(response, "status_code", ""),
            _truncate(exc, 300),
            _truncate(response.text if response is not None else "", 400),
        )
        raise
    html = response.text

    links = WEB_RESULT_LINK_RE.findall(html)
    snippets = [_strip_html(item) for item in WEB_RESULT_SNIPPET_RE.findall(html)]
    results: List[Dict[str, Any]] = []

    for index, (raw_url, raw_title) in enumerate(links[:max_results]):
        url = _normalize_result_url(raw_url)
        title = _strip_html(raw_title)
        if not url or not title:
            continue
        domain = urlparse(url).netloc
        results.append(
            {
                "rank": len(results) + 1,
                "title": title,
                "url": url,
                "domain": domain,
                "snippet": snippets[index] if index < len(snippets) else "",
            }
        )

    return {
        "query": query,
        "engine": "duckduckgo_html",
        "results": results,
    }


def _fetch_web_page(url: str) -> Dict[str, Any]:
    logger.info(
        'Web retrieval request kind=fetch url="%s"',
        _truncate(url, 500),
    )
    try:
        response = httpx.get(
            url,
            headers={"User-Agent": WEB_USER_AGENT},
            follow_redirects=True,
            timeout=20,
        )
        response.raise_for_status()
    except Exception as exc:
        response = exc.response if isinstance(exc, httpx.HTTPStatusError) else None
        logger.error(
            'Web retrieval error kind=fetch url="%s" status=%s error="%s" body="%s"',
            _truncate(url, 500),
            getattr(response, "status_code", ""),
            _truncate(exc, 300),
            _truncate(response.text if response is not None else "", 400),
        )
        raise
    html = response.text
    title_match = re.search(r"(?is)<title[^>]*>(.*?)</title>", html)
    title = _strip_html(title_match.group(1)) if title_match else ""
    cleaned = _strip_html(html)
    final_url = str(response.url)
    return {
        "url": final_url,
        "domain": urlparse(final_url).netloc,
        "title": title or urlparse(final_url).netloc,
        "text_preview": cleaned[:12000],
    }


def _execute_web_search_tool(
    *,
    tool_input: Dict[str, Any],
    state,
    conversation_id: str,
) -> Dict[str, Any]:
    action = str(tool_input.get("action") or "search").strip().lower()
    run_dir = _ensure_runtime_dirs(conversation_id)

    if action == "search":
        query = str(tool_input.get("query") or "").strip()
        if not query:
            raise RuntimeError("web_search_tool action search requires query")
        max_results = int(tool_input.get("maxResults") or 5)
        logger.info(
            'Web search start cid=%s action=search query="%s" max_results=%s',
            conversation_id,
            _truncate(query, 280),
            max(1, min(max_results, 8)),
        )
        payload = _search_web(query, max_results=max(1, min(max_results, 8)))
        results = payload.get("results") or []
        preview = []
        for item in results[:3]:
            if not isinstance(item, dict):
                continue
            title = str(item.get("title") or "").strip()
            url = str(item.get("url") or "").strip()
            if title or url:
                preview.append(f"{title} <{url}>".strip())
        logger.info(
            'Web search complete cid=%s action=search query="%s" result_count=%s preview="%s"',
            conversation_id,
            _truncate(query, 280),
            len(results),
            _truncate(" | ".join(preview), 500),
        )
        artifact_path = run_dir / "artifacts" / f"web_search_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, payload)
        record = _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="web_search_results",
            label=f"Web search: {query}",
            summary=_truncate(f"Web search for '{query}' returned {len(results)} results.", 300),
        )
        return _tool_result(
            (
                f"Web search for '{query}' returned {len(results)} results.\n"
                f"Created artifact: {record['artifact_id']}. Use sandbox to inspect or compare the results."
            ),
            {
                "kind": "web_search",
                "query": query,
                "results": results,
                "artifact_id": record["artifact_id"],
            },
        )

    if action == "fetch":
        url = str(tool_input.get("url") or "").strip()
        if not url:
            raise RuntimeError("web_search_tool action fetch requires url")
        logger.info(
            'Web search start cid=%s action=fetch url="%s"',
            conversation_id,
            _truncate(url, 500),
        )
        payload = _fetch_web_page(url)
        logger.info(
            'Web search complete cid=%s action=fetch url="%s" final_url="%s" title="%s" domain="%s"',
            conversation_id,
            _truncate(url, 500),
            _truncate(str(payload.get("url") or ""), 500),
            _truncate(str(payload.get("title") or ""), 200),
            _truncate(str(payload.get("domain") or ""), 120),
        )
        artifact_path = run_dir / "artifacts" / f"web_page_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, payload)
        record = _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="web_page",
            label=f"Web page: {payload.get('title') or payload.get('domain')}",
            summary=_truncate(f"Fetched web page {payload.get('url')}.", 300),
        )
        return _tool_result(
            (
                f"Fetched web page {payload.get('domain') or payload.get('url')}.\n"
                f"Created artifact: {record['artifact_id']}. Use sandbox to inspect the extracted text."
            ),
            {
                "kind": "web_page",
                "url": payload.get("url"),
                "domain": payload.get("domain"),
                "title": payload.get("title"),
                "text_preview": payload.get("text_preview"),
                "artifact_id": record["artifact_id"],
            },
        )

    raise RuntimeError(f"Unsupported web_search_tool action: {action}")


def _execute_provider_route_tool(
    *,
    tool_input: Dict[str, Any],
    state: HarnessState,
    conversation_id: str,
) -> Dict[str, Any]:
    route = _normalize_provider_route(tool_input.get("route"))
    if not route:
        raise RuntimeError("provider_route_tool requires route to be one of aus or macro.")
    reason = str(tool_input.get("reason") or "").strip()
    search_query = str(tool_input.get("searchQuery") or "").strip()
    if not search_query:
        raise RuntimeError("provider_route_tool requires searchQuery for both aus and macro routes.")
    logger.info(
        'Provider route selected cid=%s route=%s reason="%s" search_query="%s"',
        conversation_id,
        route,
        _truncate(reason, 220),
        _truncate(search_query, 220),
    )
    return _tool_result(
        "\n".join(
            line
            for line in [
                f"Selected provider route: {route}.",
                f"Planned retrieval query: {search_query}." if search_query else "",
                f"Reason: {reason}." if reason else "",
            ]
            if line
        ),
        {
            "kind": "provider_route_selection",
            "route": route,
            "reason": reason,
            "search_query": search_query,
        },
    )


def _execute_macro_data_tool(
    *,
    tool_input: Dict[str, Any],
    state,
    conversation_id: str,
) -> Dict[str, Any]:
    action = str(tool_input.get("action") or "").strip().lower()
    if action == "retrieve":
        action = "retrieve"
    if action not in {"query", "metadata", "retrieve"}:
        raise RuntimeError(f"Unsupported macro_data_tool action: {action}")

    query = str(tool_input.get("query") or "").strip()
    if not query:
        raise RuntimeError(f"macro_data_tool action {action} requires query")

    raw_candidate_ids = tool_input.get("candidateIds")
    candidate_ids: List[str] = []
    if isinstance(raw_candidate_ids, list):
        for item in raw_candidate_ids:
            clean = str(item or "").strip()
            if clean and clean not in candidate_ids:
                candidate_ids.append(clean)
    candidate_id = str(tool_input.get("candidateId") or "").strip()
    if candidate_id and candidate_id not in candidate_ids:
        candidate_ids.insert(0, candidate_id)
    candidate_ids = candidate_ids[:3]

    if action == "metadata":
        if not candidate_ids:
            raise RuntimeError("macro_data_tool action metadata requires candidateId")
        metadata_payload = get_macro_candidate_metadata(candidate_ids[0], query)
        run_dir = _ensure_runtime_dirs(conversation_id)
        artifact_payload = {
            "artifact_type": "macro_metadata",
            "query": query,
            "candidate_id": candidate_ids[0],
            "metadata": metadata_payload,
        }
        artifact_path = run_dir / "artifacts" / f"macro_metadata_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, artifact_payload)
        record = _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="macro_metadata",
            label=f"Macro metadata: {candidate_ids[0]}",
            summary=_truncate(f"Macro metadata for {candidate_ids[0]} with Comtrade reference dimensions.", 300),
        )
        return _tool_result(
            _truncate(
                f"Prepared macro metadata for '{query}' using candidate {candidate_ids[0]}. Artifact: {record['artifact_id']}.",
                300,
            ),
            {
                "kind": "macro_metadata",
                "query": query,
                "candidate_id": candidate_ids[0],
                "artifact_id": record["artifact_id"],
                "metadata": metadata_payload,
            },
        )

    if action == "retrieve" and not candidate_ids:
        raise RuntimeError("macro_data_tool action retrieve requires candidateId or candidateIds")
    raw_countries = tool_input.get("countries")
    countries = [
        str(item or "").strip().upper()
        for item in raw_countries
        if str(item or "").strip()
    ] if isinstance(raw_countries, list) else []
    all_countries = bool(tool_input.get("allCountries"))
    raw_start_year = tool_input.get("startYear")
    start_year = int(raw_start_year) if isinstance(raw_start_year, int) or (isinstance(raw_start_year, str) and str(raw_start_year).strip().isdigit()) else None
    raw_end_year = tool_input.get("endYear")
    end_year = int(raw_end_year) if isinstance(raw_end_year, int) or (isinstance(raw_end_year, str) and str(raw_end_year).strip().isdigit()) else None
    reporter_codes = _to_list(tool_input.get("reporterCodes"))
    partner_codes = _to_list(tool_input.get("partnerCodes"))
    hs_codes = _to_list(tool_input.get("hsCodes"))
    flow_code = str(tool_input.get("flowCode") or "").strip().upper()
    frequency_code = str(tool_input.get("frequencyCode") or "").strip().upper()

    logger.info(
        'Macro query start cid=%s action=%s query="%s" candidate_ids="%s" countries="%s" all_countries=%s years=%s:%s reporter_codes="%s" partner_codes="%s" flow=%s frequency=%s hs_codes="%s"',
        conversation_id,
        action,
        _truncate(query, 220),
        ",".join(candidate_ids),
        ",".join(countries),
        all_countries,
        start_year,
        end_year,
        ",".join(str(item or "").strip() for item in reporter_codes if str(item or "").strip()),
        ",".join(str(item or "").strip() for item in partner_codes if str(item or "").strip()),
        flow_code,
        frequency_code,
        ",".join(str(item or "").strip() for item in hs_codes if str(item or "").strip()),
    )
    try:
        if action == "retrieve":
            if len(candidate_ids) == 1:
                payload = retrieve_macro_candidate(
                    candidate_ids[0],
                    query,
                    countries=countries,
                    all_countries=all_countries,
                    start_year=start_year,
                    end_year=end_year,
                    reporter_codes=[str(item or "").strip() for item in reporter_codes if str(item or "").strip()],
                    partner_codes=[str(item or "").strip() for item in partner_codes if str(item or "").strip()],
                    flow_code=flow_code,
                    frequency_code=frequency_code,
                    hs_codes=[str(item or "").strip() for item in hs_codes if str(item or "").strip()],
                )
                payload["attempted_candidate_ids"] = list(candidate_ids)
                payload["selected_candidate_id"] = candidate_ids[0]
            else:
                results_by_id: Dict[str, Dict[str, Any]] = {}
                errors_by_id: Dict[str, str] = {}
                with ThreadPoolExecutor(max_workers=len(candidate_ids)) as executor:
                    future_map = {
                        executor.submit(
                            retrieve_macro_candidate,
                            current_candidate_id,
                            query,
                            countries=countries,
                            all_countries=all_countries,
                            start_year=start_year,
                            end_year=end_year,
                            reporter_codes=[str(item or "").strip() for item in reporter_codes if str(item or "").strip()],
                            partner_codes=[str(item or "").strip() for item in partner_codes if str(item or "").strip()],
                            flow_code=flow_code,
                            frequency_code=frequency_code,
                            hs_codes=[str(item or "").strip() for item in hs_codes if str(item or "").strip()],
                        ): current_candidate_id
                        for current_candidate_id in candidate_ids
                    }
                    for future in as_completed(future_map):
                        current_candidate_id = future_map[future]
                        try:
                            results_by_id[current_candidate_id] = future.result()
                        except Exception as exc:
                            errors_by_id[current_candidate_id] = str(exc)
                selected_payload = None
                selected_candidate_id = ""
                for current_candidate_id in candidate_ids:
                    candidate_payload = results_by_id.get(current_candidate_id)
                    if isinstance(candidate_payload, dict):
                        selected_payload = candidate_payload
                        selected_candidate_id = current_candidate_id
                        break
                if selected_payload is None:
                    primary_error = errors_by_id.get(candidate_ids[0]) or "No viable macro candidates returned usable data."
                    raise RuntimeError(primary_error)
                payload = selected_payload
                payload["attempted_candidate_ids"] = list(candidate_ids)
                payload["selected_candidate_id"] = selected_candidate_id
                payload["candidate_failures"] = errors_by_id
        else:
            payload = run_macro_query(query)
    except Exception as exc:
        if action == "retrieve":
            logger.error(
                'Macro retrieval failed cid=%s query="%s" candidate_ids="%s" countries="%s" all_countries=%s years=%s:%s reporter_codes="%s" partner_codes="%s" flow=%s frequency=%s hs_codes="%s" error="%s"',
                conversation_id,
                _truncate(query, 220),
                ",".join(candidate_ids),
                ",".join(countries),
                all_countries,
                start_year,
                end_year,
                ",".join(str(item or "").strip() for item in reporter_codes if str(item or "").strip()),
                ",".join(str(item or "").strip() for item in partner_codes if str(item or "").strip()),
                flow_code,
                frequency_code,
                ",".join(str(item or "").strip() for item in hs_codes if str(item or "").strip()),
                _truncate(exc, 500),
            )
            remaining_candidates = [
                item for item in _to_list(getattr(state, "current_macro_indicator_shortlist", []) or [])
                if str(item.get("candidate_id") or "").strip() not in set(candidate_ids)
            ]
            state.current_macro_indicator_shortlist = remaining_candidates
            if remaining_candidates:
                return _tool_result(
                    _truncate(
                        f"Macro candidates {', '.join(candidate_ids)} were unavailable for '{query}'. Remaining shortlist candidates: {len(remaining_candidates)}.",
                        300,
                    ),
                    {
                        "kind": "macro_candidate_unavailable",
                        "query": query,
                        "candidate_id": candidate_ids[0] if candidate_ids else "",
                        "candidate_ids": candidate_ids,
                        "error": str(exc),
                        "remaining_candidates": remaining_candidates,
                    },
                )
            return _tool_result(
                _truncate(
                    f"No viable macro candidates returned usable data for '{query}'.",
                    300,
                ),
                {
                    "kind": "macro_unavailable",
                    "query": query,
                    "candidate_id": candidate_ids[0] if candidate_ids else "",
                    "candidate_ids": candidate_ids,
                    "error": str(exc),
                    "remaining_candidates": [],
                },
            )
        raise RuntimeError(f"Macro query failed: {exc}") from exc

    if not isinstance(payload, dict):
        raise RuntimeError("Macro provider returned a non-object response.")

    run_dir = _ensure_runtime_dirs(conversation_id)
    sources = payload.get("source_references") if isinstance(payload.get("source_references"), list) else []
    data_entries = payload.get("series") if isinstance(payload.get("series"), list) else []
    artifact_payload = {
        "artifact_type": "macro_query_response",
        "query": query,
        "request_context": {
            "candidate_ids": list(candidate_ids),
            "countries": list(countries),
            "all_countries": all_countries,
            "start_year": start_year,
            "end_year": end_year,
        },
        "response": payload,
        "source_references": sources,
    }
    artifact_path = run_dir / "artifacts" / f"macro_query_{len(state.artifacts) + 1:03d}.json"
    _write_json(artifact_path, artifact_payload)
    record = _make_artifact_record(
        state=state,
        path=artifact_path,
        kind="macro_query_response",
        label=f"Macro query {query[:60]}".strip(),
        summary=_truncate(
            f"Macro providers returned {len(data_entries)} series for query '{query}'.",
            300,
        ),
        source_references=sources,
    )
    logger.info(
        "Macro query complete cid=%s series=%s artifact=%s providers=%s",
        conversation_id,
        len(data_entries),
        record["artifact_id"],
        [str(item.get("provider") or "") for item in sources[:4]],
    )
    source_summary = "; ".join(
        [
            ", ".join(
                part
                for part in [
                    str(item.get("provider") or "").strip(),
                    str(item.get("indicator") or "").strip(),
                    str(item.get("series_id") or "").strip(),
                ]
                if part
            )
            for item in sources[:4]
        ]
    ).strip()
    summary_lines = [
        f"Queried macro providers for: {query}",
        f"Series returned: {len(data_entries)}.",
        f"Created artifact: {record['artifact_id']}.",
    ]
    selected_candidate_id = str(payload.get("selected_candidate_id") or "").strip()
    attempted_candidate_ids = [
        str(item or "").strip()
        for item in (payload.get("attempted_candidate_ids") or [])
        if str(item or "").strip()
    ]
    if selected_candidate_id and attempted_candidate_ids:
        summary_lines.append(
            f"Selected candidate: {selected_candidate_id} from attempted candidates {', '.join(attempted_candidate_ids)}."
        )
    if source_summary:
        summary_lines.append(f"Sources: {source_summary}.")
    return _tool_result(
        "\n".join(summary_lines),
        {
            "kind": "macro_query",
            "query": query,
            "artifact_id": record["artifact_id"],
            "series_count": len(data_entries),
            "source_references": sources,
            "provider": str(payload.get("provider") or "").strip(),
            "concept_id": str(payload.get("concept_id") or "").strip(),
            "concept_label": str(payload.get("concept_label") or "").strip(),
            "selected_candidate_id": selected_candidate_id,
            "attempted_candidate_ids": attempted_candidate_ids,
        },
    )


def _bridge_error_status_code(exc: Exception) -> Optional[int]:
    if not isinstance(exc, MCPBridgeError):
        return None
    match = re.search(r"status code (\d+)", f"{exc.stderr}\n{exc}")
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _expand_abs_shortlist_query(search_query: str) -> str:
    raw_query = str(search_query or "").strip()
    if not raw_query:
        return raw_query

    lowered = raw_query.lower()
    expanded_parts: List[str] = [raw_query]

    if any(term in lowered for term in EMPLOYMENT_INTENT_TERMS):
        extras = [
            "labour force",
            "labour account",
            "employed persons",
            "employment by industry",
            "industry division",
            "monthly",
            "quarterly",
            "time series",
        ]
        for extra in extras:
            if extra not in lowered:
                expanded_parts.append(extra)

    if "manufacturing" in lowered and "anzsic division c" not in lowered:
        expanded_parts.append("ANZSIC Division C")

    if any(phrase in lowered for phrase in ("across time", "over time")) and "time series" not in lowered:
        expanded_parts.append("time series")

    return " ".join(part for part in expanded_parts if part).strip()


def _fallback_tstest_filters(filters: Dict[str, List[str]]) -> List[Dict[str, List[str]]]:
    selected = filters.get("TSEST") or []
    if len(selected) != 1:
        return []
    current = selected[0]
    fallbacks: List[str] = []
    if current == "30":
        fallbacks = ["20", "10"]
    elif current == "20":
        fallbacks = ["10"]
    else:
        return []

    attempts: List[Dict[str, List[str]]] = []
    for code in fallbacks:
        updated = {key: list(value) for key, value in filters.items()}
        updated["TSEST"] = [code]
        attempts.append(updated)
    return attempts


def _build_retrieval_summary(
    *,
    dataset_id: str,
    entry: Dict[str, Any],
    resolved_payload: Dict[str, Any],
    filters: Dict[str, List[str]],
    data_key: str,
    artifact_id: str,
    fallback_note: str = "",
) -> str:
    series = resolved_payload.get("series") if isinstance(resolved_payload, dict) else []
    observation_count = resolved_payload.get("observationCount") if isinstance(resolved_payload, dict) else None
    lines = [
        f"Retrieved and resolved {entry['title']} datasetId {dataset_id}.",
        f"Observation count: {observation_count if observation_count is not None else 'unknown'}.",
        f"Series count: {len(series) if isinstance(series, list) else 0}.",
        f"Applied data key: {data_key}.",
    ]
    retrieval_query = resolved_payload.get("query") if isinstance(resolved_payload, dict) else {}
    if isinstance(retrieval_query, dict):
        start_period = str(retrieval_query.get("startPeriod") or "").strip()
        end_period = str(retrieval_query.get("endPeriod") or "").strip()
        if start_period or end_period:
            lines.append(
                f"Applied time range: {start_period or 'open'} to {end_period or 'open'}."
            )
    if filters:
        filter_parts = [
            f"{dim_id}={'+'.join(values)}"
            for dim_id, values in filters.items()
        ]
        lines.append(f"Applied filters: {', '.join(filter_parts)}.")
    if fallback_note:
        lines.append(fallback_note)
    dimension_labels = resolved_payload.get("dimensions") if isinstance(resolved_payload, dict) else {}
    if isinstance(dimension_labels, dict):
        lines.append(
            "Resolved dimensions: "
            + ", ".join(list(dimension_labels.keys())[:8])
            + "."
        )
    lines.append(
        f"Created artifact: {artifact_id}. Use the sandbox tool to inspect, filter, aggregate or calculate."
    )
    return "\n".join(lines)


def _refresh_abs_shortlist(
    *,
    state,
    conversation_id: str,
    route: str,
    shortlist_query: str,
) -> List[Dict[str, Any]]:
    effective_query = _expand_abs_shortlist_query(shortlist_query)
    if effective_query != shortlist_query:
        logger.info(
            'ABS shortlist query expanded cid=%s original="%s" expanded="%s"',
            conversation_id,
            _truncate(shortlist_query, 220),
            _truncate(effective_query, 320),
        )
    logger.info(
        'Post-route ABS shortlist start cid=%s route=%s query="%s"',
        conversation_id,
        route,
        _truncate(effective_query, 220),
    )
    shortlist_payload = _build_discover_payload(effective_query, limit=40)
    shortlist_items = shortlist_payload.get("datasets") if isinstance(shortlist_payload, dict) else []
    refreshed_shortlist = [item for item in _to_list(shortlist_items) if isinstance(item, dict)]
    state.current_abs_dataset_shortlist = list(refreshed_shortlist)
    shortlist_record = _persist_shortlist_artifact(
        state=state,
        conversation_id=conversation_id,
        kind="abs_dataset_shortlist",
        query=effective_query,
        candidates=refreshed_shortlist,
    )
    logger.info(
        "Post-route ABS shortlist ready cid=%s route=%s count=%s artifact=%s top=%s",
        conversation_id,
        route,
        len(refreshed_shortlist),
        shortlist_record["artifact_id"],
        [str(item.get("dataset_id") or "").strip() for item in refreshed_shortlist[:3]],
    )
    return refreshed_shortlist


def _execute_abs_data_tool(
    *,
    tool_input: Dict[str, Any],
    state,
    conversation_id: str,
) -> Dict[str, Any]:
    action = str(tool_input.get("action") or "").strip()
    if action not in {"metadata", "raw_retrieve"}:
        raise RuntimeError(f"Unsupported abs_data_tool action: {action}")

    dataset_id = str(tool_input.get("datasetId") or "").strip()

    approved_plan = (
        state.pending_plan.get("plan_context")
        if isinstance(state.pending_plan, dict)
        and str(state.pending_plan.get("status") or "") == "approved"
        and isinstance(state.pending_plan.get("plan_context"), dict)
        else {}
    )

    if action == "metadata":
        if not dataset_id:
            raise RuntimeError("abs_data_tool action metadata requires datasetId")
        logger.info(
            "ABS metadata request cid=%s datasetId=%s",
            conversation_id,
            dataset_id,
        )
        try:
            metadata = get_dataflow_metadata(dataset_id, force_refresh=False)
        except Exception as exc:
            logger.error(
                'ABS metadata error cid=%s datasetId=%s error="%s"',
                conversation_id,
                dataset_id,
                _truncate(exc, 500),
            )
            raise
        if not isinstance(metadata, dict):
            raise RuntimeError(f"Live ABS metadata for {dataset_id} was not an object")
        payload = _build_raw_metadata_payload(dataset_id, metadata)
        return _tool_result(
            _summarize_raw_metadata_payload(payload),
            {
                "kind": "raw_metadata",
                "dataset_id": payload.get("dataset_id"),
                "anchor_candidates": payload.get("anchor_candidates") or [],
            },
        )

    if not dataset_id:
        raise RuntimeError(f"abs_data_tool action {action} requires datasetId")

    if action == "raw_retrieve":
        data_key = str(tool_input.get("dataKey") or "").strip()
        is_custom_domestic = _is_custom_domestic_dataset(dataset_id)
        if not data_key and not is_custom_domestic:
            raise RuntimeError("abs_data_tool action raw_retrieve requires dataKey")
        if data_key and not is_custom_domestic:
            _validate_anchor_wildcard_data_key(dataset_id, data_key)
            anchor_token = _extract_anchor_token(data_key)
            metadata_anchor_codes = _metadata_anchor_codes_for_dataset(state, dataset_id)
            if metadata_anchor_codes and anchor_token and anchor_token not in metadata_anchor_codes:
                raise RuntimeError(
                    f"Invalid ABS anchor code for {dataset_id}. The anchor token '{anchor_token}' does not appear in the metadata anchor_codes list for that dataset."
                )
        run_dir = _ensure_runtime_dirs(conversation_id)
        start_period = str(tool_input.get("startPeriod") or "").strip() or None
        end_period = str(tool_input.get("endPeriod") or "").strip() or None
        detail = str(tool_input.get("detail") or "").strip() or "full"
        dimension_at_observation = (
            str(tool_input.get("dimensionAtObservation") or "").strip()
            or "TIME_PERIOD"
        )
        query_params = {
            "detail": detail,
            "dimensionAtObservation": dimension_at_observation,
        }
        if start_period:
            query_params["startPeriod"] = start_period
        if end_period:
            query_params["endPeriod"] = end_period
        request_descriptor = ""
        if is_custom_domestic:
            custom_flow = _custom_domestic_flow(dataset_id) or {}
            request_descriptor = str(custom_flow.get("sourceUrl") or custom_flow.get("sourcePageUrl") or "").strip()
            logger.info(
                "AUS custom raw_retrieve request cid=%s datasetId=%s dataKey=%s detail=%s source=%s flowType=%s",
                conversation_id,
                dataset_id,
                data_key or "all",
                detail,
                request_descriptor,
                str(custom_flow.get("flowType") or "").strip(),
            )
        else:
            request_descriptor = (
                f"{settings.abs_api_base.rstrip('/')}/rest/data/{dataset_id}/{data_key}"
                f"?{urlencode(query_params)}"
            )
            logger.info(
                "ABS raw_retrieve request cid=%s datasetId=%s dataKey=%s detail=%s dimensionAtObservation=%s startPeriod=%s endPeriod=%s url=%s",
                conversation_id,
                dataset_id,
                data_key,
                detail,
                dimension_at_observation,
                start_period,
                end_period,
                request_descriptor,
            )
        try:
            if is_custom_domestic:
                custom_flow = _custom_domestic_flow(dataset_id) or {}
                flow_type = str(custom_flow.get("flowType") or "").strip()
                if flow_type == "dcceew_aes_xlsx":
                    logger.info(
                        "AUS custom raw_retrieve python resolve cid=%s datasetId=%s dataKey=%s flowType=%s",
                        conversation_id,
                        dataset_id,
                        data_key or "all",
                        flow_type,
                    )
                    resolved_payload = _resolve_dcceew_dataset_python(
                        flow=custom_flow,
                        data_key=data_key,
                        detail=detail,
                    )
                else:
                    logger.info(
                        "AUS custom raw_retrieve direct bridge resolve cid=%s datasetId=%s dataKey=%s",
                        conversation_id,
                        dataset_id,
                        data_key or "all",
                    )
                    resolved_payload = resolve_dataset(
                        dataset_id=dataset_id,
                        data_key=data_key,
                        start_period=start_period,
                        end_period=end_period,
                        detail=detail,
                        dimension_at_observation=dimension_at_observation,
                    )
            else:
                resolved_payload = resolve_dataset(
                    dataset_id=dataset_id,
                    data_key=data_key,
                    start_period=start_period,
                    end_period=end_period,
                    detail=detail,
                    dimension_at_observation=dimension_at_observation,
                )
        except Exception as exc:
            logger.error(
                'AUS raw_retrieve error cid=%s datasetId=%s dataKey=%s request="%s" error="%s"',
                conversation_id,
                dataset_id,
                data_key,
                request_descriptor,
                _truncate(exc, 500),
            )
            status_code = _bridge_error_status_code(exc)
            if status_code == 404:
                raise RuntimeError("Australian domestic retrieval returned no data for that call.") from exc
            raise

        artifact_payload = {
            "artifact_type": "resolved_abs_dataset",
            "catalog_entry": {
                "datasetId": dataset_id,
                "title": dataset_id,
            },
            "retrieval": {
                "datasetId": dataset_id,
                "dataKey": data_key,
                "startPeriod": start_period,
                "endPeriod": end_period,
                "detail": detail,
                "dimensionAtObservation": dimension_at_observation,
                "rawDiscovery": True,
            },
            "resolved_dataset": resolved_payload,
            "source_references": _build_domestic_source_references(
                dataset_id,
                str(((resolved_payload.get("dataset") if isinstance(resolved_payload, dict) else {}) or {}).get("name") or dataset_id),
            ),
        }
        artifact_path = run_dir / "artifacts" / f"raw_retrieve_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, artifact_payload)
        source_references = artifact_payload.get("source_references") or []
        record = _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="abs_resolved_dataset",
            label=f"Domestic raw resolved {dataset_id}",
            summary=_truncate(
                f"Resolved domestic dataset {dataset_id} with {resolved_payload.get('observationCount', 'unknown')} observations.",
                300,
            ),
            source_references=source_references,
        )
        dataset_label = "custom domestic dataset" if is_custom_domestic else "raw ABS dataset"
        return _tool_result(
            (
                f"Retrieved {dataset_label} {dataset_id}.\n"
                f"Observation count: {resolved_payload.get('observationCount', 'unknown')}.\n"
                f"Series count: {len(resolved_payload.get('series') or []) if isinstance(resolved_payload, dict) else 0}.\n"
                f"Applied data key: {data_key or 'all'}.\n"
                f"Created artifact: {record['artifact_id']}. Use the sandbox tool to inspect or verify the wildcard retrieval."
            ),
            {
                "kind": "raw_retrieve",
                "dataset_id": dataset_id,
                "data_key": data_key,
                "observation_count": resolved_payload.get("observationCount") if isinstance(resolved_payload, dict) else None,
                "series_count": len(resolved_payload.get("series") or []) if isinstance(resolved_payload, dict) else None,
                "artifact_id": record["artifact_id"],
                "source_references": source_references,
            },
        )

    raise RuntimeError(f"Unsupported abs_data_tool action: {action}")


def _execute_aus_metadata_tool(
    *,
    tool_input: Dict[str, Any],
    state,
    conversation_id: str,
) -> Dict[str, Any]:
    delegated_input = dict(tool_input or {})
    delegated_input["action"] = "metadata"
    return _execute_abs_data_tool(
        tool_input=delegated_input,
        state=state,
        conversation_id=conversation_id,
    )


def _execute_aus_raw_retrieve_tool(
    *,
    tool_input: Dict[str, Any],
    state,
    conversation_id: str,
) -> Dict[str, Any]:
    delegated_input = dict(tool_input or {})
    delegated_input["action"] = "raw_retrieve"
    return _execute_abs_data_tool(
        tool_input=delegated_input,
        state=state,
        conversation_id=conversation_id,
    )


def _summarize_sandbox_result(result: Dict[str, Any], created_ids: List[str]) -> str:
    lines: List[str] = []
    stdout = str(result.get("stdout") or "").strip()
    if stdout:
        lines.append(f"stdout: {_truncate(stdout, 300)}")

    value = result.get("result")
    if isinstance(value, dict):
        rendered = _json_text(value)
        if len(rendered) <= 35000:
            lines.append(f"result_json:\n{rendered}")
        else:
            lines.append(f"result: object with keys {list(value.keys())[:20]}")
            lines.append(f"result_json_preview:\n{rendered[:12000]}")
    elif isinstance(value, list):
        rendered = _json_text(value)
        if len(rendered) <= 35000:
            lines.append(f"result_json:\n{rendered}")
        else:
            lines.append(f"result: list length {len(value)}")
            lines.append(f"result_json_preview:\n{rendered[:12000]}")
    elif value is not None:
        lines.append(f"result: {_truncate(value, 300)}")

    if created_ids:
        lines.append(f"created artifacts: {', '.join(created_ids)}")

    if not lines:
        lines.append("Sandbox completed without a structured result.")
    return "\n".join(lines)


def _create_sandbox_artifacts(
    *,
    state,
    conversation_id: str,
    runner_result: Dict[str, Any],
) -> List[str]:
    run_dir = _ensure_runtime_dirs(conversation_id)
    created_ids: List[str] = []

    raw_result = runner_result.get("result")
    if isinstance(raw_result, (dict, list)):
        path = run_dir / "artifacts" / f"sandbox_result_{len(state.artifacts) + 1:03d}.json"
        _write_json(path, raw_result)
        record = _make_artifact_record(
            state=state,
            path=path,
            kind="sandbox_result",
            label="Sandbox result",
            summary=_truncate(json.dumps(raw_result, ensure_ascii=False), 600),
        )
        created_ids.append(record["artifact_id"])

    for created in runner_result.get("created_artifacts") or []:
        if not isinstance(created, dict):
            continue
        path_value = created.get("path")
        if not isinstance(path_value, str):
            continue
        path = Path(path_value)
        if not path.exists():
            continue
        summary = f"Sandbox created file {path.name}."
        record = _make_artifact_record(
            state=state,
            path=path,
            kind=str(created.get("kind") or "sandbox_output"),
            label=f"Sandbox {created.get('name') or path.name}",
            summary=summary,
        )
        created_ids.append(record["artifact_id"])

    return created_ids


def _validate_sandbox_artifact_references(code: str, artifact_ids: List[str]) -> None:
    referenced = sorted(set(re.findall(r"\bartifact-\d{3}\b", str(code or ""))))
    if not referenced:
        return
    allowed = {str(item).strip() for item in artifact_ids if str(item).strip()}
    invalid = [artifact_id for artifact_id in referenced if artifact_id not in allowed]
    if invalid:
        raise RuntimeError(
            "sandbox_tool code referenced artifact ids that were not passed in artifact_ids: "
            + ", ".join(invalid)
        )


def _lint_sandbox_code(code: str) -> Optional[str]:
    source = str(code or "")

    if re.search(r"\bavailable_artifacts\b", source):
        return (
            "The sandbox step referenced available_artifacts as if it were a runtime variable. "
            "It is prompt context only. Use the passed artifact_ids plus load_artifact(...), "
            "get_series_rows(...), or inspect_artifact_schema(...)."
        )

    invented_artifact_helper = re.search(
        r"\b(create_artifact|save_artifact|make_artifact|create_json_artifact|create_text_artifact)\s*\(",
        source,
        re.IGNORECASE,
    )
    if invented_artifact_helper:
        return (
            "The sandbox step called an artifact helper that does not exist. "
            "Do not invent helper names. If you need to save output, use save_json(...) or save_text(...). "
            "Otherwise write normal Python."
        )

    positional_filter_rows = re.search(
        r"\bfilter_rows\s*\(\s*[^,\n\)]+\s*,\s*[^A-Za-z_\n\)][^,\n\)]*",
        source,
        re.IGNORECASE,
    )
    if positional_filter_rows:
        return (
            "The sandbox step called filter_rows(...) with unsupported positional arguments. "
            "Use filter_rows(rows, FIELD_code='X', OTHER_code='Y') with keyword equality filters only. "
            "If you need more complex filtering, write normal Python."
        )

    brittle_descending_sort = re.search(
        r"key\s*=\s*lambda\s+[^:]+:\s*-\s*[^\n,)]{3,}",
        source,
        re.IGNORECASE,
    )
    if brittle_descending_sort:
        return (
            "The sandbox step uses a handwritten descending lambda with unary minus. "
            "That pattern is brittle on ABS rows because fields may be nested or non-numeric. "
            "Use sort_by_numeric(...), top_n_by_numeric(...), or extract numbers with get_numeric()/coerce_number() first."
        )

    custom_uniq = re.search(r"^\s*def\s+uniq\s*\(", source, re.IGNORECASE | re.MULTILINE)
    if custom_uniq:
        return (
            "The sandbox step defines a custom uniq helper. "
            "Prefer distinct_values(...), group_rows(...), or index_rows(...) instead of handwritten dedupe logic."
        )

    raw_value_arithmetic = re.search(
        r"(\.get\(\s*['\"]value['\"]\s*\)|\[['\"]value['\"]\]|get_value\s*\([^)]+\)|get_numeric\s*\([^)]+\))\s*[-+]\s*"
        r"(\.get\(\s*['\"]value['\"]\s*\)|\[['\"]value['\"]\]|get_value\s*\([^)]+\)|get_numeric\s*\([^)]+\))",
        source,
        re.IGNORECASE,
    )
    if raw_value_arithmetic:
        return (
            "The sandbox step performs raw arithmetic on potentially nullable ABS values. "
            "Use safe_float()/get_value()/get_numeric() first and guard missing values before + or -."
        )

    return None


def _generate_sandbox_code(
    *,
    tool_input: Dict[str, Any],
    state,
    loop_payload: Dict[str, Any],
    usage_callback: Optional[Callable[[Dict[str, int]], None]] = None,
) -> str:
    code = str(tool_input.get("code") or "").strip()
    if code:
        return code

    sandbox_request = str(tool_input.get("sandbox_request") or "").strip()
    if not sandbox_request:
        raise RuntimeError("sandbox_tool requires either code or sandbox_request")

    artifact_ids = [
        str(item).strip()
        for item in (tool_input.get("artifact_ids") or [])
        if isinstance(item, str) and item.strip()
    ]
    sandbox_request = _normalize_sandbox_request_handoff(
        sandbox_request=sandbox_request,
        artifact_ids=artifact_ids,
    )
    tool_input["sandbox_request"] = sandbox_request

    logger.info(
        'Sandbox codegen request brief="%s" artifacts=%s',
        _truncate(sandbox_request, 280),
        artifact_ids,
    )

    try:
        generated = _call_model_text(
            build_sandbox_codegen_messages(
                payload=loop_payload,
                sandbox_request=sandbox_request,
                artifact_ids=artifact_ids,
            ),
            reasoning_effort="low",
            usage_callback=usage_callback,
        ).strip()
    except Exception as exc:
        logger.exception(
            'Sandbox codegen failed brief="%s" artifacts=%s error=%s',
            _truncate(sandbox_request, 280),
            artifact_ids,
            exc,
        )
        raise RuntimeError(
            "Sandbox code generation failed. "
            f"Brief: {_truncate(sandbox_request, 220)}. "
            f"Artifacts: {artifact_ids}. "
            f"Underlying error: {exc}"
        ) from exc
    if generated.startswith("```"):
        generated = re.sub(r"^```(?:python)?\s*", "", generated)
        generated = re.sub(r"\s*```$", "", generated)
        generated = generated.strip()
    if not generated:
        raise RuntimeError("Sandbox code generator returned empty code.")
    logger.info(
        'Sandbox codegen output preview="%s"',
        _truncate(generated, 280),
    )
    return generated


def _normalize_sandbox_request_handoff(*, sandbox_request: str, artifact_ids: List[str]) -> str:
    text = str(sandbox_request or "").strip()
    if not text:
        return text

    normalized = re.sub(r"\bavailable_artifacts\b", "selected artifact ids", text, flags=re.IGNORECASE)
    normalized = re.sub(r"\bartifacts?\b", "selected artifact ids", normalized)

    prefix = ""
    if artifact_ids:
        joined = ", ".join(artifact_ids[:4])
        prefix = (
            f"Use only the passed artifact_ids ({joined}). "
            "Work in this order: inspect first, narrow second, calculate third. "
        )

    return (prefix + normalized).strip()


def _execute_sandbox_tool(
    *,
    tool_input: Dict[str, Any],
    state,
    conversation_id: str,
    loop_payload: Dict[str, Any],
    status_callback: Optional[Callable[[str], None]] = None,
    usage_callback: Optional[Callable[[Dict[str, int]], None]] = None,
) -> Dict[str, Any]:
    status_callback = status_callback or (lambda _message: None)
    artifact_ids = tool_input.get("artifact_ids")
    code = _generate_sandbox_code(
        tool_input=tool_input,
        state=state,
        loop_payload=loop_payload,
        usage_callback=usage_callback,
    )
    tool_input["code"] = code
    if not isinstance(artifact_ids, list) or not all(isinstance(item, str) and item.strip() for item in artifact_ids):
        raise RuntimeError("sandbox_tool requires a non-empty artifact_ids array of strings")
    _validate_sandbox_artifact_references(code, artifact_ids)
    lint_error = _lint_sandbox_code(code)
    if lint_error:
        raise RuntimeError(lint_error)
    try:
        compile(code, "<sandbox>", "exec")
    except SyntaxError as exc:
        raise RuntimeError(
            "Sandbox code generation produced invalid Python. "
            f"{exc.__class__.__name__}: {exc}"
        ) from exc

    allowed_artifacts = {
        artifact["artifact_id"]: artifact
        for artifact in state.artifacts
        if isinstance(artifact, dict) and artifact.get("artifact_id") in artifact_ids
    }
    if len(allowed_artifacts) != len(artifact_ids):
        raise RuntimeError("sandbox_tool requested unknown artifact ids")

    run_dir = _ensure_runtime_dirs(conversation_id)
    payload_path = run_dir / "sandbox" / f"payload_{len(state.loop_history) + 1:03d}.json"
    output_dir = run_dir / "sandbox" / f"output_{len(state.loop_history) + 1:03d}"
    payload = {
        "code": code,
        "output_dir": str(output_dir),
        "artifacts": allowed_artifacts,
    }
    _write_json(payload_path, payload)

    runner_path = Path(__file__).resolve().parent / "harness" / "sandbox_runner.py"
    completed = subprocess.run(
        [settings.python_binary, str(runner_path), str(payload_path)],
        capture_output=True,
        text=True,
        cwd=str(run_dir),
        timeout=90,
    )
    if completed.returncode != 0:
        raise RuntimeError(
            "Sandbox execution failed:\n"
            f"STDOUT: {completed.stdout[:1000]}\nSTDERR: {completed.stderr[:1000]}"
        )
    try:
        runner_result = json.loads(completed.stdout.strip())
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"Sandbox returned invalid JSON: {completed.stdout[:1000]}"
        ) from exc

    if runner_result.get("error"):
        raise RuntimeError(f"Sandbox error:\n{runner_result['error']}")

    created_ids = _create_sandbox_artifacts(
        state=state,
        conversation_id=conversation_id,
        runner_result=runner_result,
    )
    return _tool_result(
        _summarize_sandbox_result(runner_result, created_ids),
        {
            "kind": "sandbox",
            "result": runner_result.get("result"),
            "stdout": str(runner_result.get("stdout") or "").strip()[:4000],
            "created_artifact_ids": created_ids,
            "generated_code_preview": code[:600],
        },
    )


def _build_loop_handoff_summary(
    *,
    step: Dict[str, Any],
    progress_note: str,
    result_summary: str,
    result_data: Any = None,
) -> str:
    step_id = str(step.get("id") or "").strip()
    attempt = str(step.get("summary") or progress_note or "").strip()
    if not attempt:
        attempt = f"Ran step {step_id or 'unknown'}."

    result_dict = result_data if isinstance(result_data, dict) else {}
    kind = str(result_dict.get("kind") or "").strip()

    outcome = "Completed."
    known = str(result_summary or "").strip()
    remaining = ""

    if kind == "tool_error":
        error_text = str(result_dict.get("error") or "").strip()
        retry_guidance = str(result_dict.get("retry_guidance") or "").strip()
        outcome = "Failed."
        known = error_text or "The step did not complete successfully."
        remaining = retry_guidance or "Choose a safer next step using the failure details."
    elif kind in {"retrieve", "raw_retrieve"}:
        dataset_id = str(result_dict.get("dataset_id") or "").strip()
        artifact_id = str(result_dict.get("artifact_id") or "").strip()
        observation_count = result_dict.get("observation_count")
        series_count = result_dict.get("series_count")
        outcome_bits = [f"Retrieved {dataset_id}." if dataset_id else "Retrieved the dataset."]
        if observation_count is not None:
            outcome_bits.append(f"Observations: {observation_count}.")
        if series_count is not None:
            outcome_bits.append(f"Series: {series_count}.")
        outcome = " ".join(outcome_bits)
        known_bits = []
        if artifact_id:
            known_bits.append(f"Artifact available: {artifact_id}.")
        source_refs = result_dict.get("source_references") if isinstance(result_dict.get("source_references"), list) else []
        if source_refs:
            first_ref = source_refs[0] if isinstance(source_refs[0], dict) else {}
            ref_provider = str(first_ref.get("provider") or "").strip()
            ref_dataset = str(first_ref.get("dataset_id") or "").strip()
            if ref_provider or ref_dataset:
                known_bits.append(f"Source: {' '.join(item for item in [ref_provider, ref_dataset] if item)}.")
        known = " ".join(known_bits) or known
        remaining = "Inspect or narrow the retrieved artifact before further analysis."
    elif kind == "raw_metadata":
        dataset_id = str(result_dict.get("dataset_id") or "").strip()
        anchor_candidates = result_dict.get("anchor_candidates") if isinstance(result_dict.get("anchor_candidates"), list) else []
        outcome = f"Inspected metadata for {dataset_id or 'the dataset'}."
        known_bits: List[str] = []
        if anchor_candidates:
            candidate_labels = []
            for item in anchor_candidates[:3]:
                if not isinstance(item, dict):
                    continue
                anchor_type = str(item.get("anchor_type") or "").strip()
                anchor_description = str(item.get("anchor_description") or "").strip()
                if anchor_type and anchor_description:
                    candidate_labels.append(f"{anchor_type} ({anchor_description})")
                elif anchor_type:
                    candidate_labels.append(anchor_type)
            if candidate_labels:
                known_bits.append(f"Anchor candidates: {', '.join(candidate_labels)}.")
        known = " ".join(known_bits) or known
        remaining = "Choose one anchor and build the exact wildcard data key from the metadata."
    elif kind == "macro_query":
        artifact_id = str(result_dict.get("artifact_id") or "").strip()
        series_count = result_dict.get("series_count")
        source_refs = result_dict.get("source_references") if isinstance(result_dict.get("source_references"), list) else []
        provider = str(result_dict.get("provider") or "").strip()
        concept_label = str(result_dict.get("concept_label") or "").strip()
        outcome_bits = ["Queried macro providers and received upstream data."]
        if provider:
            outcome_bits.append(f"Provider: {provider}.")
        if concept_label:
            outcome_bits.append(f"Concept: {concept_label}.")
        if series_count is not None:
            outcome_bits.append(f"Series: {series_count}.")
        outcome = " ".join(outcome_bits)
        known_bits = []
        if artifact_id:
            known_bits.append(f"Artifact available: {artifact_id}.")
        if source_refs:
            preview = []
            for item in source_refs[:3]:
                if not isinstance(item, dict):
                    continue
                label = ", ".join(
                    part
                    for part in [
                        str(item.get("provider") or "").strip(),
                        str(item.get("indicator") or "").strip(),
                        str(item.get("series_id") or "").strip(),
                    ]
                    if part
                )
                if label:
                    preview.append(label)
            if preview:
                known_bits.append(f"Upstream sources: {'; '.join(preview)}.")
        known = " ".join(known_bits) or known
        remaining = "Inspect or narrow the retrieved artifact before comparing, ranking, or answering."
    elif kind == "macro_metadata":
        candidate_id = str(result_dict.get("candidate_id") or "").strip()
        metadata = result_dict.get("metadata") if isinstance(result_dict.get("metadata"), dict) else {}
        counts = metadata.get("fullDimensionCounts") if isinstance(metadata.get("fullDimensionCounts"), dict) else {}
        outcome = f"Inspected macro metadata for {candidate_id or 'the selected candidate'}."
        count_bits = []
        for key in ["reporters", "partners", "hs_4digit"]:
            value = counts.get(key)
            if isinstance(value, int):
                count_bits.append(f"{key}={value}")
        known = f"Available dimension counts: {', '.join(count_bits)}." if count_bits else known
        remaining = "Choose exact reporter, partner, flow, frequency, and HS codes from metadata before retrieval."
    elif kind == "sandbox":
        created_ids = result_dict.get("created_artifact_ids") if isinstance(result_dict.get("created_artifact_ids"), list) else []
        sandbox_result = result_dict.get("result")
        outcome = "Sandbox step completed."
        if isinstance(sandbox_result, dict):
            keys = [str(key) for key in list(sandbox_result.keys())[:8]]
            known_bits = [f"Result keys: {', '.join(keys)}."] if keys else []
            if created_ids:
                known_bits.append(f"Created artifacts: {', '.join(str(item) for item in created_ids[:4])}.")
            known = " ".join(known_bits) or known
        elif isinstance(sandbox_result, list):
            known = f"Result list length: {len(sandbox_result)}."
            if created_ids:
                known += f" Created artifacts: {', '.join(str(item) for item in created_ids[:4])}."
        elif created_ids:
            known = f"Created artifacts: {', '.join(str(item) for item in created_ids[:4])}."
        remaining = "Use this result to decide whether the next step is more narrowing or the final calculation."
    elif kind in {"web_search", "web_page"}:
        outcome = "Gathered supporting web context."
        remaining = "Use it only as supporting context, with ABS evidence primary."
    elif kind in {"model_call_error", "harness_parse_error"}:
        outcome = "The loop failed before a valid decision completed."
        remaining = "Retry with a valid loop decision and stricter output formatting."

    sections = [
        f"Tried: {_truncate(attempt, 240)}",
        f"Outcome: {_truncate(outcome, 320)}",
    ]
    if known:
        sections.append(f"Known now: {_truncate(known, 480)}")
    if remaining:
        sections.append(f"Still unresolved: {_truncate(remaining, 320)}")
    return "\n".join(sections)


def _record_loop_feedback(
    state,
    *,
    step: Dict[str, Any],
    progress_note: str,
    result_summary: str,
    result_data: Any = None,
) -> None:
    handoff_summary = _build_loop_handoff_summary(
        step=step,
        progress_note=progress_note,
        result_summary=result_summary,
        result_data=result_data,
    )
    entry = {
        "step": step,
        "progress_note": progress_note,
        "handoff_summary": handoff_summary,
        "result_summary": str(result_summary or "").strip()[:12000],
    }
    if result_data is not None:
        entry["result_data"] = result_data
    state.loop_history.append(entry)


def _normalize_sandbox_code(code: Any) -> str:
    if not isinstance(code, str):
        return ""
    normalized_lines = []
    for raw_line in code.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        normalized_lines.append(line)
    return re.sub(r"\s+", " ", "\n".join(normalized_lines)).strip()


def _recent_failed_sandbox_entry(state) -> Optional[Dict[str, Any]]:
    for item in reversed(state.loop_history):
        if not isinstance(item, dict):
            continue
        step = item.get("step") if isinstance(item.get("step"), dict) else {}
        if str(step.get("id") or "").strip() != "sandbox_tool":
            continue
        result_data = item.get("result_data") if isinstance(item.get("result_data"), dict) else {}
        if str(result_data.get("kind") or "").strip() != "tool_error":
            continue
        if str(result_data.get("tool_step_id") or "").strip() != "sandbox_tool":
            continue
        return item
    return None


def _classify_tool_failure(step_id: str, error_text: str, tool_input: Dict[str, Any]) -> Dict[str, Any]:
    clean_error = str(error_text or "").strip()
    tool_name = {
        "provider_route_tool": "provider_route_tool",
        "aus_metadata_tool": "aus_metadata_tool",
        "aus_raw_retrieve_tool": "aus_raw_retrieve_tool",
        "macro_data_tool": "macro_data_tool",
        "web_search_tool": "web_search_tool",
        "sandbox_tool": "sandbox_tool",
    }.get(step_id, "")

    result = {
        "kind": "tool_error",
        "tool_step_id": step_id,
        "tool_name": tool_name,
        "error": clean_error[:4000],
    }

    if step_id != "sandbox_tool":
        lowered = clean_error.lower()
        if "exactly one anchored segment" in lowered or "one anchor code only" in lowered:
            result["error_class"] = "invalid_raw_anchor_wildcard_key"
            result["retry_guidance"] = (
                "For raw_retrieve, rebuild dataKey as exactly one anchored segment plus wildcard dots in every other position. "
                "Do not fix geography, frequency, adjustment, industry, or any extra segment in the key."
            )
        else:
            result["retry_guidance"] = "Choose a different valid step or adjust the tool input to address the reported error."
        return result

    sandbox_code = str(tool_input.get("code") or "")
    sandbox_request = str(tool_input.get("sandbox_request") or "")
    result["sandbox_code_preview"] = sandbox_code[:600]
    result["sandbox_code_normalized"] = _normalize_sandbox_code(sandbox_code)[:2000]
    result["sandbox_request"] = sandbox_request[:1000]
    result["artifact_ids"] = [
        str(item).strip()
        for item in (tool_input.get("artifact_ids") or [])
        if isinstance(item, str) and item.strip()
    ]

    lowered = clean_error.lower()
    if "nonetype' object is not subscriptable" in lowered:
        result["error_class"] = "missing_match_or_null_lookup"
        result["retry_guidance"] = (
            "A sandbox lookup returned None and the code indexed into it. "
            "Do not reuse the same lookup chain. Inspect rows first, then use find_row/safe_get or require_row/require_fields."
        )
    elif "no row matched:" in lowered:
        result["error_class"] = "missing_row"
        result["retry_guidance"] = (
            "The requested row was not present. Inspect available rows and codes before retrying, "
            "or retrieve data at a compatible level instead of assuming the join target exists."
        )
    elif "missing required fields:" in lowered:
        result["error_class"] = "missing_fields"
        result["retry_guidance"] = (
            "The target row exists but required fields were absent or null. "
            "Inspect schema/rows first and adjust the calculation to fields that are actually populated."
        )
    elif "numeric_change requires two non-null numeric values" in lowered:
        result["error_class"] = "missing_numeric_inputs"
        result["retry_guidance"] = (
            "The calculation attempted to use null numeric values. "
            "Verify the numerator and denominator exist before computing the metric."
        )
    elif "timed out" in lowered or "timeout" in lowered:
        result["error_class"] = "sandbox_timeout"
        result["retry_guidance"] = (
            "The sandbox step was too heavy. Retry with a smaller inspection step or simplify the transformation before joining artifacts."
        )
    elif "handwritten descending lambda with unary minus" in lowered:
        result["error_class"] = "sandbox_brittle_sort"
        result["retry_guidance"] = (
            "Do not hand-write descending sort lambdas over ABS rows. "
            "Use sort_by_numeric/top_n_by_numeric or inspect and extract numeric fields first."
        )
    elif "defines a custom uniq helper" in lowered:
        result["error_class"] = "sandbox_custom_dedupe"
        result["retry_guidance"] = (
            "Do not write custom uniq/dedupe helpers. "
            "Use distinct_values, group_rows, or index_rows to inspect or deduplicate rows."
        )
    elif "raw arithmetic on potentially nullable abs values" in lowered:
        result["error_class"] = "sandbox_null_arithmetic"
        result["retry_guidance"] = (
            "Do not add or subtract raw ABS values directly. "
            "Use get_value/get_numeric/safe_float first and guard missing values before arithmetic."
        )
    else:
        result["error_class"] = "sandbox_generic"
        result["retry_guidance"] = (
            "Do not repeat the same sandbox code. Inspect the artifact shape again and choose a narrower or safer analysis step."
        )
    return result


def _validate_anchor_wildcard_data_key(dataset_id: str, data_key: str) -> None:
    segments = str(data_key or "").split(".")
    fixed_segments: List[tuple[int, str]] = []

    for index, segment in enumerate(segments):
        token = str(segment or "").strip()
        if not token:
            continue
        fixed_segments.append((index, token))

    if len(fixed_segments) != 1:
        raise RuntimeError(
            "Invalid raw ABS dataKey. raw_retrieve must use exactly one anchored segment and wildcard every other segment. "
            f"Received datasetId={dataset_id}, dataKey={data_key}. "
            "Rebuild the key as one anchor code in the correct position with dots for all remaining positions."
        )

    _, anchor_token = fixed_segments[0]
    if "+" in anchor_token:
        raise RuntimeError(
            "Invalid raw ABS dataKey. raw_retrieve must use exactly one anchor code, not multiple codes in one segment. "
            f"Received datasetId={dataset_id}, dataKey={data_key}. "
            "Choose one anchor code only and wildcard every other position."
        )


def _extract_anchor_token(data_key: str) -> str:
    for segment in str(data_key or "").split("."):
        token = str(segment or "").strip()
        if token:
            return token
    return ""


def _metadata_anchor_codes_for_dataset(state, dataset_id: str) -> set[str]:
    clean_dataset_id = str(dataset_id or "").strip()
    if not clean_dataset_id:
        return set()
    for item in reversed(state.loop_history):
        if not isinstance(item, dict):
            continue
        result_data = item.get("result_data") if isinstance(item.get("result_data"), dict) else {}
        if str(result_data.get("kind") or "").strip() != "raw_metadata":
            continue
        if str(result_data.get("dataset_id") or "").strip() != clean_dataset_id:
            continue
        codes: set[str] = set()
        for candidate in result_data.get("anchor_candidates") or []:
            if not isinstance(candidate, dict):
                continue
            for code_item in candidate.get("anchor_codes") or []:
                if not isinstance(code_item, dict):
                    continue
                code = str(code_item.get("code") or "").strip()
                if code:
                    codes.add(code)
        return codes
    return set()


def _last_provider_route_context(state) -> Dict[str, str]:
    stored_route = _normalize_provider_route(getattr(state, "last_provider_route", ""))
    stored_query = str(getattr(state, "last_provider_search_query", "") or "").strip()
    if stored_route:
        payload: Dict[str, str] = {"selected_route": stored_route}
        if stored_query:
            payload["selected_search_query"] = stored_query
        return payload
    for item in reversed(state.loop_history):
        if not isinstance(item, dict):
            continue
        result_data = item.get("result_data") if isinstance(item.get("result_data"), dict) else {}
        if str(result_data.get("kind") or "").strip() != "provider_route_selection":
            continue
        route = _normalize_provider_route(result_data.get("route"))
        search_query = str(result_data.get("search_query") or "").strip()
        reason = str(result_data.get("reason") or "").strip()
        if route:
            payload: Dict[str, str] = {"selected_route": route}
            if search_query:
                payload["selected_search_query"] = search_query
            if reason:
                payload["routing_reason"] = reason
            return payload
    return {}


def _sandbox_retry_conflict(state, tool_input: Dict[str, Any]) -> Optional[str]:
    recent_failure = _recent_failed_sandbox_entry(state)
    if recent_failure is None:
        return None

    prior_data = recent_failure.get("result_data") if isinstance(recent_failure.get("result_data"), dict) else {}
    prior_normalized = str(prior_data.get("sandbox_code_normalized") or "").strip()
    current_normalized = _normalize_sandbox_code(tool_input.get("code"))
    prior_request = str(prior_data.get("sandbox_request") or "").strip()
    current_request = str(tool_input.get("sandbox_request") or "").strip()

    same_attempt = False
    if prior_normalized and current_normalized and prior_normalized == current_normalized:
        same_attempt = True
    elif prior_request and current_request and prior_request == current_request:
        same_attempt = True

    if same_attempt:
        guidance = str(prior_data.get("retry_guidance") or "").strip()
        return (
            "The proposed sandbox step is materially the same as the most recent failed sandbox attempt. "
            "Choose a different approach instead of retrying identical code. "
            f"Previous guidance: {guidance}"
        ).strip()

    return None


def _count_recent_recovery_failures(state) -> int:
    count = 0
    for item in reversed(state.loop_history):
        if not isinstance(item, dict):
            break
        result_data = item.get("result_data") if isinstance(item.get("result_data"), dict) else {}
        kind = str(result_data.get("kind") or "").strip()
        if kind not in {"model_call_error", "harness_parse_error"}:
            break
        count += 1
    return count


def _count_recent_harness_parse_failures(state) -> int:
    count = 0
    for item in reversed(state.loop_history):
        if not isinstance(item, dict):
            break
        result_data = item.get("result_data") if isinstance(item.get("result_data"), dict) else {}
        if str(result_data.get("kind") or "").strip() != "harness_parse_error":
            break
        count += 1
    return count


def _retry_reasoning_effort(state) -> Optional[str]:
    return None


def _recent_completed_runs(state, keep_runs: int = PRIOR_RUN_WINDOW) -> List[Dict[str, int]]:
    runs = getattr(state, "completed_runs", [])
    if not isinstance(runs, list):
        return []
    normalized: List[Dict[str, int]] = []
    for item in runs:
        if not isinstance(item, dict):
            continue
        message_count = item.get("message_count")
        loop_count = item.get("loop_count")
        artifact_count = item.get("artifact_count")
        if not all(isinstance(value, int) and value >= 0 for value in [message_count, loop_count, artifact_count]):
            continue
        normalized.append(
            {
                "message_count": int(message_count),
                "loop_count": int(loop_count),
                "artifact_count": int(artifact_count),
            }
        )
    return normalized[-keep_runs:] if keep_runs > 0 else []


def _slice_from_recent_runs(items: List[Any], completed_runs: List[Dict[str, int]], count_key: str) -> List[Any]:
    if not isinstance(items, list):
        return []
    if len(completed_runs) <= PRIOR_RUN_WINDOW:
        return list(items)
    start_index = int(completed_runs[-(PRIOR_RUN_WINDOW + 1)].get(count_key) or 0)
    start_index = max(0, min(start_index, len(items)))
    return items[start_index:]


def _build_recent_chat_history_payload(state) -> List[Dict[str, str]]:
    completed_runs = _recent_completed_runs(state, keep_runs=PRIOR_RUN_WINDOW + 1)
    recent_messages = _slice_from_recent_runs(state.messages, completed_runs, "message_count")
    return build_chat_history_payload(recent_messages, recent_full_limit=8, older_compact_limit=4)


def _payload_loop_history(state, run_loop_start_index: int) -> tuple[List[Dict[str, Any]], int]:
    completed_runs = _recent_completed_runs(state, keep_runs=PRIOR_RUN_WINDOW + 1)
    prior = _slice_from_recent_runs(state.loop_history[:run_loop_start_index], completed_runs, "loop_count")
    current = state.loop_history[run_loop_start_index:]
    prior_payload = compact_loop_history(prior, limit=4) if prior else []
    current_payload = compact_loop_history(current, limit=max(len(current), 1)) if current else []
    return prior_payload + current_payload, len(current_payload)


def _payload_artifacts(state, run_artifact_start_index: int) -> tuple[List[Dict[str, Any]], int]:
    completed_runs = _recent_completed_runs(state, keep_runs=PRIOR_RUN_WINDOW + 1)
    prior = _slice_from_recent_runs(state.artifacts[:run_artifact_start_index], completed_runs, "artifact_count")
    current = state.artifacts[run_artifact_start_index:]
    prior_payload = compact_artifacts(prior, limit=4) if prior else []
    current_payload = compact_artifacts(current, limit=max(len(current), 1)) if current else []
    return prior_payload + current_payload, len(current_payload)


def generate_response(
    conversation_id: str,
    user_content: str,
    store: ConversationStore,
    status_callback: Optional[Callable[[str], None]] = None,
) -> str:
    cancel_event = _acquire_cancellation_event(conversation_id)
    status_callback = status_callback or (lambda _message: None)

    try:
        state = store.load(conversation_id)
        _ensure_runtime_dirs(conversation_id)
        saved_progress_messages: list[str] = []

        def emit_status(message: str) -> None:
            normalized = str(message or "").strip()
            if not normalized:
                return
            status_callback(normalized)
            if not saved_progress_messages or saved_progress_messages[-1] != normalized:
                saved_progress_messages.append(normalized)

        run_input_tokens = 0
        run_output_tokens = 0
        run_cached_input_tokens = 0

        def record_usage(usage: Dict[str, int]) -> None:
            nonlocal run_input_tokens, run_output_tokens, run_cached_input_tokens
            if not isinstance(usage, dict):
                return
            run_input_tokens += _safe_int(usage.get("input_tokens"))
            run_output_tokens += _safe_int(usage.get("output_tokens"))
            run_cached_input_tokens += _safe_int(usage.get("cached_input_tokens"))

        def persist_completed_turn(assistant_content: str) -> None:
            state.messages.append({"role": "user", "content": user_content})
            for progress_message in saved_progress_messages:
                state.messages.append({"role": "progress", "content": progress_message})
            state.messages.append(
                {
                    "role": "assistant",
                    "content": assistant_content,
                    "run_cost": _build_run_cost_payload(
                        input_tokens=run_input_tokens,
                        cached_input_tokens=run_cached_input_tokens,
                        output_tokens=run_output_tokens,
                    ),
                }
            )
            state.completed_runs.append(
                {
                    "message_count": len(state.messages),
                    "loop_count": len(state.loop_history),
                    "artifact_count": len(state.artifacts),
                }
            )
            state.active_run_message_count = len(state.messages)
            state.active_run_loop_count = len(state.loop_history)
            state.active_run_artifact_count = len(state.artifacts)
            store.save(state)

        active_user_message = user_content
        payload_chat_history = _build_recent_chat_history_payload(state)
        pre_run_dataset_shortlist: List[Dict[str, Any]] = []
        pre_run_provider_route: Dict[str, Any] = {}
        clarification_followup = False
        pending_plan = state.pending_plan if isinstance(state.pending_plan, dict) else None
        if pending_plan and str(pending_plan.get("status") or "") == "awaiting_approval":
            plan_context = _normalize_plan_context(
                pending_plan.get("plan_context"),
                fallback_question=user_content,
            )
            if bool(plan_context.get("await_user_input")):
                clarification_followup = True
                state.pending_plan = None
                active_user_message = user_content
            else:
                plan_reply = _detect_plan_reply(user_content)
                if plan_reply == "approve":
                    plan_context["allow_raw_discovery"] = True
                    state.pending_plan = {
                        "status": "approved",
                        "plan_markdown": str(pending_plan.get("plan_markdown") or "").strip(),
                        "plan_context": plan_context,
                    }
                    active_user_message = user_content
                else:
                    state.pending_plan = None

        if _should_reset_after_user_correction(state, user_content):
            active_user_message = _reset_context_after_user_correction(state, user_content)
            payload_chat_history = _build_recent_chat_history_payload(state)
            state.pending_plan = None
        state.current_abs_dataset_shortlist = []
        state.current_macro_indicator_shortlist = []
        store.save(state)

        route_hint = _detect_provider_route(active_user_message)
        pre_run_provider_route = dict(route_hint or {})
        logger.info(
            'Pre-run provider hint cid=%s route=%s preferred_tool=%s reason="%s" query="%s"',
            conversation_id,
            pre_run_provider_route.get("provider_route") or "",
            pre_run_provider_route.get("preferred_tool") or "",
            pre_run_provider_route.get("routing_reason") or "",
            _truncate(active_user_message, 220),
        )
        pre_run_dataset_shortlist: List[Dict[str, Any]] = []
        pre_run_macro_indicator_shortlist: List[Dict[str, Any]] = []
        selected_provider_route = ""
        selected_route_query = ""
        carried_provider_route = _last_provider_route_context(state)
        if carried_provider_route:
            pre_run_provider_route.update(
                {key: value for key, value in carried_provider_route.items() if value}
            )
            selected_provider_route = _normalize_provider_route(
                carried_provider_route.get("selected_route")
            ) or ""
            selected_route_query = str(
                carried_provider_route.get("selected_search_query") or ""
            ).strip()
        if clarification_followup and not selected_provider_route:
            hinted_route = _normalize_provider_route(pre_run_provider_route.get("provider_route"))
            if hinted_route:
                selected_provider_route = hinted_route
                selected_route_query = active_user_message
                pre_run_provider_route["selected_route"] = hinted_route
                pre_run_provider_route["selected_search_query"] = active_user_message

        run_loop_start_index = len(state.loop_history)
        run_artifact_start_index = len(state.artifacts)

        for loop_index in range(1, settings.max_loops + 1):
            _ensure_not_cancelled(conversation_id, cancel_event, f"loop_{loop_index}_start")

            pending_mode = str(getattr(state, "pending_user_mode", "") or "").strip().lower()
            pending_message = str(getattr(state, "pending_user_message", "") or "").strip()
            if pending_mode == "steer" and pending_message:
                emit_status("User steer received. Adapting to the new request.")
                logger.info(
                    'Loop steer inject cid=%s loop=%s message="%s"',
                    conversation_id,
                    loop_index,
                    _truncate(pending_message, 220),
                )
                active_user_message = pending_message
                payload_chat_history = _build_recent_chat_history_payload(state)
                state.pending_user_message = ""
                state.pending_user_mode = ""
                state.pending_plan = None
                pre_run_dataset_shortlist = []
                pre_run_macro_indicator_shortlist = []
                state.current_abs_dataset_shortlist = []
                state.current_macro_indicator_shortlist = []
                selected_provider_route = ""
                selected_route_query = ""
                route_hint = _detect_provider_route(active_user_message)
                pre_run_provider_route = dict(route_hint or {})
                store.save(state)

            if not pre_run_dataset_shortlist and selected_provider_route == "aus":
                shortlist_query = selected_route_query or active_user_message
                pre_run_dataset_shortlist = _refresh_abs_shortlist(
                    state=state,
                    conversation_id=conversation_id,
                    route=selected_provider_route,
                    shortlist_query=shortlist_query,
                )
            if not pre_run_macro_indicator_shortlist and selected_provider_route == "macro":
                shortlist_query = selected_route_query or active_user_message
                logger.info(
                    'Post-route macro shortlist start cid=%s route=%s query="%s"',
                    conversation_id,
                    selected_provider_route,
                    _truncate(shortlist_query, 220),
                )
                macro_shortlist_payload = build_macro_shortlist(shortlist_query, limit=40)
                macro_candidates = (
                    macro_shortlist_payload.get("candidates")
                    if isinstance(macro_shortlist_payload, dict)
                    else []
                )
                pre_run_macro_indicator_shortlist = [
                    item for item in _to_list(macro_candidates) if isinstance(item, dict)
                ]
                state.current_macro_indicator_shortlist = list(pre_run_macro_indicator_shortlist)
                shortlist_record = _persist_shortlist_artifact(
                    state=state,
                    conversation_id=conversation_id,
                    kind="macro_indicator_shortlist",
                    query=shortlist_query,
                    candidates=pre_run_macro_indicator_shortlist,
                )
                logger.info(
                    "Post-route macro shortlist ready cid=%s route=%s count=%s artifact=%s top=%s",
                    conversation_id,
                    selected_provider_route,
                    len(pre_run_macro_indicator_shortlist),
                    shortlist_record["artifact_id"],
                    [str(item.get("candidate_id") or "").strip() for item in pre_run_macro_indicator_shortlist[:3]],
                )

            payload_loop_history, protected_loop_history_count = _payload_loop_history(
                state,
                run_loop_start_index,
            )
            payload_artifacts, protected_artifact_count = _payload_artifacts(
                state,
                run_artifact_start_index,
            )

            current_provider_route_payload = dict(pre_run_provider_route or {})
            if selected_provider_route:
                current_provider_route_payload["selected_route"] = selected_provider_route
            if selected_route_query:
                current_provider_route_payload["selected_search_query"] = selected_route_query

            payload = build_loop_payload(
                user_message=active_user_message,
                chat_history=payload_chat_history,
                loop_history=payload_loop_history,
                artifacts=payload_artifacts,
                plan_state=_build_plan_state(state, user_message=active_user_message),
                pre_run_provider_route=current_provider_route_payload,
                pre_run_dataset_shortlist=pre_run_dataset_shortlist,
                pre_run_macro_indicator_shortlist=pre_run_macro_indicator_shortlist,
                loop_index=loop_index,
                max_loops=settings.max_loops,
                protected_loop_history_count=protected_loop_history_count,
                protected_artifact_count=protected_artifact_count,
            )
            try:
                raw_model_response = _call_model(
                    build_model_messages(payload),
                    reasoning_effort=_retry_reasoning_effort(state),
                    usage_callback=record_usage,
                )
            except Exception as exc:
                logger.exception(
                    "Model call failed cid=%s loop=%s error=%s",
                    conversation_id,
                    loop_index,
                    exc,
                )
                _record_loop_feedback(
                    state,
                    step={
                        "id": "model_call_error",
                        "summary": "Model API call failed before a loop decision was returned",
                    },
                    progress_note="Recovering from a model call failure.",
                    result_summary=(
                        "The model API call failed before the harness received a valid loop decision.\n"
                        f"Error: {str(exc)}"
                    ),
                    result_data={
                        "kind": "model_call_error",
                        "error": str(exc),
                    },
                )
                store.save(state)
                if _count_recent_recovery_failures(state) >= MAX_CONSECUTIVE_RECOVERY_FAILURES:
                    raise RuntimeError(
                        "The harness hit repeated model-call failures and stopped after 3 recovery attempts. "
                        "Please retry or ask a narrower follow-up."
                    )
                continue
            try:
                parsed = parse_harness_loop_output(raw_model_response)
            except HarnessParserError as exc:
                diagnostics = exc.diagnostics if isinstance(getattr(exc, "diagnostics", None), dict) else {}
                logger.warning(
                    "Harness parse failed cid=%s loop=%s error=%s failure_class=%s candidate_count=%s validation_path=%s top_level_keys=%s truncated_suspected=%s raw_len=%s raw=%s",
                    conversation_id,
                    loop_index,
                    str(exc),
                    diagnostics.get("failure_class"),
                    diagnostics.get("candidate_count"),
                    diagnostics.get("validation_path"),
                    diagnostics.get("top_level_keys_detected") or diagnostics.get("normalized_top_level_keys"),
                    diagnostics.get("truncated_suspected"),
                    diagnostics.get("raw_length"),
                    _truncate(raw_model_response, 600),
                )
                next_parse_retry = _count_recent_harness_parse_failures(state) + 1
                if next_parse_retry == 1:
                    progress_text = "That step failed output validation. Retrying with stricter formatting."
                elif next_parse_retry == 2:
                    progress_text = "That step failed output validation again. Retrying with ultra-strict formatting."
                else:
                    progress_text = "That step failed output validation repeatedly."

                _record_loop_feedback(
                    state,
                    step={
                        "id": "invalid_model_output",
                        "summary": "Model returned malformed harness JSON",
                    },
                    progress_note=progress_text,
                    result_summary=(
                        "The model returned malformed harness JSON.\n"
                        f"Parse error: {str(exc)}\n"
                        "Correction required: stay with the intended loop decision, but return one literal top-level JSON object with "
                        "`step`, `progress_note`, and `model_output` only. "
                        "Do not include prose, markdown fences, quoted JSON, escaped JSON, or nested wrapper objects.\n"
                        f"Raw output preview: {_truncate(raw_model_response, 600)}"
                    ),
                    result_data={
                        "kind": "harness_parse_error",
                        "parse_error": str(exc),
                        "parse_diagnostics": diagnostics,
                        "retry_stage": next_parse_retry,
                        "raw_output_preview": _truncate(raw_model_response, 600),
                        "correction_instructions": (
                            "Return one literal top-level JSON object only. "
                            "Required keys: step, progress_note, model_output. "
                            "For tool steps, model_output must contain tool_name and tool_input as an object. "
                            "Do not rethink the task; fix the output shape."
                        ),
                    },
                )
                store.save(state)
                if _count_recent_recovery_failures(state) >= MAX_CONSECUTIVE_RECOVERY_FAILURES:
                    raise RuntimeError(
                        "The harness hit repeated malformed model outputs and stopped after 3 recovery attempts. "
                        "Please retry or ask a narrower follow-up."
                    )
                continue

            step = parsed["step"]
            progress_note = parsed["progress_note"]
            model_output = parsed["model_output"]

            logger.info(
                'Loop decision cid=%s loop=%s step=%s summary="%s" progress="%s"',
                conversation_id,
                loop_index,
                step.get("id"),
                _truncate(step.get("summary") or "", 220),
                _truncate(progress_note, 220),
            )
            if step["id"] in {"provider_route_tool", "aus_metadata_tool", "aus_raw_retrieve_tool", "macro_data_tool", "web_search_tool", "sandbox_tool"}:
                logger.info(
                    'Loop tool input cid=%s loop=%s step=%s input="%s"',
                    conversation_id,
                    loop_index,
                    step.get("id"),
                    _summarize_tool_input(model_output.get("tool_input") if isinstance(model_output, dict) else {}),
                )

            if step["id"] not in {"propose_plan", "compose_final"}:
                emit_status(progress_note)
            _ensure_not_cancelled(conversation_id, cancel_event, f"loop_{loop_index}_after_parse")

            if step["id"] == "propose_plan":
                plan_markdown = str(model_output.get("plan_markdown") or "").strip()
                plan_context = _normalize_plan_context(
                    model_output.get("plan_context"),
                    fallback_question=active_user_message,
                )
                state.pending_plan = {
                    "status": "awaiting_approval",
                    "plan_markdown": plan_markdown,
                    "plan_context": plan_context,
                }
                persist_completed_turn(plan_markdown)
                logger.info(
                    'Loop plan cid=%s loop=%s plan="%s"',
                    conversation_id,
                    loop_index,
                    _truncate(plan_markdown, 280),
                )
                return plan_markdown

            if step["id"] == "compose_final":
                final_answer = str(model_output.get("final_answer_markdown") or "").strip()
                state.pending_plan = None
                state.latest_export_artifact_id = ""
                if _run_has_exportable_support(state, run_loop_start_index):
                    state.latest_export_status = "processing"
                    state.latest_export_request = {
                        "user_message": active_user_message,
                        "final_answer": final_answer,
                        "run_loop_start_index": run_loop_start_index,
                    }
                else:
                    state.latest_export_status = ""
                    state.latest_export_request = None
                persist_completed_turn(final_answer)
                logger.info(
                    'Loop final cid=%s loop=%s preview="%s"',
                    conversation_id,
                    loop_index,
                    _truncate(final_answer, 280),
                )
                return final_answer

            try:
                if step["id"] == "sandbox_tool":
                    retry_conflict = _sandbox_retry_conflict(state, model_output["tool_input"])
                    if retry_conflict:
                        raise RuntimeError(retry_conflict)
                if step["id"] == "provider_route_tool":
                    tool_result = _execute_provider_route_tool(
                        tool_input=model_output["tool_input"],
                        state=state,
                        conversation_id=conversation_id,
                    )
                elif step["id"] == "aus_metadata_tool":
                    tool_result = _execute_aus_metadata_tool(
                        tool_input=model_output["tool_input"],
                        state=state,
                        conversation_id=conversation_id,
                    )
                elif step["id"] == "aus_raw_retrieve_tool":
                    tool_result = _execute_aus_raw_retrieve_tool(
                        tool_input=model_output["tool_input"],
                        state=state,
                        conversation_id=conversation_id,
                    )
                elif step["id"] == "macro_data_tool":
                    tool_result = _execute_macro_data_tool(
                        tool_input=model_output["tool_input"],
                        state=state,
                        conversation_id=conversation_id,
                    )
                elif step["id"] == "web_search_tool":
                    tool_result = _execute_web_search_tool(
                        tool_input=model_output["tool_input"],
                        state=state,
                        conversation_id=conversation_id,
                    )
                elif step["id"] == "sandbox_tool":
                    tool_result = _execute_sandbox_tool(
                        tool_input=model_output["tool_input"],
                        state=state,
                        conversation_id=conversation_id,
                        loop_payload=payload,
                        status_callback=emit_status,
                        usage_callback=record_usage,
                    )
                else:
                    raise RuntimeError(f"Unsupported step id: {step['id']}")
            except Exception as exc:
                logger.exception(
                    "Tool step failed cid=%s loop=%s step=%s error=%s",
                    conversation_id,
                    loop_index,
                    step.get("id"),
                    exc,
                )
                result_summary = (
                    "Tool execution failed. Adjust the next step using this feedback.\n"
                    f"Error: {str(exc)}"
                )
                failure_data = _classify_tool_failure(
                    str(step.get("id") or "").strip(),
                    str(exc),
                    model_output.get("tool_input") if isinstance(model_output, dict) else {},
                )
                retry_guidance = str(failure_data.get("retry_guidance") or "").strip()
                if retry_guidance:
                    result_summary += f"\nRecovery guidance: {retry_guidance}"
                _record_loop_feedback(
                    state,
                    step=step,
                    progress_note=progress_note,
                    result_summary=result_summary,
                    result_data=failure_data,
                )
                store.save(state)
                continue

            result_summary = str(tool_result.get("summary") or "").strip()
            result_data = tool_result.get("result_data")

            if step["id"] == "provider_route_tool" and isinstance(result_data, dict):
                resolved_route = _normalize_provider_route(result_data.get("route"))
                if resolved_route:
                    selected_provider_route = resolved_route
                    pre_run_provider_route["selected_route"] = resolved_route
                    state.last_provider_route = resolved_route
                    selected_route_query = str(result_data.get("search_query") or "").strip()
                    if selected_route_query:
                        pre_run_provider_route["selected_search_query"] = selected_route_query
                        state.last_provider_search_query = selected_route_query
                    else:
                        state.last_provider_search_query = ""
                    if resolved_route == "macro":
                        pre_run_dataset_shortlist = []
                        state.current_abs_dataset_shortlist = []
                    if resolved_route == "aus":
                        pre_run_macro_indicator_shortlist = []
                        state.current_macro_indicator_shortlist = []
                        refresh_query = selected_route_query or active_user_message
                        logger.info(
                            'Provider route requested ABS shortlist refresh cid=%s route=%s query="%s"',
                            conversation_id,
                            resolved_route,
                            _truncate(refresh_query, 220),
                        )
                        pre_run_dataset_shortlist = _refresh_abs_shortlist(
                            state=state,
                            conversation_id=conversation_id,
                            route=resolved_route,
                            shortlist_query=refresh_query,
                        )
            if step["id"] == "macro_data_tool" and isinstance(result_data, dict):
                result_kind = str(result_data.get("kind") or "").strip()
                if result_kind == "macro_indicator_shortlist":
                    candidates = result_data.get("candidates") if isinstance(result_data.get("candidates"), list) else []
                    pre_run_macro_indicator_shortlist = [item for item in candidates if isinstance(item, dict)]
                    state.current_macro_indicator_shortlist = list(pre_run_macro_indicator_shortlist)
                elif result_kind == "macro_candidate_unavailable":
                    candidates = result_data.get("remaining_candidates") if isinstance(result_data.get("remaining_candidates"), list) else []
                    pre_run_macro_indicator_shortlist = [item for item in candidates if isinstance(item, dict)]
                    state.current_macro_indicator_shortlist = list(pre_run_macro_indicator_shortlist)
                elif result_kind == "macro_unavailable":
                    pre_run_macro_indicator_shortlist = []
                    state.current_macro_indicator_shortlist = []

            _record_loop_feedback(
                state,
                step=step,
                progress_note=progress_note,
                result_summary=result_summary,
                result_data=result_data,
            )
            store.save(state)
            logger.info(
                'Loop result cid=%s loop=%s step=%s result="%s"',
                conversation_id,
                loop_index,
                step.get("id"),
                _truncate(result_summary, 280),
            )

        final_answer = _compose_best_effort_final(
            conversation_id,
            active_user_message,
            state,
            usage_callback=record_usage,
        )
        state.pending_plan = None
        persist_completed_turn(final_answer)
        logger.info(
            'Loop max best-effort final cid=%s preview="%s"',
            conversation_id,
            _truncate(final_answer, 280),
        )
        return final_answer
    finally:
        _release_cancellation_event(conversation_id)
