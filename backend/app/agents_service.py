from __future__ import annotations

import asyncio
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
import json
import logging
import re
import shutil
import sys
import time
from pathlib import Path
from threading import Event, Lock
from typing import Any, Callable, Dict, List, Optional

from agents import (
    Agent,
    CodeInterpreterTool,
    ModelSettings,
    ModelRetrySettings,
    Runner,
    SQLiteSession,
    function_tool,
    retry_policies,
    set_default_openai_key,
)
from agents.run_context import RunContextWrapper
from agents.mcp import MCPServerStdio, create_static_tool_filter
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .config import get_settings
from .storage import ConversationStore


settings = get_settings()
logger = logging.getLogger("abs.backend.agents")
if not logger.handlers:
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(
        logging.Formatter("[%(asctime)s] %(levelname)s %(name)s - %(message)s")
    )
    logger.addHandler(stream_handler)
logger.setLevel(logging.INFO)
logger.propagate = False

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SESSION_DB_PATH = settings.runtime_dir / "agent_sdk_sessions.sqlite3"
SESSION_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
GPT_5_4_INPUT_PRICE_PER_MILLION = 2.50
GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION = 0.625
GPT_5_4_OUTPUT_PRICE_PER_MILLION = 10.00
AI_COST_SURCHARGE_RATE = 0.10

_CANCELLATION_LOCK = Lock()
_CANCELLATION_EVENTS: Dict[str, Event] = {}


@dataclass
class AgentRuntimeContext:
    conversation_id: str
    store: ConversationStore
    code_container_id: str
    status_callback: Callable[[str], None]


class ConversationCancelled(RuntimeError):
    """Raised when a conversation is cancelled mid-generation."""


def _acquire_cancellation_event(conversation_id: str) -> Event:
    with _CANCELLATION_LOCK:
        event = _CANCELLATION_EVENTS.get(conversation_id)
        if event is None:
            event = Event()
            _CANCELLATION_EVENTS[conversation_id] = event
        event.clear()
        return event


def cancel_conversation_processing(conversation_id: str) -> None:
    with _CANCELLATION_LOCK:
        event = _CANCELLATION_EVENTS.get(conversation_id)
        if event is None:
            event = Event()
            _CANCELLATION_EVENTS[conversation_id] = event
        event.set()


def _release_cancellation_event(conversation_id: str) -> None:
    with _CANCELLATION_LOCK:
        _CANCELLATION_EVENTS.pop(conversation_id, None)


def _ensure_not_cancelled(conversation_id: str, event: Event, stage: str) -> None:
    if event.is_set():
        logger.info("Conversation cancelled cid=%s stage=%s", conversation_id, stage)
        raise ConversationCancelled(f"Conversation {conversation_id} cancelled during {stage}")


def _conversation_runtime_dir(conversation_id: str) -> Path:
    safe_id = "".join(ch for ch in conversation_id if ch.isalnum() or ch in {"-", "_"})
    if not safe_id:
        safe_id = "conversation"
    return settings.runtime_dir / "conversations" / safe_id


def conversation_session(conversation_id: str) -> SQLiteSession:
    return SQLiteSession(conversation_id, db_path=SESSION_DB_PATH)


def _artifact_file_path(conversation_id: str, artifact_id: str) -> Path:
    return _conversation_runtime_dir(conversation_id) / "artifacts" / f"{artifact_id}.json"


def _trace_file_path(conversation_id: str) -> Path:
    return _conversation_runtime_dir(conversation_id) / "agent_trace.jsonl"


def _ensure_runtime_dirs(conversation_id: str) -> Path:
    run_dir = _conversation_runtime_dir(conversation_id)
    (run_dir / "artifacts").mkdir(parents=True, exist_ok=True)
    return run_dir


def _clear_runtime_dir(conversation_id: str) -> None:
    run_dir = _conversation_runtime_dir(conversation_id)
    if run_dir.exists():
        shutil.rmtree(run_dir, ignore_errors=True)


def reset_conversation_runtime(conversation_id: str) -> None:
    _clear_runtime_dir(conversation_id)


def _session_items_from_state_messages(messages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for message in messages:
        if not isinstance(message, dict):
            continue
        role = str(message.get("role") or "").strip().lower()
        content = str(message.get("content") or "").strip()
        if not content:
            continue
        if role == "user":
            items.append({"type": "message", "role": "user", "content": content})
        elif role == "assistant":
            items.append(
                {
                    "type": "message",
                    "role": "assistant",
                    "phase": "final_answer",
                    "content": content,
                }
            )
    return items


async def clear_agent_session(conversation_id: str) -> None:
    await conversation_session(conversation_id).clear_session()


async def sync_agent_session_from_state(conversation_id: str, state: Any) -> None:
    session = conversation_session(conversation_id)
    await session.clear_session()
    items = _session_items_from_state_messages(list(getattr(state, "messages", []) or []))
    if items:
        await session.add_items(items)


def _truncate(text: Any, length: int = 280) -> str:
    clean = str(text or "").replace("\n", " ").strip()
    return clean if len(clean) <= length else clean[: length - 1] + "…"


def _safe_int(value: Any) -> int:
    try:
        numeric = int(value)
    except Exception:
        return 0
    return numeric if numeric > 0 else 0


def _compute_run_cost_breakdown(
    *,
    input_tokens: int,
    output_tokens: int,
    cached_input_tokens: int = 0,
) -> Dict[str, float]:
    cached_tokens = min(max(cached_input_tokens, 0), max(input_tokens, 0))
    uncached_input_tokens = max(input_tokens, 0) - cached_tokens
    ai_cost = (
        (uncached_input_tokens / 1_000_000) * GPT_5_4_INPUT_PRICE_PER_MILLION
        + (cached_tokens / 1_000_000) * GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION
        + (max(output_tokens, 0) / 1_000_000) * GPT_5_4_OUTPUT_PRICE_PER_MILLION
    )
    surcharge = ai_cost * AI_COST_SURCHARGE_RATE
    final_cost = ai_cost + surcharge
    return {
        "ai_cost_usd": round(ai_cost, 6),
        "surcharge_usd": round(surcharge, 6),
        "final_cost_usd": round(final_cost, 6),
    }


def _build_run_cost_payload(
    *,
    input_tokens: int,
    output_tokens: int,
    cached_input_tokens: int = 0,
    model: str | None = None,
) -> Dict[str, Any]:
    cached_tokens = min(max(cached_input_tokens, 0), max(input_tokens, 0))
    breakdown = _compute_run_cost_breakdown(
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        cached_input_tokens=cached_tokens,
    )
    return {
        "model": str(model or settings.openai_model or "gpt-5.4"),
        "input_tokens": max(input_tokens, 0),
        "cached_input_tokens": cached_tokens,
        "output_tokens": max(output_tokens, 0),
        "pricing": {
            "input_per_million_usd": GPT_5_4_INPUT_PRICE_PER_MILLION,
            "cached_input_per_million_usd": GPT_5_4_CACHED_INPUT_PRICE_PER_MILLION,
            "output_per_million_usd": GPT_5_4_OUTPUT_PRICE_PER_MILLION,
            "surcharge_rate": AI_COST_SURCHARGE_RATE,
        },
        **breakdown,
    }


def _system_instructions() -> str:
    return """
You are Nisaba, an AI economic analyst for Australian public data with supporting global macro context.
You should feel like a precise economic scribe: measured, precise, calm, intellectually honest, economical with words, and quietly confident.
Prefer verified data over guesses, structure over flourish, plain explanation over hype, clean charts over decorative ones, and evidence-led judgment over forced certainty.
Economic statistics measure specific things in specific ways. Name what the data shows. Name what it does not show. If the data does not support a conclusion, say so clearly.

Tooling:
- Web search is disabled for now. Use MCP tools plus the python tool only.
- Use `report_progress` frequently, including after most meaningful steps and whenever the plan materially changes.
- Keep each progress update to one short plain-English sentence saying what you just did and what you will do next, for example: `Checked the shortlist. Next I’m opening the metadata.`
- Do not reveal chain-of-thought or hidden reasoning. Keep updates operational and factual.
- Use the domestic MCP server for Australian domestic data, including ABS and curated custom Australian sources.
- Use the macro MCP server for global macro and trade data via the `macro_search_catalog`, `macro_get_metadata`, and `macro_retrieve` tools.
- Retrieval tools save raw data as server-side artifacts and return compact manifests.
- Domestic MCP provides `inspect_artifact` and `narrow_artifact`.
- Macro MCP provides `macro_inspect_artifact` and `macro_narrow_artifact`.
- If narrowing returns `analysis_file`, open that file in the python tool and use it for calculations, comparisons, and chart preparation.
- For charting or any answer that depends on exact numeric comparisons, use the python tool on the best available analysis-ready artifact before writing the final response.

Retrieval rules:
- For domestic data, search the catalog first unless the dataset is already clear.
- For ABS datasets, follow this flow: shortlist the dataset, inspect metadata when needed to determine retrieval, retrieve the data, inspect the artifact, narrow only if needed, then analyze.
- For curated custom domestic sources, metadata may not be necessary; if retrieval works directly, do not force an artificial metadata step.
- For ABS datasets, `get_metadata` returns `anchor_candidates`. Choose one anchor code only from that metadata and then call `retrieve` with `anchorType` and `anchorCode` so the server constructs the wildcard key.
- Do not invent dataset IDs, filters, anchor codes, or data keys.
- Once MCP retrieval succeeds, stay on the MCP/artifact path.
- Do not analyze large raw retrieval artifacts directly by default. Inspect first, then decide whether the artifact is already narrow enough to use directly or needs narrowing.
- Use the size information returned by inspect to decide whether the artifact can go directly to python or needs narrowing first.
- If inspect shows the artifact is already narrow enough for the user's request, use it directly rather than narrowing for its own sake.
- If an artifact is too large for direct python handoff, narrow it enough to get under the handoff limit.
- When data contains multiple frequencies, countries, or series variants, narrow to the exact comparable slice before answering.
- For comparisons over time, use one comparable definition and one frequency before charting.
- For ABS datasets, use exactly one anchor and wildcard every other segment, matching the metadata-derived template. Do not invent alternate frequency variants or annualized keys unless metadata clearly supports them.
- After a successful ABS retrieve plus inspect, prefer narrowing and analysis over issuing another retrieve.
- For matrix, workbook, supply-use, or input-output style datasets, retrieve the broad published table first, inspect the returned structure, and do not use MCP narrowing as the default next step.
- For supply-use, input-output, and other matrix-style tables, prefer using the full retrieved table directly after inspect when it fits the python handoff limit.
- If a matrix-style artifact is too large, narrow to one correct full matrix or one correct metric/anchor, not to partial rows or columns inside that matrix.
- For matrix-style data, treat totals carefully: do not manually sum rows or columns that already include published total entries, or you will double count.
- For matrix-style data, exclude total rows and total columns before manual summing unless the user explicitly wants the published total itself.

Analysis rules:
- Ground claims in retrieved data.
- Be explicit when comparing periods, units, countries, or series definitions.
- Do not fabricate missing values or missing source coverage.
- This is an economic analyst. Answer with data. For empirical questions, give concrete values, rankings, periods, or charted points from retrieved data; if the data is insufficient, say so plainly instead of substituting generic prose. If you use a proxy, name it explicitly.

Response rules:
- Write a clean final answer in markdown.
- Unless the user explicitly asks for detailed analysis, keep the full final answer to about 150 words or fewer, excluding the chart JSON block and the short source line.
- Keep the answer tight and relevant to the user's request.
- Include source links when available from tool outputs.
- Prefer charts by default whenever the retrieved data can reasonably be visualized.
- Unless the user says otherwise, prefer a chart over a table.
- Do not include both a table and a chart for the same data unless the user explicitly asks for both or a table is clearly necessary for precision.
- For trends over time, comparisons across categories, or shares/compositions, default to a chart if the data supports it.
- When a chart is used, treat it as the main output: chart first, then a very short direct read of what the data shows, then broader interpretation only if genuinely helpful.
- Do not force the answer into rigid titled sections like `Chart`, `Analysis`, or `Interpretation`.
- Let the prose flow naturally and read organically.
- Keep the direct data read tightly grounded in the retrieved data.
- If you offer broader interpretation or a possible explanation of why, make it clearly separate from the direct data read. Use phrasing that distinguishes what the data shows from your interpretation, for example `The data shows ...` versus `One possible explanation is ...`.
- If a chart is appropriate, include a fenced chart block with valid JSON using this schema:
```chart
{
  "type": "line",
  "title": "Short title",
  "xLabel": "X axis",
  "yLabel": "Y axis",
  "series": [
    {
      "name": "Series name",
      "points": [{"x": "2020", "y": 123.4}]
    }
  ]
}
```
- Only include chart blocks when the underlying data is already retrieved and the chart improves the answer.
- Use a table only when the user asks for one, the output is too small to merit a chart, or exact tabular values are the clearest form.
- End with a short `Source` or `Sources` line at the bottom.
- Keep the source line tight: source name plus a clean link where possible, not a long bibliography.
""".strip()


@lru_cache(maxsize=1)
def _openai_client() -> OpenAI:
    return OpenAI(api_key=settings.openai_api_key, timeout=settings.openai_timeout_seconds)


def _create_code_container(conversation_id: str) -> str:
    container = _openai_client().containers.create(
        name=f"nisaba-{conversation_id[:24]}",
        memory_limit="1g",
    )
    return str(container.id)


@function_tool
def report_progress(
    ctx: RunContextWrapper[AgentRuntimeContext],
    message: str,
) -> Dict[str, Any]:
    """Send a short user-facing progress update at a meaningful point in the workflow. Use one plain-English sentence. Keep the user oriented as you go, and send an extra update if you pivot because the data, slice, or source was not right."""
    normalized = _truncate(message, 220)
    if normalized:
        ctx.context.status_callback(normalized)
    return {"ok": True, "message": normalized}


def _build_agent(code_container_id: str) -> Agent[Any]:
    return Agent(
        name="Nisaba",
        model=settings.openai_model,
        instructions=_system_instructions(),
        tools=[
            report_progress,
            CodeInterpreterTool(
                tool_config={
                    "type": "code_interpreter",
                    "container": code_container_id,
                }
            ),
        ],
        model_settings=ModelSettings(
            reasoning={"effort": settings.openai_reasoning_effort},
            retry=ModelRetrySettings(
                max_retries=2,
                policy=retry_policies.provider_suggested(),
                backoff={
                    "initial_delay": 0.5,
                    "max_delay": 2.0,
                    "multiplier": 2.0,
                    "jitter": True,
                },
            ),
        ),
        mcp_config={
            "convert_schemas_to_strict": True,
        },
    )


def _domestic_mcp_server_for_conversation(conversation_id: str, code_container_id: str) -> MCPServerStdio:
    return MCPServerStdio(
        params={
            "command": settings.node_binary,
            "args": [str(PROJECT_ROOT / "build" / "index.js")],
            "cwd": str(PROJECT_ROOT),
            "env": {
                "NISABA_CONVERSATION_ID": conversation_id,
                "NISABA_RUNTIME_DIR": str(settings.runtime_dir),
                "NISABA_CODE_CONTAINER_ID": code_container_id,
                "NISABA_PYTHON_BINARY": settings.python_binary,
                "OPENAI_API_KEY": settings.openai_api_key,
            },
        },
        name="domestic",
        client_session_timeout_seconds=max(120, settings.macro_timeout_seconds),
        cache_tools_list=True,
        tool_filter=create_static_tool_filter(
            allowed_tool_names=["search_catalog", "get_metadata", "retrieve", "inspect_artifact", "narrow_artifact"]
        ),
    )


def _macro_mcp_server_for_conversation(conversation_id: str, code_container_id: str) -> MCPServerStdio:
    return MCPServerStdio(
        params={
            "command": settings.python_binary,
            "args": ["-m", "backend.app.macro_mcp_server"],
            "cwd": str(PROJECT_ROOT),
            "env": {
                "NISABA_CONVERSATION_ID": conversation_id,
                "NISABA_RUNTIME_DIR": str(settings.runtime_dir),
                "NISABA_CODE_CONTAINER_ID": code_container_id,
                "OPENAI_API_KEY": settings.openai_api_key,
            },
        },
        name="macro",
        client_session_timeout_seconds=max(60, settings.macro_timeout_seconds),
        cache_tools_list=True,
        tool_filter=create_static_tool_filter(
            allowed_tool_names=[
                "macro_search_catalog",
                "macro_get_metadata",
                "macro_retrieve",
                "macro_inspect_artifact",
                "macro_narrow_artifact",
            ]
        ),
    )


def _next_artifact_id(artifacts: List[Dict[str, Any]]) -> str:
    return f"artifact-{len(artifacts) + 1:03d}"


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _tool_args_summary(tool_args: Dict[str, Any]) -> Dict[str, Any]:
    summary: Dict[str, Any] = {}
    for key in (
        "query",
        "searchQuery",
        "datasetId",
        "candidateId",
        "dataKey",
        "startPeriod",
        "endPeriod",
        "startYear",
        "endYear",
        "detail",
        "limit",
    ):
        value = tool_args.get(key)
        if value in (None, "", [], {}):
            continue
        if isinstance(value, str):
            summary[key] = _truncate(value, 120)
        else:
            summary[key] = value
    for key in ("countries", "reporterCodes", "partnerCodes", "hsCodes"):
        value = tool_args.get(key)
        if isinstance(value, list) and value:
            summary[key] = value[:6]
            if len(value) > 6:
                summary[f"{key}_count"] = len(value)
    return summary


def _tool_output_summary(payload: Any) -> Dict[str, Any]:
    if isinstance(payload, dict):
        summary: Dict[str, Any] = {"keys": sorted(str(key) for key in list(payload.keys())[:12])}
        if payload.get("artifact_id"):
            summary["artifact_id"] = _truncate(payload.get("artifact_id"), 120)
        if payload.get("kind"):
            summary["kind"] = _truncate(payload.get("kind"), 120)
        analysis_file = payload.get("analysis_file")
        if isinstance(analysis_file, dict):
            summary["analysis_filename"] = _truncate(analysis_file.get("filename") or "", 120)
        if isinstance(payload.get("dataflows"), list):
            summary["dataflows"] = len(payload["dataflows"])
        if isinstance(payload.get("candidates"), list):
            summary["candidates"] = len(payload["candidates"])
        if isinstance(payload.get("series"), list):
            summary["series"] = len(payload["series"])
        manifest = payload.get("manifest")
        if isinstance(manifest, dict):
            if manifest.get("series_count") is not None:
                summary["series_count"] = manifest.get("series_count")
            if manifest.get("observation_count") is not None:
                summary["observation_count"] = manifest.get("observation_count")
            if manifest.get("point_count") is not None:
                summary["point_count"] = manifest.get("point_count")
        dataset = payload.get("dataset")
        if isinstance(dataset, dict):
            summary["datasetId"] = _truncate(dataset.get("id") or dataset.get("dataset_id") or "", 120)
            summary["datasetName"] = _truncate(dataset.get("name") or "", 120)
        selected = payload.get("selected_indicator")
        if isinstance(selected, dict):
            summary["indicator"] = _truncate(
                selected.get("indicator_label") or selected.get("provider_name") or "",
                120,
            )
        if payload.get("provider"):
            summary["provider"] = _truncate(payload.get("provider"), 120)
        return summary
    if isinstance(payload, list):
        return {"type": "list", "count": len(payload)}
    if isinstance(payload, str):
        return {"type": "text", "preview": _truncate(payload, 160)}
    return {"type": type(payload).__name__}


def _display_tool_args_summary(state, tool_name: str, tool_args: Dict[str, Any]) -> Dict[str, Any]:
    summary = _tool_args_summary(tool_args)
    clean_name = str(tool_name or "").strip().lower()
    if summary:
        return summary
    if clean_name in {"inspect_artifact", "narrow_artifact", "macro_inspect_artifact", "macro_narrow_artifact"}:
        record = _latest_artifact_record(state)
        if record:
            return {
                "artifactId": str(record.get("artifact_id") or "").strip(),
                "kind": str(record.get("kind") or "").strip(),
                "label": _truncate(str(record.get("label") or "").strip(), 120),
                "inferred": "latest_artifact",
            }
    if clean_name == "code_interpreter":
        record = _latest_artifact_record(state)
        if record:
            summary = {
                "artifactId": str(record.get("artifact_id") or "").strip(),
                "kind": str(record.get("kind") or "").strip(),
                "label": _truncate(str(record.get("label") or "").strip(), 120),
            }
            analysis_filename = str(record.get("analysis_filename") or "").strip()
            if analysis_filename:
                summary["analysis_filename"] = _truncate(analysis_filename, 120)
            summary["inferred"] = "latest_analysis_artifact"
            return summary
    return summary


def _event_payload_preview(value: Any, length: int = 400) -> str:
    coerced = _coerce_jsonable(value)
    if coerced is None:
        return ""
    try:
        if isinstance(coerced, (dict, list)):
            return _truncate(json.dumps(coerced, ensure_ascii=False, sort_keys=True), length)
        return _truncate(str(coerced), length)
    except Exception:
        return _truncate(repr(coerced), length)


def _artifact_record_by_id(state, artifact_id: str) -> Optional[Dict[str, Any]]:
    target = str(artifact_id or "").strip()
    if not target:
        return None
    for item in reversed(state.artifacts):
        if not isinstance(item, dict):
            continue
        if str(item.get("artifact_id") or "").strip() == target:
            return item
    return None


def _latest_artifact_record(state) -> Optional[Dict[str, Any]]:
    for item in reversed(state.artifacts):
        if isinstance(item, dict):
            return item
    return None


def _artifact_payload_from_record(record: Dict[str, Any]) -> Any:
    path = Path(str(record.get("path") or ""))
    if not path.exists() or path.suffix.lower() != ".json":
        raise RuntimeError(f"Artifact file is not available for {record.get('artifact_id')}.")
    return json.loads(path.read_text(encoding="utf-8"))


def _domestic_preview_rows(payload: Dict[str, Any], limit: int = 8) -> List[Dict[str, Any]]:
    headers, rows = _flatten_domestic_payload(payload)
    preview: List[Dict[str, Any]] = []
    for row in rows[:limit]:
        preview.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return preview


def _macro_preview_rows(payload: Dict[str, Any], limit: int = 8) -> List[Dict[str, Any]]:
    headers, rows = _flatten_macro_payload(payload)
    preview: List[Dict[str, Any]] = []
    for row in rows[:limit]:
        preview.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return preview


def _artifact_manifest_summary(record: Dict[str, Any], payload: Dict[str, Any]) -> Dict[str, Any]:
    kind = str(record.get("kind") or "").strip()
    summary: Dict[str, Any] = {
        "artifact_id": str(record.get("artifact_id") or "").strip(),
        "kind": kind,
        "label": str(record.get("label") or "").strip(),
        "summary": str(record.get("summary") or "").strip(),
    }
    if kind.startswith("domestic"):
        dataset = payload.get("dataset") if isinstance(payload.get("dataset"), dict) else {}
        series_items = payload.get("series") if isinstance(payload.get("series"), list) else []
        observation_count = 0
        dimension_values: Dict[str, List[str]] = {}
        for series in series_items:
            if not isinstance(series, dict):
                continue
            observations = series.get("observations") if isinstance(series.get("observations"), list) else []
            observation_count += len(observations)
            series_dims = series.get("dimensions") if isinstance(series.get("dimensions"), dict) else {}
            for key, value in series_dims.items():
                label = value.get("label") if isinstance(value, dict) else value
                if label is None:
                    continue
                clean = str(label).strip()
                if not clean:
                    continue
                dimension_values.setdefault(str(key), [])
                if clean not in dimension_values[str(key)] and len(dimension_values[str(key)]) < 6:
                    dimension_values[str(key)].append(clean)
        summary.update(
            {
                "dataset_id": str(dataset.get("id") or "").strip(),
                "series_count": len(series_items),
                "observation_count": observation_count,
                "dimensions": dimension_values,
                "preview_rows": _domestic_preview_rows(payload),
            }
        )
    elif kind.startswith("macro"):
        series_items = payload.get("series") if isinstance(payload.get("series"), list) else []
        point_count = 0
        countries: List[str] = []
        frequencies: List[str] = []
        for series in series_items:
            if not isinstance(series, dict):
                continue
            point_count += len(series.get("points") if isinstance(series.get("points"), list) else [])
            country = str(series.get("country_code") or series.get("country") or "").strip()
            if country and country not in countries and len(countries) < 12:
                countries.append(country)
            frequency = str(series.get("frequency") or "").strip()
            if frequency and frequency not in frequencies:
                frequencies.append(frequency)
        summary.update(
            {
                "provider": str(payload.get("provider") or payload.get("provider_key") or "").strip(),
                "series_count": len(series_items),
                "point_count": point_count,
                "countries": countries,
                "frequencies": frequencies,
                "preview_rows": _macro_preview_rows(payload),
            }
        )
    parent_artifact_id = str(record.get("parent_artifact_id") or "").strip()
    if parent_artifact_id:
        summary["parent_artifact_id"] = parent_artifact_id
    analysis_filename = str(record.get("analysis_filename") or "").strip()
    analysis_container_id = str(record.get("analysis_container_id") or "").strip()
    if analysis_filename and analysis_container_id:
        summary["analysis_file"] = {
            "filename": analysis_filename,
            "container_id": analysis_container_id,
            "artifact_id": str(record.get("artifact_id") or "").strip(),
        }
    return summary


def _make_artifact_record(
    *,
    state,
    path: Path,
    kind: str,
    label: str,
    summary: str,
    artifact_id: str | None = None,
    source_references: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    record = {
        "artifact_id": str(artifact_id or "").strip() or _next_artifact_id(state.artifacts),
        "path": str(path),
        "kind": kind,
        "label": label,
        "summary": summary,
    }
    if source_references:
        record["source_references"] = source_references
    state.artifacts.append(record)
    return record


def _coerce_jsonable(value: Any) -> Any:
    if hasattr(value, "model_dump") and callable(value.model_dump):
        try:
            return value.model_dump(mode="json", exclude_none=True)
        except TypeError:
            return value.model_dump()
    if isinstance(value, list):
        return [_coerce_jsonable(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _coerce_jsonable(item) for key, item in value.items()}
    return value


def _extract_text_output(value: Any) -> str:
    value = _coerce_jsonable(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        fragments: List[str] = []
        for item in value:
            text = _extract_text_output(item)
            if text:
                fragments.append(text)
        return "\n".join(fragment for fragment in fragments if fragment).strip()
    if isinstance(value, dict):
        if value.get("type") == "text" and value.get("text") is not None:
            return str(value.get("text") or "").strip()
        if isinstance(value.get("content"), list):
            return _extract_text_output(value.get("content"))
        if value.get("text") is not None and len(value.keys()) <= 3:
            return str(value.get("text") or "").strip()
    return ""


def _extract_json_payload(value: Any) -> Any:
    value = _coerce_jsonable(value)
    if isinstance(value, dict):
        text = _extract_text_output(value).strip()
        if text:
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                match = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", text)
                if match:
                    try:
                        return json.loads(match.group(1))
                    except json.JSONDecodeError:
                        pass
        output_value = value.get("output")
        if output_value is not None:
            nested = _extract_json_payload(output_value)
            if nested is not None:
                return nested
        content_value = value.get("content")
        if content_value is not None:
            nested = _extract_json_payload(content_value)
            if nested is not None:
                return nested
        return value
    if isinstance(value, list):
        if len(value) == 1:
            nested = _extract_json_payload(value[0])
            if nested is not None:
                return nested
        return value
    text = _extract_text_output(value).strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", text)
        if match:
            try:
                return json.loads(match.group(1))
            except json.JSONDecodeError:
                return None
    return None


def _extract_raw_item(value: Any) -> Any:
    raw_item = getattr(value, "raw_item", None)
    return raw_item if raw_item is not None else value


def _extract_raw_item_type(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        candidate = raw.get("type")
        return str(candidate).strip() if candidate is not None else ""
    candidate = getattr(raw, "type", None)
    return str(candidate).strip() if candidate is not None else ""


def _extract_mcp_server_label(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        candidate = raw.get("server_label") or raw.get("serverLabel")
        return str(candidate).strip() if candidate is not None else ""
    for key in ("server_label", "serverLabel"):
        candidate = getattr(raw, key, None)
        if candidate is not None:
            return str(candidate).strip()
    return ""


def _extract_call_id(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        for key in ("call_id", "id"):
            item = raw.get(key)
            if isinstance(item, str) and item.strip():
                return item.strip()
        return ""
    for key in ("call_id", "id"):
        item = getattr(raw, key, None)
        if isinstance(item, str) and item.strip():
            return item.strip()
    return ""


def _extract_tool_name(value: Any) -> str:
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        for key in ("name", "tool_name"):
            item = raw.get(key)
            if isinstance(item, str) and item.strip():
                return item.strip()
        raw_type = str(raw.get("type") or "").strip().lower()
        if raw_type in {"web_search_call", "web_search_preview"}:
            return "web_search"
        if raw_type in {"code_interpreter_call", "code_interpreter_tool_call"}:
            return "code_interpreter"
        if raw_type in {"mcp_call"}:
            item = raw.get("name")
            if isinstance(item, str) and item.strip():
                return item.strip()
        return ""
    for key in ("name", "tool_name"):
        item = getattr(raw, key, None)
        if isinstance(item, str) and item.strip():
            return item.strip()
    raw_type = str(getattr(raw, "type", "") or "").strip().lower()
    if raw_type in {"web_search_call", "web_search_preview"}:
        return "web_search"
    if raw_type in {"code_interpreter_call", "code_interpreter_tool_call"}:
        return "code_interpreter"
    return ""


def _extract_tool_arguments(value: Any) -> Dict[str, Any]:
    raw = _extract_raw_item(value)
    arguments = None
    if isinstance(raw, dict):
        arguments = raw.get("arguments")
    else:
        arguments = getattr(raw, "arguments", None)
    if isinstance(arguments, dict):
        return arguments
    if isinstance(arguments, str) and arguments.strip():
        try:
            parsed = json.loads(arguments)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    input_payload = getattr(raw, "input", None)
    if isinstance(input_payload, dict):
        return input_payload
    return {}


def _extract_tool_output_payload(value: Any) -> Any:
    direct_output = getattr(value, "output", None)
    payload = _extract_json_payload(direct_output)
    if payload is not None:
        return payload
    raw = _extract_raw_item(value)
    if isinstance(raw, dict):
        payload = _extract_json_payload(raw.get("output"))
        if payload is not None:
            return payload
    else:
        payload = _extract_json_payload(getattr(raw, "output", None))
        if payload is not None:
            return payload
    return None


def _tool_transport(tool_name: str, item: Any) -> str:
    clean_name = str(tool_name or "").strip().lower()
    raw_type = _extract_raw_item_type(item).lower()
    if clean_name == "web_search" or "web_search" in raw_type:
        return "web"
    if clean_name == "code_interpreter" or "code_interpreter" in raw_type:
        return "code"
    if clean_name.startswith("macro_") or clean_name in {
        "search_catalog",
        "get_metadata",
        "retrieve",
        "inspect_artifact",
        "narrow_artifact",
    }:
        return "mcp"
    if raw_type.startswith("mcp"):
        return "mcp"
    return "tool"


def _append_trace_event(conversation_id: str, payload: Dict[str, Any]) -> None:
    path = _trace_file_path(conversation_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    entry = {"ts": datetime.utcnow().isoformat(timespec="milliseconds") + "Z", **payload}
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False, sort_keys=True) + "\n")


def _looks_like_domestic_dataset(payload: Any) -> bool:
    return (
        isinstance(payload, dict)
        and isinstance(payload.get("dataset"), dict)
        and isinstance(payload.get("series"), list)
    )


def _looks_like_macro_result(payload: Any) -> bool:
    return (
        isinstance(payload, dict)
        and isinstance(payload.get("series"), list)
        and (
            payload.get("provider")
            or payload.get("provider_key")
            or isinstance(payload.get("selected_indicator"), dict)
        )
        and not isinstance(payload.get("dataset"), dict)
    )


def _persist_retrieval_artifact(
    *,
    state,
    conversation_id: str,
    payload: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    run_dir = _ensure_runtime_dirs(conversation_id)

    if (
        isinstance(payload, dict)
        and str(payload.get("artifact_id") or "").strip()
        and str(payload.get("kind") or "").strip() in {
            "domestic_retrieve",
            "macro_retrieve",
            "domestic_narrowed",
            "macro_narrowed",
        }
    ):
        artifact_id = str(payload.get("artifact_id") or "").strip()
        existing = _artifact_record_by_id(state, artifact_id)
        if existing is not None:
            return existing
        path = _artifact_file_path(conversation_id, artifact_id)
        record = _make_artifact_record(
            state=state,
            path=path,
            kind=str(payload.get("kind") or "").strip(),
            label=str(payload.get("label") or artifact_id).strip(),
            summary=_truncate(str(payload.get("summary") or "").strip() or f"Stored artifact {artifact_id}.", 300),
            artifact_id=artifact_id,
            source_references=payload.get("source_references") if isinstance(payload.get("source_references"), list) else None,
        )
        parent_artifact_id = str(payload.get("parent_artifact_id") or "").strip()
        if parent_artifact_id:
            record["parent_artifact_id"] = parent_artifact_id
        for key in ("analysis_container_id", "analysis_file_id", "analysis_filename", "analysis_local_path"):
            value = str(payload.get(key) or "").strip()
            if value:
                record[key] = value
        return record

    if _looks_like_domestic_dataset(payload):
        dataset = payload.get("dataset") if isinstance(payload.get("dataset"), dict) else {}
        label = str(dataset.get("name") or dataset.get("id") or "Domestic dataset").strip()
        artifact_path = run_dir / "artifacts" / f"domestic_retrieve_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, payload)
        return _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="domestic_retrieve",
            label=label,
            summary=_truncate(f"Retrieved domestic dataset '{label}'.", 300),
            source_references=payload.get("source_references") if isinstance(payload.get("source_references"), list) else None,
        )

    if _looks_like_macro_result(payload):
        selected = payload.get("selected_indicator") if isinstance(payload.get("selected_indicator"), dict) else {}
        label = str(
            selected.get("indicator_label")
            or payload.get("concept_label")
            or payload.get("provider")
            or "Macro dataset"
        ).strip()
        artifact_path = run_dir / "artifacts" / f"macro_retrieve_{len(state.artifacts) + 1:03d}.json"
        _write_json(artifact_path, payload)
        return _make_artifact_record(
            state=state,
            path=artifact_path,
            kind="macro_retrieve",
            label=label,
            summary=_truncate(f"Retrieved macro dataset '{label}'.", 300),
            source_references=payload.get("source_references") if isinstance(payload.get("source_references"), list) else None,
        )

    return None


def _safe_sheet_text(value: Any) -> str:
    text = str(value or "").strip()
    return text if len(text) <= 32000 else text[:31997] + "..."


def _parse_chart_spec_from_markdown(markdown: str) -> Dict[str, Any] | None:
    text = str(markdown or "").strip()
    if not text:
        return None
    match = re.search(r"```chart\s*(\{.*?\})\s*```", text, flags=re.DOTALL)
    if not match:
        return None
    try:
        parsed = json.loads(match.group(1))
    except Exception:
        return None
    if not isinstance(parsed, dict):
        return None
    series = parsed.get("series")
    if not isinstance(series, list) or not series:
        return None
    normalized_series: List[Dict[str, Any]] = []
    for entry in series:
        if not isinstance(entry, dict):
            continue
        points = entry.get("points")
        if not isinstance(points, list) or not points:
            continue
        normalized_points = []
        for point in points:
            if not isinstance(point, dict):
                continue
            x = str(point.get("x") or "").strip()
            y = point.get("y")
            if not x:
                continue
            try:
                numeric_y = float(y)
            except Exception:
                continue
            normalized_points.append({"x": x, "y": numeric_y})
        if normalized_points:
            normalized_series.append(
                {
                    "name": str(entry.get("name") or "Series").strip() or "Series",
                    "points": normalized_points,
                }
            )
    if not normalized_series:
        return None
    return {
        "type": str(parsed.get("type") or "line").strip() or "line",
        "title": str(parsed.get("title") or "").strip(),
        "xLabel": str(parsed.get("xLabel") or "").strip(),
        "yLabel": str(parsed.get("yLabel") or "").strip(),
        "series": normalized_series,
    }


def _chart_table(chart_spec: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    series = chart_spec.get("series") if isinstance(chart_spec.get("series"), list) else []
    if not series:
        return [], []
    x_values: List[str] = []
    for entry in series:
        if not isinstance(entry, dict):
            continue
        for point in entry.get("points") or []:
            if not isinstance(point, dict):
                continue
            x = str(point.get("x") or "").strip()
            if x and x not in x_values:
                x_values.append(x)
    headers = [str(chart_spec.get("xLabel") or "Metric").strip() or "Metric"]
    headers.extend(str(entry.get("name") or "Series").strip() or "Series" for entry in series if isinstance(entry, dict))
    rows: List[List[Any]] = []
    for x in x_values:
        row: List[Any] = [x]
        for entry in series:
            point_map = {
                str(point.get("x") or "").strip(): point.get("y")
                for point in (entry.get("points") or [])
                if isinstance(point, dict)
            }
            row.append(point_map.get(x))
        rows.append(row)
    return headers, rows


def _safe_sheet_name(value: str, used: set[str]) -> str:
    cleaned = re.sub(r'[:\\/*?\[\]]+', " ", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned)[:31].strip() or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f" {suffix}"
        candidate = (cleaned[: 31 - len(tail)] + tail).strip()
        suffix += 1
    used.add(candidate)
    return candidate


def _safe_export_filename(user_message: str) -> str:
    words = re.findall(r"[A-Za-z0-9]+", str(user_message or "").lower())
    meaningful = [word for word in words if len(word) > 2][:6]
    stem = "-".join(meaningful) or "analysis-export"
    stem = stem[:64].strip("-") or "analysis-export"
    return f"{stem}.xlsx"


def _write_table(sheet, headers: List[Any], rows: List[List[Any]]) -> None:
    if headers:
        sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _flatten_domestic_payload(payload: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    series_items = payload.get("series")
    if not isinstance(series_items, list):
        return [], []

    dimension_keys: List[str] = []
    attribute_keys: List[str] = []
    for series in series_items:
        if not isinstance(series, dict):
            continue
        for key in (series.get("dimensions") or {}).keys():
            if key not in dimension_keys:
                dimension_keys.append(str(key))
        for observation in series.get("observations") or []:
            if not isinstance(observation, dict):
                continue
            for key in (observation.get("dimensions") or {}).keys():
                if key not in dimension_keys:
                    dimension_keys.append(str(key))
            for key in (observation.get("attributes") or {}).keys():
                if key not in attribute_keys:
                    attribute_keys.append(str(key))
        for key in (series.get("attributes") or {}).keys():
            if key not in attribute_keys:
                attribute_keys.append(str(key))

    headers = ["seriesKey"] + dimension_keys + ["observationKey", "value"] + attribute_keys
    rows: List[List[Any]] = []

    def _label_or_value(value: Any) -> Any:
        if isinstance(value, dict):
            if value.get("label") is not None:
                return value.get("label")
            if value.get("code") is not None:
                return value.get("code")
        return value

    for series in series_items:
        if not isinstance(series, dict):
            continue
        series_dims = series.get("dimensions") if isinstance(series.get("dimensions"), dict) else {}
        series_attrs = series.get("attributes") if isinstance(series.get("attributes"), dict) else {}
        observations = series.get("observations") if isinstance(series.get("observations"), list) else []
        for observation in observations:
            if not isinstance(observation, dict):
                continue
            obs_dims = observation.get("dimensions") if isinstance(observation.get("dimensions"), dict) else {}
            obs_attrs = observation.get("attributes") if isinstance(observation.get("attributes"), dict) else {}
            row: List[Any] = [series.get("seriesKey")]
            for key in dimension_keys:
                value = obs_dims.get(key)
                if value is None:
                    value = series_dims.get(key)
                row.append(_label_or_value(value))
            row.append(observation.get("observationKey"))
            row.append(observation.get("value"))
            for key in attribute_keys:
                value = obs_attrs.get(key)
                if value is None:
                    value = series_attrs.get(key)
                row.append(_label_or_value(value))
            rows.append(row)

    return headers, rows


def _flatten_macro_payload(payload: Dict[str, Any]) -> tuple[List[str], List[List[Any]]]:
    series_items = payload.get("series")
    if not isinstance(series_items, list):
        return [], []
    headers = ["provider", "country", "country_code", "indicator", "series_id", "frequency", "unit", "x", "y"]
    rows: List[List[Any]] = []
    for series in series_items:
        if not isinstance(series, dict):
            continue
        for point in series.get("points") or []:
            if not isinstance(point, dict):
                continue
            rows.append(
                [
                    series.get("provider"),
                    series.get("country"),
                    series.get("country_code"),
                    series.get("indicator"),
                    series.get("series_id"),
                    series.get("frequency"),
                    series.get("unit"),
                    point.get("x"),
                    point.get("y"),
                ]
            )
    return headers, rows


def _write_raw_artifact_sheet(sheet, record: Dict[str, Any]) -> None:
    path = Path(str(record.get("path") or ""))
    payload: Any = None
    if path.exists() and path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
    source_refs = record.get("source_references") if isinstance(record.get("source_references"), list) else []
    source_line = ""
    if source_refs:
        first = source_refs[0]
        if isinstance(first, dict):
            source_line = " | ".join(
                part
                for part in [
                    str(first.get("provider") or "").strip(),
                    str(first.get("dataset_id") or first.get("series_id") or "").strip(),
                    str(first.get("title") or first.get("indicator") or "").strip(),
                    str(first.get("url") or "").strip(),
                ]
                if part
            )
    sheet["A1"] = f"Source: {source_line}".strip() if source_line else "Source data"
    sheet["A2"] = f"Artifact: {str(record.get('label') or record.get('artifact_id') or '').strip()}".strip()
    sheet["A3"] = ""
    sheet["A4"] = "Returned data"

    headers: List[str] = []
    rows: List[List[Any]] = []
    kind = str(record.get("kind") or "").strip()
    if isinstance(payload, dict):
        if kind.startswith("domestic"):
            headers, rows = _flatten_domestic_payload(payload)
        elif kind.startswith("macro"):
            headers, rows = _flatten_macro_payload(payload)

    if headers and rows:
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=5, column=col_index, value=header)
        for row_offset, row in enumerate(rows, start=6):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_offset, column=col_index, value=value)
    else:
        raw = json.dumps(payload, ensure_ascii=False, indent=2) if isinstance(payload, (dict, list)) else str(payload or "")
        sheet.cell(row=5, column=1, value=_safe_sheet_text(raw))
        sheet.column_dimensions["A"].width = 140

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")


def _apply_export_theme(workbook) -> None:
    summary = workbook["Summary"] if "Summary" in workbook.sheetnames else None
    if summary is None:
        return
    title_font = Font(color="234233", size=16)
    section_font = Font(color="8F6A3A", size=12)
    header_font = Font(color="54745F", size=11)
    for cell in summary[1]:
        cell.font = title_font
    for row in range(1, summary.max_row + 1):
        first_value = str(summary.cell(row=row, column=1).value or "").strip()
        if first_value in {"Presented data", "Retrieved data"}:
            summary.cell(row=row, column=1).font = section_font
            continue
        values = [str(summary.cell(row=row, column=col).value or "").strip() for col in range(1, summary.max_column + 1)]
        non_empty = [value for value in values if value]
        if len(non_empty) >= 2 and row > 1:
            for col in range(1, len(values) + 1):
                summary.cell(row=row, column=col).font = header_font
            break


def get_latest_export_artifact_path(state) -> Path | None:
    target_id = str(getattr(state, "latest_export_artifact_id", "") or "").strip()
    if not target_id:
        return None
    for item in reversed(state.artifacts):
        if not isinstance(item, dict):
            continue
        if str(item.get("artifact_id") or "").strip() != target_id:
            continue
        path_value = str(item.get("path") or "").strip()
        if not path_value:
            return None
        path = Path(path_value)
        return path if path.exists() else None
    return None


def generate_latest_export(conversation_id: str, store: ConversationStore) -> Path | None:
    state = store.load(conversation_id)
    request = state.latest_export_request if isinstance(state.latest_export_request, dict) else None
    if not request:
        return get_latest_export_artifact_path(state)

    try:
        _build_answer_export(
            state=state,
            conversation_id=conversation_id,
            user_message=str(request.get("user_message") or "").strip(),
            final_answer=str(request.get("final_answer") or "").strip(),
            run_artifact_start_index=int(request.get("run_artifact_start_index") or 0),
        )
        state.latest_export_status = "ready"
        state.latest_export_request = None
        store.save(state)
    except Exception:
        state.latest_export_status = "failed"
        state.latest_export_request = None
        store.save(state)
        raise

    return get_latest_export_artifact_path(state)


def _build_answer_export(
    *,
    state,
    conversation_id: str,
    user_message: str,
    final_answer: str,
    run_artifact_start_index: int,
) -> str:
    chart_spec = _parse_chart_spec_from_markdown(final_answer)
    run_artifacts = [
        item
        for item in state.artifacts[run_artifact_start_index:]
        if isinstance(item, dict) and str(item.get("kind") or "").strip() in {"domestic_retrieve", "macro_retrieve"}
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["Summary"])
    sheet.append(["Question", user_message])
    sheet.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")])

    source_lines: List[str] = []
    seen_sources: set[str] = set()
    for artifact in run_artifacts:
        refs = artifact.get("source_references") if isinstance(artifact.get("source_references"), list) else []
        for item in refs:
            if not isinstance(item, dict):
                continue
            line = " | ".join(
                part
                for part in [
                    str(item.get("provider") or "").strip(),
                    str(item.get("dataset_id") or item.get("series_id") or "").strip(),
                    str(item.get("title") or item.get("indicator") or "").strip(),
                    str(item.get("url") or "").strip(),
                ]
                if part
            ).strip()
            if line and line not in seen_sources:
                seen_sources.add(line)
                source_lines.append(line)

    if source_lines:
        sheet.append(["Sources", _safe_sheet_text(source_lines[0])])
        for line in source_lines[1:12]:
            sheet.append(["", _safe_sheet_text(line)])

    sheet.append([])

    chart_headers, chart_rows = _chart_table(chart_spec or {})
    if chart_headers and chart_rows:
        sheet.append(["Presented data"])
        _write_table(sheet, chart_headers, chart_rows)
        sheet.append([])

    if run_artifacts:
        sheet.append(["Retrieved data"])
        _write_table(
            sheet,
            ["Artifact", "Kind", "Summary"],
            [
                [
                    str(item.get("label") or item.get("artifact_id") or "").strip(),
                    str(item.get("kind") or "").strip(),
                    str(item.get("summary") or "").strip(),
                ]
                for item in run_artifacts
            ],
        )

    widths = {"A": 18, "B": 44, "C": 24, "D": 24, "E": 20, "F": 20}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    used_sheet_names = {"Summary"}
    for artifact in run_artifacts:
        raw_sheet = workbook.create_sheet(
            title=_safe_sheet_name(
                str(artifact.get("label") or artifact.get("artifact_id") or "Raw data"),
                used_sheet_names,
            )
        )
        _write_raw_artifact_sheet(raw_sheet, artifact)

    _apply_export_theme(workbook)
    run_dir = _ensure_runtime_dirs(conversation_id)
    download_filename = _safe_export_filename(user_message)
    path = run_dir / "artifacts" / f"answer_export_{len(state.artifacts) + 1:03d}.xlsx"
    workbook.save(path)
    source_references: List[Dict[str, Any]] = []
    for artifact in run_artifacts:
        refs = artifact.get("source_references") if isinstance(artifact.get("source_references"), list) else []
        source_references.extend([item for item in refs if isinstance(item, dict)])
    record = _make_artifact_record(
        state=state,
        path=path,
        kind="answer_export",
        label="Excel export",
        summary=_truncate(f"Summary workbook with retrieved data for '{user_message}'.", 300),
        source_references=source_references[:12],
    )
    record["download_filename"] = download_filename
    state.latest_export_artifact_id = record["artifact_id"]
    return record["artifact_id"]


async def _generate_response_async(
    conversation_id: str,
    user_input: str,
    store: ConversationStore,
    status_callback: Callable[[str], None],
) -> str:
    set_default_openai_key(settings.openai_api_key, use_for_tracing=False)
    cancel_event = _acquire_cancellation_event(conversation_id)
    state = store.load(conversation_id)
    run_dir = _ensure_runtime_dirs(conversation_id)
    run_artifact_start_index = len(state.artifacts)
    processed_tool_output_call_ids: set[str] = set()
    tool_call_names: Dict[str, str] = {}
    tool_call_started_at: Dict[str, float] = {}
    last_status = ""
    saved_progress_messages: List[str] = []
    _append_trace_event(
        conversation_id,
        {
            "event": "run_started",
            "user_input": _truncate(user_input, 400),
            "model": settings.openai_model,
            "reasoning_effort": settings.openai_reasoning_effort,
        },
    )

    def emit_status(message: str) -> None:
        nonlocal last_status
        normalized = str(message or "").strip()
        if not normalized or normalized == last_status:
            return
        last_status = normalized
        if not saved_progress_messages or saved_progress_messages[-1] != normalized:
            saved_progress_messages.append(normalized)
        status_callback(normalized)

    state.messages.append({"role": "user", "content": user_input})
    state.latest_export_artifact_id = ""
    state.latest_export_request = None
    state.latest_export_status = ""
    store.save(state)
    session = conversation_session(conversation_id)

    code_container_id = _create_code_container(conversation_id)
    _append_trace_event(
        conversation_id,
        {
            "event": "code_container_created",
            "container_id": code_container_id,
        },
    )

    agent = _build_agent(code_container_id)
    domestic_server = _domestic_mcp_server_for_conversation(conversation_id, code_container_id)
    macro_server = _macro_mcp_server_for_conversation(conversation_id, code_container_id)
    agent.mcp_servers = [domestic_server, macro_server]
    runtime_context = AgentRuntimeContext(
        conversation_id=conversation_id,
        store=store,
        code_container_id=code_container_id,
        status_callback=emit_status,
    )

    _ensure_not_cancelled(conversation_id, cancel_event, "before_run")

    try:
        async with domestic_server, macro_server:
            result = Runner.run_streamed(
                agent,
                user_input,
                context=runtime_context,
                session=session,
                max_turns=20,
            )

            async for event in result.stream_events():
                if cancel_event.is_set():
                    result.cancel(mode="after_turn")
                if getattr(event, "type", "") != "run_item_stream_event":
                    continue

                item = getattr(event, "item", None)
                event_name = str(getattr(event, "name", "") or "").strip()
                if event_name in {"tool_search_called", "tool_search_output_created"}:
                    payload_preview = _event_payload_preview(_extract_raw_item(item))
                    logger.info(
                        "Tool search event cid=%s event=%s payload=%s",
                        conversation_id,
                        event_name,
                        payload_preview or "-",
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": event_name,
                            "payload_preview": payload_preview,
                        },
                    )
                    continue

                if event_name == "tool_called":
                    tool_name = _extract_tool_name(item)
                    call_id = _extract_call_id(item)
                    tool_args = _extract_tool_arguments(item)
                    raw_item_type = _extract_raw_item_type(item)
                    server_label = _extract_mcp_server_label(item)
                    transport = _tool_transport(tool_name, item)
                    if call_id and tool_name:
                        tool_call_names[call_id] = tool_name
                        tool_call_started_at[call_id] = time.perf_counter()
                    logger.info(
                        "Tool call start cid=%s call_id=%s transport=%s tool=%s raw_type=%s server=%s args=%s",
                        conversation_id,
                        call_id or "-",
                        transport,
                        tool_name or "-",
                        raw_item_type or "-",
                        server_label or "-",
                        json.dumps(
                            _display_tool_args_summary(state, tool_name, tool_args),
                            ensure_ascii=False,
                            sort_keys=True,
                        ),
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": "tool_called",
                            "call_id": call_id or "",
                            "tool_name": tool_name or "",
                            "transport": transport,
                            "raw_item_type": raw_item_type,
                            "server_label": server_label,
                            "args": tool_args,
                            "args_summary": _display_tool_args_summary(state, tool_name, tool_args),
                            "raw_item_preview": _event_payload_preview(_extract_raw_item(item)),
                        },
                    )
                    continue

                if event_name != "tool_output":
                    continue

                call_id = _extract_call_id(item)
                if call_id and call_id in processed_tool_output_call_ids:
                    continue

                tool_name = tool_call_names.get(call_id or "", "") or _extract_tool_name(item)
                output_payload = _extract_tool_output_payload(item)
                raw_item_type = _extract_raw_item_type(item)
                server_label = _extract_mcp_server_label(item)
                transport = _tool_transport(tool_name, item)
                duration_ms = None
                if call_id:
                    started_at = tool_call_started_at.pop(call_id, None)
                    if started_at is not None:
                        duration_ms = int((time.perf_counter() - started_at) * 1000)
                if isinstance(output_payload, dict):
                    logger.info(
                        "Tool call success cid=%s call_id=%s transport=%s tool=%s raw_type=%s server=%s duration_ms=%s summary=%s",
                        conversation_id,
                        call_id or "-",
                        transport,
                        tool_name or "-",
                        raw_item_type or "-",
                        server_label or "-",
                        duration_ms if duration_ms is not None else -1,
                        json.dumps(_tool_output_summary(output_payload), ensure_ascii=False, sort_keys=True),
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": "tool_output",
                            "call_id": call_id or "",
                            "tool_name": tool_name or "",
                            "transport": transport,
                            "raw_item_type": raw_item_type,
                            "server_label": server_label,
                            "duration_ms": duration_ms,
                            "output_summary": _tool_output_summary(output_payload),
                            "output_preview": _event_payload_preview(output_payload),
                        },
                    )
                    record = _persist_retrieval_artifact(
                        state=state,
                        conversation_id=conversation_id,
                        payload=output_payload,
                    )
                    if record:
                        store.save(state)
                        _append_trace_event(
                            conversation_id,
                            {
                                "event": "artifact_registered",
                                "call_id": call_id or "",
                                "tool_name": tool_name or "",
                                "artifact_id": str(record.get("artifact_id") or ""),
                                "artifact_kind": str(record.get("kind") or ""),
                                "artifact_label": str(record.get("label") or ""),
                                "artifact_path": str(record.get("path") or ""),
                            },
                        )
                else:
                    logger.info(
                        "Tool call complete cid=%s call_id=%s transport=%s tool=%s raw_type=%s server=%s duration_ms=%s summary=%s",
                        conversation_id,
                        call_id or "-",
                        transport,
                        tool_name or "-",
                        raw_item_type or "-",
                        server_label or "-",
                        duration_ms if duration_ms is not None else -1,
                        json.dumps(_tool_output_summary(output_payload), ensure_ascii=False, sort_keys=True),
                    )
                    _append_trace_event(
                        conversation_id,
                        {
                            "event": "tool_output",
                            "call_id": call_id or "",
                            "tool_name": tool_name or "",
                            "transport": transport,
                            "raw_item_type": raw_item_type,
                            "server_label": server_label,
                            "duration_ms": duration_ms,
                            "output_summary": _tool_output_summary(output_payload),
                            "output_preview": _event_payload_preview(getattr(item, "output", None) or _extract_raw_item(item)),
                        },
                    )
                if call_id:
                    processed_tool_output_call_ids.add(call_id)

            _ensure_not_cancelled(conversation_id, cancel_event, "after_stream")
            final_answer = str(result.final_output or "").strip()
            usage = getattr(getattr(result, "context_wrapper", None), "usage", None)
    finally:
        _release_cancellation_event(conversation_id)

    if not final_answer:
        raise RuntimeError("The agent returned an empty response.")

    _append_trace_event(
        conversation_id,
        {
            "event": "run_completed",
            "final_answer_preview": _truncate(final_answer, 600),
        },
    )

    usage_input_tokens = _safe_int(getattr(usage, "input_tokens", 0))
    usage_output_tokens = _safe_int(getattr(usage, "output_tokens", 0))
    usage_input_details = getattr(usage, "input_tokens_details", None)
    usage_cached_input_tokens = _safe_int(getattr(usage_input_details, "cached_tokens", 0))
    run_cost = _build_run_cost_payload(
        input_tokens=usage_input_tokens,
        cached_input_tokens=usage_cached_input_tokens,
        output_tokens=usage_output_tokens,
        model=settings.openai_model,
    )

    state = store.load(conversation_id)
    for progress_message in saved_progress_messages:
        state.messages.append({"role": "progress", "content": progress_message})
    state.messages.append({"role": "assistant", "content": final_answer, "run_cost": run_cost})
    has_exportable_artifacts = len(state.artifacts) > run_artifact_start_index
    has_chart = _parse_chart_spec_from_markdown(final_answer) is not None
    if has_exportable_artifacts or has_chart:
        state.latest_export_status = "processing"
        state.latest_export_request = {
            "user_message": user_input,
            "final_answer": final_answer,
            "run_artifact_start_index": run_artifact_start_index,
        }
    else:
        state.latest_export_status = ""
        state.latest_export_request = None
    store.save(state)
    return final_answer


def generate_response(
    conversation_id: str,
    user_input: str,
    store: ConversationStore,
    status_callback: Callable[[str], None],
) -> str:
    return asyncio.run(
        _generate_response_async(
            conversation_id=conversation_id,
            user_input=user_input,
            store=store,
            status_callback=status_callback,
        )
    )
